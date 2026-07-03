# -*- coding: utf-8 -*-
"""
OPMS 1.6 Slave Board - Test App
================================
GUI (tkinter) kiem tra board OPMS 1.6 qua cong Console (RS232/USB, 115200).
Giao dien dang the (card) theo phong cach app Smart PDU: nut BAT/TAT co icon +
mau ro rang, chot trang thai, panel test tu dong + xuat Excel.

Gui lenh CLI xuong firmware, doc gia tri tra ve, so nguong -> PASS/FAIL.
Ke hoach test tu dong nap tu app_config.json.

Cac nhom test: GPO, Quat 48V, GPI, Relay AC, Quat nho (DEV fan), Flash,
NTC nhiet do, do am, nguon 5V/12V, RS485 -> deu chay trong Run test.

Phu thuoc:  pip install pyserial openpyxl
Chay:       python opms_test_app.py
Build exe:  build_opms_exe.bat
"""

import os
import re
import sys
import json
import time
import threading
import subprocess
import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    import serial
    import serial.tools.list_ports
except ImportError:
    serial = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ===== CONFIG =====
APP_TITLE = "OPMS 1.6 Slave Board - Test App"
DEVICE_ID = "OPMS_1.6_SLAVE"
DEFAULT_BAUD = 115200
CONFIG_FILE = "app_config.json"
REPORT_NAME = "test_report.xlsx"   # 1 file co dinh, moi thiet bi ghi them 1 hang

# ===== Theme dung chung (mau + font + nut) tu opms_theme.py =====
from opms_theme import *   # noqa: F401,F403  (WHITE, HDR_BG, PRIMARY, FONT*, flat_btn, sec_btn, resource_path...)

HDR_LOGO_H = 52
OPMS_DEVICE_IMG = "OPMS_Device.png"
FPT_LOGO_IMG = "Vivoo_logo.png"   # logo goc phai header

# handle.control: 1=ON, 2=OFF, 0=CHECK, None=READ
CONTROL_LABEL = {None: "READ", 0: "CHECK", 1: "ON", 2: "OFF"}

EXPECTED_HINT = {
    "Nhiệt độ": "Khong cam:255 | Cam:18-40°C",
    "Nguồn 5V độ ẩm": "Bat:5V | Tat:0V (do tay)",
    "Độ ẩm": "Khong cam:~0 | Cam:30-90 %RH",
    "Đầu vào": "Khong cam:1 | Cam:0",
    "Đầu ra": "Bat:150-400mA | Tat:<100mA",
    "Quạt": "Bat:200-3500rpm | Tat:0 (48V)",
    "Relay AC": "Bat:50-1000mA | Tat:0-50",
    "Quạt nhỏ": "Bat:200-3500rpm | Tat:0 (12V)",
    "Nguồn 12V rs485": "Bat:12V | Tat:0V (do tay)",
    "Cổng RS485": "OK neu doc duoc",
    "Flash": "FLASH=OK",
    "Feedback": "Doc chan feedback",
}

# Thẻ thông báo chuẩn bị trước mỗi nhóm (wizard, theo Production Test v3.0).
PREP_PROMPT = {
    "Nhiệt độ": "Cắm đủ 4 cảm biến nhiệt độ NTC1–NTC4, rồi bấm Tiếp tục.\n"
                "(Bước 'không cắm' sẽ kiểm trước — để trống nếu app yêu cầu.)",
    "Nguồn 5V độ ẩm": "Chuẩn bị Volt kế đo tại cổng H1/H2 (đo tay xác nhận 5V), "
                      "rồi bấm Tiếp tục.",
    "Độ ẩm": "Cắm 2 cảm biến độ ẩm vào H1 và H2, rồi bấm Tiếp tục.\n"
             "(App tự bật nguồn 5V khi đọc và tự tắt sau đó.)",
    "Đầu vào": "Kết nối Test Jig Digital Input.\nKhi chạy bước từng cổng, hãy kích "
               "lần lượt I1→I8 — app tự xác nhận khi đủ. Bấm Tiếp tục.",
    "Đầu ra": "TẮT TẤT CẢ ngõ ra, mắc tải vào O1–O4 theo sơ đồ, rồi bấm Tiếp tục.",
    "Quạt": "Cắm 4 quạt 48V vào jig, rồi bấm Tiếp tục.",
    "Relay AC": "Mắc tải AC + đồng hồ đo dòng, rồi bấm Tiếp tục.",
    "Quạt nhỏ": "Cắm quạt nhỏ thiết bị, rồi bấm Tiếp tục.",
    "Nguồn 12V rs485": "Chuẩn bị đo cổng RS485 (12V), rồi bấm Tiếp tục.",
    "Cổng RS485": "Nối dây loopback (chập A–B) hoặc thiết bị Modbus vào cổng RS485, "
                  "rồi bấm Tiếp tục (bấm Test từng cổng A1–A4).",
    "Nhiệt độ": "Chuẩn bị 4 cảm biến nhiệt độ NTC (CHƯA cắm), rồi bấm Tiếp tục.",
    "Độ ẩm": "Chuẩn bị 2 cảm biến độ ẩm (CHƯA cắm), rồi bấm Tiếp tục.",
    "Relay AC": "Đấu tải AC vào relay, rồi bấm Tiếp tục.",
}


def report_path():
    """Duong dan file report co dinh: canh ben .exe (khi build) hoac canh .py."""
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, REPORT_NAME)


def load_test_config(path=None):
    path = path or resource_path(CONFIG_FILE)
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f).get("test_config", [])
    except Exception:
        return []


def parse_kv(text, key):
    m = re.search(re.escape(key) + r"\s*=\s*(-?\d+)", text)
    return int(m.group(1)) if m else None


# Nguong chap nhan (PASS) mac dinh cho trang thai "Bat/Cam" - theo HDSD.
# Nguoi dung chinh duoc tren app o the "Nguong chap nhan (PASS)".
DEFAULT_TH = {
    "Đầu ra":   (150, 400),   # GPO_CUR (mA) khi bat
    "Quạt":     (200, 3500),  # FANn_RPS (rpm) khi bat [Quat 48V, dinh muc 3500]
    "Quạt nhỏ": (200, 3500),  # DFANn_RPS (rpm) khi bat [Quat 12V]
    "Relay AC": (50, 1000),   # AC_CUR (mA) khi dong relay
    "Nhiệt độ": (18, 40),     # NTC (°C) khi cam cam bien
    "Độ ẩm":    (30, 90),     # %RH (SHT30-ARP) khi cam cam bien
}
TH_UNIT = {"Đầu ra": "mA", "Quạt": "rpm", "Quạt nhỏ": "rpm", "Relay AC": "mA",
           "Nhiệt độ": "°C", "Độ ẩm": "%RH"}
# Ten hien thi (Run test / Setup): nhom noi bo -> ten than thien
DISPLAY_NAME = {"Quạt": "Quạt 48V", "Quạt nhỏ": "Quạt 12V"}

# RS485: chuoi mong doi AK MCU KIT tra ve (kit dap "AK_MCU_RS485_OK").
# Cong DAT khi chuoi nhan ve CHUA token nay (so noi dung, khong chi dem byte).
RS485_EXPECT = "OK"


def rs485_rx_str(text):
    """Trich noi dung trong RS485RX="..." tu phan hoi firmware. "" neu khong co."""
    m = re.search(r'RS485RX="([^"]*)"', text or "")
    return m.group(1) if m else ""


def rs485_pass(text):
    """True neu phan hoi RS485 chua token mong doi (vd AK_MCU_RS485_OK)."""
    return RS485_EXPECT in rs485_rx_str(text).upper()

# Nhóm test -> loại popup bán tự động (Run test). Mỗi nhóm gộp thành 1 dòng/bảng.
PANEL_GROUPS = {"Đầu vào": "gpi", "Quạt nhỏ": "dfan", "Quạt": "fan",
                "Đầu ra": "gpo", "Cổng RS485": "rs485",
                "Nhiệt độ": "ntc", "Độ ẩm": "hum", "Relay AC": "ac"}
PANEL_LABEL = {"gpi": "Kích & kiểm tra 8 cổng vào (I1–I8)",
               "fan": "Cắm & kiểm tra 4 quạt 48V (Fan1–4)",
               "dfan": "Cắm & kiểm tra 2 quạt 12V (FC1–2)",
               "gpo": "Đấu tải & kiểm tra 4 cổng ra (O1–O4)",
               "rs485": "Kiểm tra 4 cổng RS485 (A1–A4)",
               "ntc": "Kiểm tra 4 cảm biến nhiệt độ (NTC1–4)",
               "hum": "Kiểm tra 2 cảm biến độ ẩm (H1–2)",
               "ac": "Kiểm tra Relay AC (bật/tắt)"}


def resolve_plan(m, TH=None, fan_pct=100):
    """metadata buoc -> ke hoach test tu dong.
    TH: dict nguong chap nhan (PASS) cho trang thai Bat/Cam, chinh tren app.
    fan_pct: % toc do khi bat Quat 48V (cai trong Setup).
    """
    TH = TH or DEFAULT_TH
    def band(key):
        return TH.get(key, DEFAULT_TH[key])
    g = m["group"]
    ctrl = m["control"]
    port = m["port"] if m["port"] is not None else 0
    allp = m["all"]
    on = (ctrl == 1)
    if g == "Đầu ra":
        ch = port + 1
        return {"cmds": [f"gpo {ch} {'on' if on else 'off'}"], "read": "gpoi",
                "key": "GPO_CUR", "mode": "numeric",
                "expect": band("Đầu ra") if on else (0, 100), "settle": 0.4}
    if g == "Quạt":
        ch = port + 1
        on_cmd = f"fan {ch} on {fan_pct}" if on else f"fan {ch} off"
        return {"cmds": [on_cmd], "read": f"fanr {ch}",
                "key": f"FAN{ch}_RPS", "mode": "numeric",
                "expect": band("Quạt") if on else (0, 5), "settle": 1.0}
    if g == "Đầu vào":
        if allp:
            return {"cmds": [], "read": "gpi", "key": "GPI", "mode": "gpi_all",
                    "expect": 255, "settle": 0.1,
                    "prompt": "KHÔNG cắm cổng đầu vào nào (I1–I8)"}
        return {"cmds": [], "read": "gpi", "key": "GPI", "mode": "gpi_bit",
                "bit": port, "expect": 0, "settle": 0.1}
    if g == "Relay AC":
        return {"cmds": [f"ac {'on' if on else 'off'}"], "read": "aci",
                "key": "AC_CUR", "mode": "numeric",
                "expect": band("Relay AC") if on else (0, 50), "settle": 0.4}
    if g == "Quạt nhỏ":
        return {"cmds": [f"dfan {i} {'on' if on else 'off'}" for i in (1, 2)],
                "reads": [("dfanr 1", "DFAN1_RPS"), ("dfanr 2", "DFAN2_RPS")],
                "mode": "multi",
                "expect": band("Quạt nhỏ") if on else (0, 5), "settle": 1.0}
    if g == "Flash":
        return {"cmds": [], "mode": "system", "settle": 0.2}
    if g == "Feedback":
        return {"cmds": [], "read": "dfanfb 1", "key": "DFAN1_FB",
                "mode": "read_ok", "settle": 0.1}
    if g in ("Giao tiếp Orange Pi", "Orange Pi"):   # scaffold: loopback UART3
        return {"cmds": [], "mode": "pi", "settle": 0.2}

    # ===== Cam bien / nguon / RS485 =====
    if g == "Nhiệt độ":     # 4 NTC, doc 0-255 (ho ~255, cam theo nguong app)
        exp = band("Nhiệt độ") if on else (245, 255)
        return {"cmds": [], "reads": [(f"ntc {i}", f"NTC{i}") for i in (1, 2, 3, 4)],
                "mode": "multi", "expect": exp, "settle": 0.2,
                "prompt": ("CẮM TẤT CẢ 4 cảm biến nhiệt độ (NTC)" if on
                           else "KHÔNG cắm cổng nhiệt độ nào")}
    if g == "Độ ẩm":        # 2 cong, can bat nguon 5V cam bien truoc -> doc -> tat
        # Chua cam: firmware xa tu loc -> duong tha noi ~0 (KHONG treo len 3V3)
        exp = band("Độ ẩm") if on else (0, 20)
        return {"cmds": ["humpwr 1 on", "humpwr 2 on"],
                "reads": [(f"hum {i}", f"HUM{i}") for i in (1, 2)],
                "post_cmds": ["humpwr 1 off", "humpwr 2 off"],
                "mode": "multi", "expect": exp, "settle": 0.4,
                "prompt": ("CẮM cả 2 cảm biến độ ẩm" if on
                           else "KHÔNG cắm cổng độ ẩm nào")}
    if g == "Nguồn 5V độ ẩm":   # bo qua trong Auto Test (theo yeu cau)
        return {"mode": "skip"}
    if g == "Nguồn 12V rs485":  # bat/tat nguon + doc power-good
        return {"cmds": [f"rspwr {'on' if on else 'off'}"], "read": "rspg",
                "key": "RS485_PG", "mode": "flag",
                "expect": 1 if on else 0, "settle": 0.3}
    if g == "Cổng RS485":       # probe tung cong (can jig + Modbus / loopback)
        p = port + 1
        return {"cmds": [], "read": f"rs485 {p}", "key": f"RS485{p}_RX",
                "mode": "rs485", "settle": 0.1,
                "prompt": (f"Cổng RS485 A{p}B{p}:\n"
                           f"• Nối thiết bị Modbus được cấp (hoặc dây loopback chập "
                           f"A{p}–B{p}) vào cổng này.\n"
                           f"• Bấm OK → app gửi chuỗi và đếm byte phản hồi (PASS nếu >0).\n\n"
                           f"Test thủ công: mở Terminal gõ   rs485 {p} <chuỗi>")}

    return {"mode": "manual"}


# =============================================================
#  Nut gat BAT/TAT co icon + mau
# =============================================================
class Toggle:
    def __init__(self, parent, on_change, width=10):
        self.state = False
        self.on_change = on_change
        self.btn = tk.Button(parent, width=width, font=FONT_B, fg="white",
                             relief="raised", bd=2, cursor="hand2",
                             activeforeground="white", command=self._click)
        self._render()

    def _click(self):
        self.state = not self.state
        self._render()
        if self.on_change:
            self.on_change(self.state)

    def set(self, s):
        self.state = bool(s)
        self._render()

    def _render(self):
        if self.state:   # BAT = primary (chu trang)
            self.btn.config(text="●  BẬT", bg=ON_BG, activebackground=ON_HOV,
                            fg="white", activeforeground="white")
        else:            # TAT = secondary (chu toi)
            self.btn.config(text="○  TẮT", bg=OFF_BG, activebackground=OFF_HOV,
                            fg=SEC_TX, activeforeground=SEC_TX)

    def pack(self, **kw):
        self.btn.pack(**kw)
        return self

    def grid(self, **kw):
        self.btn.grid(**kw)
        return self


def make_dot(parent, size=14, color=DOT_OFF):
    c = tk.Canvas(parent, width=size, height=size, bg=WHITE, highlightthickness=0)
    oid = c.create_oval(2, 2, size - 2, size - 2, fill=color, outline="")
    return c, oid


# =============================================================
#  App  (flat_btn / sec_btn / mau / font lay tu opms_theme)
# =============================================================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1280x860")
        self.configure(bg=MAIN_BG)
        self.ser = None
        self.test_groups = load_test_config()
        self.step_items = []
        self._meta = {}
        self._paused = False
        self._cancel = False
        self._busy = False              # dang trong 1 lenh blocking (chan idle reader)
        self._cmd_history = []          # lich su lenh terminal
        self._cmd_hist_idx = -1
        self._fw_ver = ""               # phien ban FW doc luc ket noi
        self._connecting = False        # dang mo cong / xac thuc ID
        self._connect_cancelled = False
        self._connect_timeout_id = None
        self.thresholds = dict(DEFAULT_TH)   # nguong PASS hien hanh (chinh tren app)
        self.fan_speed = 80                  # % toc do khi bat Quat 48V (Setup)
        self._buzzer_on = True               # còi báo bước test (tắt khi test im lặng)
        self._build_ui()
        self._populate_auto_plan()
        self._set_controls_enabled(False)
        self._refresh_progress()
        self.update_idletasks()
        self.minsize(1040, 600)
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.after(150, self._poll_idle)   # terminal: doc du lieu async khi ranh roi

    # ---------------- UI helpers ----------------
    def _load_hdr_image(self, name, h=HDR_LOGO_H):
        """Nap PNG, scale ve chieu cao h, dan len nen header (tranh vien trang)."""
        try:
            from PIL import Image as _Img, ImageTk as _ImgTk
            im = _Img.open(resource_path(name)).convert("RGBA")
            w0, h0 = im.size
            nw = max(1, int(w0 * h / h0))
            im = im.resize((nw, h), _Img.LANCZOS)
            bgc = tuple(int(HDR_BG.lstrip("#")[i:i+2], 16) for i in (0, 2, 4))
            bg = _Img.new("RGB", (nw, h), bgc)
            bg.paste(im, mask=im.split()[3])
            return _ImgTk.PhotoImage(bg)
        except Exception:
            return None

    def _draw_icon(self, parent, kind, size=26):
        """Ve icon vector don sac (navy = nhan, xam = phu) - theo he 3 mau."""
        c = tk.Canvas(parent, width=size, height=size, bg=WHITE,
                      highlightthickness=0)
        s = size
        P = NAVY     # nhan chinh
        G = GREY     # phu
        if kind == "bulb":          # ngo ra (bong den)
            c.create_oval(s*0.22, s*0.08, s*0.78, s*0.64, fill="", outline=P, width=2)
            c.create_rectangle(s*0.4, s*0.6, s*0.6, s*0.8, fill=G, outline="")
            c.create_line(s*0.4, s*0.72, s*0.6, s*0.72, fill=WHITE)
        elif kind == "fan":         # quat 48V
            r = s*0.42
            cx = cy = s/2
            for a in (45, 135, 225, 315):
                c.create_arc(cx-r, cy-r, cx+r, cy+r, start=a, extent=64,
                             style="pieslice", fill=P, outline="")
            c.create_oval(cx-3, cy-3, cx+3, cy+3, fill=WHITE, outline="")
        elif kind == "ac":          # relay AC (may lanh)
            c.create_rectangle(s*0.12, s*0.28, s*0.88, s*0.6, fill=P, outline="")
            c.create_line(s*0.2, s*0.5, s*0.8, s*0.5, fill=WHITE)
            c.create_arc(s*0.28, s*0.64, s*0.46, s*0.84, start=200, extent=140,
                         style="arc", outline=G, width=2)
            c.create_arc(s*0.54, s*0.64, s*0.72, s*0.84, start=200, extent=140,
                         style="arc", outline=G, width=2)
        elif kind == "wind":        # quat nho (gio)
            for y in (s*0.34, s*0.52, s*0.7):
                c.create_arc(s*0.12, y-s*0.13, s*0.62, y+s*0.13, start=270,
                             extent=200, style="arc", outline=P, width=2)
        elif kind == "input":       # dau vao GPI
            c.create_rectangle(s*0.18, s*0.72, s*0.82, s*0.82, fill=P, outline="")
            c.create_line(s*0.5, s*0.16, s*0.5, s*0.58, fill=P, width=2)
            c.create_polygon(s*0.36, s*0.5, s*0.64, s*0.5, s*0.5, s*0.68,
                             fill=P, outline="")
        elif kind == "chip":        # flash & rtc (chip nho)
            c.create_rectangle(s*0.26, s*0.26, s*0.74, s*0.74, fill=P, outline="")
            for i in (0.36, 0.5, 0.64):
                c.create_line(s*i, s*0.16, s*i, s*0.26, fill=G)
                c.create_line(s*i, s*0.74, s*i, s*0.84, fill=G)
                c.create_line(s*0.16, s*i, s*0.26, s*i, fill=G)
                c.create_line(s*0.74, s*i, s*0.84, s*i, fill=G)
            c.create_oval(s*0.45, s*0.45, s*0.55, s*0.55, fill=WHITE, outline="")
        elif kind == "temp":        # nhiet do (nhiet ke)
            c.create_rectangle(s*0.44, s*0.12, s*0.56, s*0.64, fill="", outline=P)
            c.create_rectangle(s*0.47, s*0.22, s*0.53, s*0.64, fill=P, outline="")
            c.create_oval(s*0.36, s*0.58, s*0.64, s*0.86, fill=P, outline="")
        elif kind == "humid":       # do am (giot nuoc)
            c.create_oval(s*0.3, s*0.42, s*0.7, s*0.84, fill=P, outline="")
            c.create_polygon(s*0.5, s*0.1, s*0.33, s*0.52, s*0.67, s*0.52,
                             fill=P, outline="")
        elif kind == "power":       # nguon (nut power)
            c.create_arc(s*0.2, s*0.2, s*0.8, s*0.8, start=70, extent=320,
                         style="arc", outline=P, width=3)
            c.create_line(s*0.5, s*0.12, s*0.5, s*0.5, fill=P, width=3)
        elif kind == "rs485":       # cong RS485 (2 chieu)
            c.create_line(s*0.18, s*0.38, s*0.82, s*0.38, fill=P,
                          width=2, arrow="last")
            c.create_line(s*0.82, s*0.62, s*0.18, s*0.62, fill=G,
                          width=2, arrow="last")
        return c

    def _card(self, parent, title, subtitle="", icon=None):
        frm = tk.Frame(parent, bg=WHITE, relief="raised", bd=1,
                       highlightbackground=CARD_BD, highlightthickness=1)
        hf = tk.Frame(frm, bg=WHITE, padx=12, pady=7)
        hf.pack(fill="x")
        if icon:
            self._draw_icon(hf, icon).pack(side="left", padx=(0, 8))
        # Tieu de + phu de xep DOC (tranh tran/mat chu khi card hep)
        txt = tk.Frame(hf, bg=WHITE)
        txt.pack(side="left", fill="x", expand=True)
        tk.Label(txt, text=title, bg=WHITE, fg=BLUE_ACC, font=FONT_CARD,
                 anchor="w", justify="left").pack(anchor="w")
        if subtitle:
            tk.Label(txt, text=subtitle, bg=WHITE, fg=HDR_SUB, font=FONT_SM,
                     anchor="w", justify="left").pack(anchor="w")
        tk.Frame(frm, bg=CARD_BD, height=1).pack(fill="x")
        body = tk.Frame(frm, bg=WHITE, padx=12, pady=8)
        body.pack(fill="both", expand=True)
        return frm, body

    # ---------------- BUILD UI ----------------
    def _build_ui(self):
        # ===== HEADER =====
        hdr = tk.Frame(self, bg=HDR_BG)
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=HDR_ACC, width=6).pack(side="left", fill="y")

        # Anh thiet bi OPMS (ben trai)
        self._dev_img = self._load_hdr_image(OPMS_DEVICE_IMG)
        if self._dev_img:
            tk.Label(hdr, image=self._dev_img, bg=HDR_BG).pack(
                side="left", padx=(12, 10), pady=8)

        tit = tk.Frame(hdr, bg=HDR_BG, padx=4, pady=10)
        tit.pack(side="left")
        tk.Label(tit, text="OPMS 1.6  SLAVE BOARD", bg=HDR_BG, fg="white",
                 font=("Segoe UI", 15, "bold")).pack(anchor="w")
        tk.Label(tit, text="Functional Test  •  STM32F303VCT6  •  Console 115200",
                 bg=HDR_BG, fg=HDR_SUB, font=FONT_SM).pack(anchor="w")

        # Logo FPT (ben phai)
        self._fpt_img = self._load_hdr_image(FPT_LOGO_IMG)
        if self._fpt_img:
            tk.Label(hdr, image=self._fpt_img, bg=HDR_BG).pack(
                side="right", padx=(8, 16), pady=8)

        # ===== CONTROL / CONNECTION BAR =====
        bar = tk.Frame(self, bg=CTRL_BG)
        bar.pack(fill="x")
        bar_in = tk.Frame(bar, bg=CTRL_BG, padx=12, pady=8)
        bar_in.pack(fill="x")

        # ===== 1 hàng setup duy nhất: kết nối (trái) + Serial/Lưu/Còi (phải) =====
        sec_btn(bar_in, "⚙ Setup", padx=8,
                command=self._open_thresholds).pack(side="left", padx=(0, 6))
        sec_btn(bar_in, "🔌 Console CH348", padx=8,
                command=self._open_console_test).pack(side="left", padx=(0, 12))

        self._conn_dot, self._conn_oid = make_dot(bar_in, 14, GREY)
        self._conn_dot.config(bg=CTRL_BG)
        self._conn_dot.pack(side="left")
        self.lbl_conn = tk.Label(bar_in, text=" Chưa kết nối", bg=CTRL_BG,
                                 fg=HDR_SUB, font=FONT_SM, anchor="w", width=24)
        self.lbl_conn.pack(side="left", padx=(2, 10))

        tk.Label(bar_in, text="COM:", bg=CTRL_BG, fg=HDR_SUB,
                 font=FONT_SM).pack(side="left")
        self.cmb_port = ttk.Combobox(bar_in, width=30, values=self._list_ports())
        self.cmb_port.pack(side="left", padx=4)
        sec_btn(bar_in, "↻", width=2,
                command=self._reload_ports).pack(side="left")
        tk.Label(bar_in, text="  Baud:", bg=CTRL_BG, fg=HDR_SUB,
                 font=FONT_SM).pack(side="left")
        self.cmb_baud = ttk.Combobox(bar_in, width=7,
                                     values=[9600, 19200, 38400, 57600, 115200])
        self.cmb_baud.set(DEFAULT_BAUD)
        self.cmb_baud.pack(side="left", padx=4)
        self.btn_conn = flat_btn(bar_in, "Kết nối", READ_BG, READ_HOV, padx=10,
                                 width=14, command=self.toggle_connect)
        self.btn_conn.pack(side="left", padx=8)

        # --- Bên phải cùng hàng (gộp từ hàng Serial cũ) ---
        self.btn_buzz = sec_btn(bar_in, "🔊 Còi: BẬT", padx=10, pady=4,
                                command=self._toggle_buzzer)
        self.btn_buzz.pack(side="right", padx=(6, 0))
        sec_btn(bar_in, "💾 Lưu Excel", padx=10, pady=4,
                command=lambda: self._save_report(silent=False)).pack(
                    side="right", padx=6)
        # Ô Serial board đã chuyển xuống panel Run test.

        # ===== BODY: trai = the dieu khien, phai = panel auto =====
        body = tk.Frame(self, bg=MAIN_BG, padx=10, pady=10)
        body.pack(fill="both", expand=True)

        # Trai: dashboard - CỐ ĐỊNH bề rộng (không nở ra khi mở rộng cửa sổ)
        left_outer = tk.Frame(body, bg=MAIN_BG, width=720)
        left_outer.pack(side="left", fill="y")
        left_outer.pack_propagate(False)

        # Phải: Run test - NỞ RA chiếm phần còn lại khi phóng to cửa sổ
        right = tk.Frame(body, bg=MAIN_BG)
        right.pack(side="right", fill="both", expand=True, padx=(10, 0))
        lcanvas = tk.Canvas(left_outer, bg=MAIN_BG, highlightthickness=0)
        lvsb = ttk.Scrollbar(left_outer, orient="vertical", command=lcanvas.yview)
        lcanvas.configure(yscrollcommand=lvsb.set)
        lvsb.pack(side="right", fill="y")
        lcanvas.pack(side="left", fill="both", expand=True)
        left = tk.Frame(lcanvas, bg=MAIN_BG)
        lwin = lcanvas.create_window((0, 0), window=left, anchor="nw")
        left.bind("<Configure>",
                  lambda e: lcanvas.configure(scrollregion=lcanvas.bbox("all")))
        lcanvas.bind("<Configure>",
                     lambda e: lcanvas.itemconfig(lwin, width=e.width))

        def _wheel(e):
            lcanvas.yview_scroll(int(-e.delta / 120), "units")
        lcanvas.bind("<Enter>", lambda e: lcanvas.bind_all("<MouseWheel>", _wheel))
        lcanvas.bind("<Leave>", lambda e: lcanvas.unbind_all("<MouseWheel>"))

        self._build_control_cards(left)
        self._build_auto_panel(right)

        # ===== STATUS BAR =====
        self.status = tk.StringVar(value="San sang. Mo cong COM de bat dau.")
        tk.Label(self, textvariable=self.status, bg=GREY_LT, fg=INK,
                 anchor="w", padx=10, pady=4, font=FONT_SM).pack(fill="x")

        self._control_widgets = []   # de enable/disable theo ket noi

    # ----- The dieu khien -----
    def _build_control_cards(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.columnconfigure(1, weight=1)

        # --- GPO (ngang: O1..O4, dong tong ben duoi) ---
        card, b = self._card(parent, "Đầu ra GPO (O1–O4)", "Bật/tắt + đo dòng tổng",
                             icon="bulb")
        card.grid(row=1, column=1, sticky="nsew", padx=4, pady=4)
        self.gpo_toggles = []
        gprow = tk.Frame(b, bg=WHITE)
        gprow.pack(fill="x")
        for ch in range(1, 5):
            cell = tk.Frame(gprow, bg=WHITE)
            cell.grid(row=0, column=ch - 1, padx=6)
            tk.Label(cell, text=f"O{ch}", bg=WHITE, fg=GREY_TX, font=FONT_SM).pack()
            tg = Toggle(cell, lambda s, c=ch: self._on_gpo(c, s), width=6)
            tg.pack()
            self.gpo_toggles.append(tg)
        cur = tk.Frame(b, bg=WHITE)
        cur.pack(fill="x", pady=(8, 0))
        flat_btn(cur, "⚡ Đo dòng tổng", READ_BG, READ_HOV, padx=8,
                 command=self._read_gpo_cur).pack(side="left")
        self.lbl_gpo_cur = tk.Label(cur, text="—", bg=WHITE, fg=BLUE_ACC, font=FONT_B,
                                    width=8, anchor="w")   # be rong co dinh
        self.lbl_gpo_cur.pack(side="left", padx=10)

        # --- Quat 48V (ngang: Fan1..Fan4, rpm ben duoi) ---
        card, b = self._card(parent, "Quạt 48V (Fan1–4)", "Bật/tắt + đọc tốc độ (rpm)",
                             icon="fan")
        card.grid(row=1, column=0, sticky="nsew", padx=4, pady=4)
        self.fan_toggles = []
        self.fan_rps_lbl = []
        frow = tk.Frame(b, bg=WHITE)
        frow.pack(fill="x")
        for ch in range(1, 5):
            cell = tk.Frame(frow, bg=WHITE)
            cell.grid(row=0, column=ch - 1, padx=5)
            tk.Label(cell, text=f"Fan{ch}", bg=WHITE, fg=GREY_TX, font=FONT_SM).pack()
            tg = Toggle(cell, lambda s, c=ch: self._on_fan(c, s), width=6)
            tg.pack(pady=1)
            flat_btn(cell, "rpm", READ_BG, READ_HOV, padx=6,
                     command=lambda c=ch: self._read_fan(c)).pack(pady=1)
            lb = tk.Label(cell, text="—", bg=WHITE, fg=BLUE_ACC, font=FONT_B,
                          width=6, anchor="center")   # be rong CO DINH -> the khong dich, 4 o van vua
            lb.pack()
            self.fan_toggles.append(tg)
            self.fan_rps_lbl.append(lb)

        # --- Relay AC ---
        card, b = self._card(parent, "Relay AC", "Bật/tắt + đo dòng", icon="ac")
        card.grid(row=4, column=0, sticky="nsew", padx=4, pady=4)
        row = tk.Frame(b, bg=WHITE)
        row.pack(fill="x", pady=3)
        tk.Label(row, text="AC", bg=WHITE, fg=GREY_TX, font=FONT_B,
                 width=4, anchor="w").pack(side="left")
        self.ac_toggle = Toggle(row, self._on_ac)
        self.ac_toggle.pack(side="left")
        flat_btn(row, "⚡ Đo", READ_BG, READ_HOV, padx=8,
                 command=self._read_ac).pack(side="left", padx=8)
        self.lbl_ac_cur = tk.Label(row, text="—", bg=WHITE, fg=BLUE_ACC, font=FONT_B,
                                   width=8, anchor="w")   # be rong co dinh
        self.lbl_ac_cur.pack(side="left")

        # --- Quat 12V (quat nho 1-2, ngang, doc rpm) -> giong the Quat 48V ---
        card, b = self._card(parent, "Quạt 12V (FC1–2)", "Bật/tắt + đọc tốc độ (rpm)",
                             icon="fan")
        card.grid(row=0, column=1, sticky="nsew", padx=4, pady=4)
        self.dfan_toggles = []
        self.dfan_rps_lbl = []
        dfrow = tk.Frame(b, bg=WHITE)
        dfrow.pack(fill="x")
        for ch in range(1, 3):
            cell = tk.Frame(dfrow, bg=WHITE)
            cell.grid(row=0, column=ch - 1, padx=14)
            tk.Label(cell, text=f"FC{ch}", bg=WHITE, fg=GREY_TX, font=FONT_SM).pack()
            tg = Toggle(cell, lambda s, c=ch: self._on_dfan(c, s), width=6)
            tg.pack(pady=1)
            flat_btn(cell, "rpm", READ_BG, READ_HOV, padx=6,
                     command=lambda c=ch: self._read_dfan_rps(c)).pack(pady=1)
            lb = tk.Label(cell, text="—", bg=WHITE, fg=BLUE_ACC, font=FONT_B,
                          width=6, anchor="center")   # be rong CO DINH -> the khong dich, 4 o van vua
            lb.pack()
            self.dfan_toggles.append(tg)
            self.dfan_rps_lbl.append(lb)

        # --- GPI ---
        card, b = self._card(parent, "Đầu vào GPI (I1–I8)", "74HC165 – đọc 8 ngõ vào",
                             icon="input")
        card.grid(row=2, column=0, sticky="nsew", padx=4, pady=4)
        grid = tk.Frame(b, bg=WHITE)
        grid.pack(fill="x")
        self.gpi_dots = []
        for i in range(8):
            cell = tk.Frame(grid, bg=WHITE)
            cell.grid(row=0, column=i, padx=4)
            dc, oid = make_dot(cell, 16, DOT_OFF)
            dc.pack()
            tk.Label(cell, text=f"I{i+1}", bg=WHITE, fg=GREY_TX,
                     font=FONT_SM).pack()
            self.gpi_dots.append((dc, oid))
        rr = tk.Frame(b, bg=WHITE)
        rr.pack(fill="x", pady=(8, 0))
        flat_btn(rr, "🔍 Đọc GPI", READ_BG, READ_HOV, padx=8,
                 command=self._read_gpi).pack(side="left")
        self.lbl_gpi = tk.Label(rr, text="GPI = —", bg=WHITE, fg=BLUE_ACC,
                                font=FONT_B)
        self.lbl_gpi.pack(side="left", padx=10)

        # --- Flash ---
        card, b = self._card(parent, "Flash (W25Q80)", "SPI3 – đọc JEDEC ID",
                             icon="chip")
        card.grid(row=0, column=0, sticky="nsew", padx=4, pady=4)
        rr = tk.Frame(b, bg=WHITE)
        rr.pack(fill="x", pady=6)
        flat_btn(rr, "🔍 Kiểm tra Flash", READ_BG, READ_HOV, padx=10,
                 command=self._read_flash).pack(side="left")
        self.lbl_flash = tk.Label(rr, text="—", bg=WHITE, fg=BLUE_ACC, font=FONT_B)
        self.lbl_flash.pack(side="left", padx=10)
        # ===== Thẻ cảm biến / nguồn / RS485 =====
        # Nhiệt độ
        card, b = self._card(parent, "Nhiệt độ (NTC1–4)", "Nhiệt độ (°C)", icon="temp")
        card.grid(row=2, column=1, sticky="nsew", padx=4, pady=4)
        self.ntc_lbl = []
        rowf = tk.Frame(b, bg=WHITE)
        rowf.pack(fill="x")
        for ch in range(1, 5):
            cell = tk.Frame(rowf, bg=WHITE)
            cell.grid(row=0, column=ch - 1, padx=8)
            tk.Label(cell, text=f"NTC{ch}", bg=WHITE, fg=GREY_TX,
                     font=FONT_SM).pack()
            lb = tk.Label(cell, text="—", bg=WHITE, fg=BLUE_ACC, font=FONT_B,
                          width=6, anchor="center")   # be rong CO DINH -> the khong dich, 4 o van vua
            lb.pack()
            self.ntc_lbl.append(lb)
        flat_btn(b, "Đọc nhiệt độ", READ_BG, READ_HOV, padx=8,
                 command=self._read_ntc).pack(anchor="w", pady=(8, 0))

        # Độ ẩm (ngang H1,H2 + nguon 5V cam bien gop chung)
        card, b = self._card(parent, "Độ ẩm (H1–2)", "%RH (SHT30-ARP) + nguồn cảm biến",
                             icon="humid")
        card.grid(row=3, column=0, sticky="nsew", padx=4, pady=4)
        self.hum_lbl = []
        self.humpwr_toggles = []
        hrow = tk.Frame(b, bg=WHITE)
        hrow.pack(fill="x")
        for ch in range(1, 3):
            cell = tk.Frame(hrow, bg=WHITE)
            cell.grid(row=0, column=ch - 1, padx=12)
            tk.Label(cell, text=f"H{ch}", bg=WHITE, fg=GREY_TX, font=FONT_B).pack()
            tk.Label(cell, text="5V", bg=WHITE, fg=HDR_SUB, font=FONT_SM).pack()
            tg = Toggle(cell, lambda s, c=ch: self._on_humpwr(c, s), width=6)
            tg.pack(pady=1)
            lb = tk.Label(cell, text="—", bg=WHITE, fg=BLUE_ACC, font=FONT_B,
                          width=6, anchor="center")   # be rong CO DINH -> the khong dich, 4 o van vua
            lb.pack()
            self.hum_lbl.append(lb)
            self.humpwr_toggles.append(tg)
        flat_btn(b, "Đọc độ ẩm", READ_BG, READ_HOV, padx=8,
                 command=self._read_hum).pack(anchor="w", pady=(8, 0))

        # (Thẻ "Nguồn 5V/12V" đã bỏ: 5V gộp vào thẻ Độ ẩm, 12V gộp vào thẻ RS485)

        # Cổng RS485 (gop nguon 12V cap cho cong RS485 1,2 + PG)
        card, b = self._card(parent, "Cổng RS485 (RS1–RS4)", "nguồn 12V + probe",
                             icon="rs485")
        card.grid(row=3, column=1, sticky="nsew", padx=4, pady=4)
        prow = tk.Frame(b, bg=WHITE)
        prow.pack(fill="x", pady=(0, 4))
        tk.Label(prow, text="Nguồn 12V", bg=WHITE, fg=GREY_TX,
                 font=FONT_B).pack(side="left")
        self.rspwr_toggle = Toggle(prow, self._on_rspwr, width=6)
        self.rspwr_toggle.pack(side="left", padx=6)
        tk.Label(prow, text="PG:", bg=WHITE, fg=GREY_TX,
                 font=FONT_SM).pack(side="left")
        dc, oid = make_dot(prow, 14, DOT_OFF)
        dc.pack(side="left", padx=(2, 0))
        self.pg_dot = (dc, oid)
        self.rs485_lbl = []
        grid = tk.Frame(b, bg=WHITE)
        grid.pack(fill="x")
        for p in range(1, 5):
            cell = tk.Frame(grid, bg=WHITE)
            cell.grid(row=0, column=p - 1, padx=6)
            flat_btn(cell, f"RS{p}", READ_BG, READ_HOV, padx=8,
                     command=lambda pp=p: self._probe_rs485(pp)).pack()
            lb = tk.Label(cell, text="—", bg=WHITE, fg=BLUE_ACC, font=FONT_SM)
            lb.pack()
            self.rs485_lbl.append(lb)
        tk.Label(b, text="12V cấp cho cổng RS485 1,2 · (cần jig Modbus)", bg=WHITE,
                 fg=HDR_SUB, font=FONT_SM).pack(anchor="w", pady=(6, 0))

    def _current_thresholds(self):
        """Nguong PASS hien hanh (da luu tu hop thoai Nguong)."""
        return dict(self.thresholds)

    def _open_console_test(self):
        """Phuong an A trong 1 app duy nhat: mo cua so test mach USB 8-COM CH348
        bang cach chay lai CHINH chuong trinh nay voi co '--console' -> tien trinh
        rieng, cua so rieng (COM/serial/ket qua doc lap)."""
        try:
            if getattr(sys, "frozen", False):
                subprocess.Popen([sys.executable, "--console"])   # cung 1 file .exe
            else:
                subprocess.Popen([sys.executable,
                                  os.path.abspath(__file__), "--console"])
            self.status.set("Đã mở cửa sổ Test Console CH348.")
        except Exception as e:
            messagebox.showerror("Lỗi mở cửa sổ", str(e))

    def _open_thresholds(self):
        """Hop thoai nhap/chinh nguong PASS (min–max) cho cac muc test."""
        dlg = tk.Toplevel(self)
        dlg.title("Setup — Ngưỡng chấp nhận (PASS)")
        dlg.configure(bg=WHITE)
        dlg.transient(self)
        dlg.resizable(False, False)
        tk.Label(dlg, text="Setup — Ngưỡng chấp nhận (PASS) khi Bật/Cắm", bg=WHITE,
                 fg=BLUE_ACC, font=FONT_CARD).grid(row=0, column=0, columnspan=4,
                 sticky="w", padx=16, pady=(14, 2))
        tk.Label(dlg, text="Để trống/sai sẽ dùng mặc định. Min>Max sẽ tự đảo.",
                 bg=WHITE, fg=HDR_SUB, font=FONT_SM).grid(row=1, column=0,
                 columnspan=4, sticky="w", padx=16, pady=(0, 10))
        # header
        for j, t0 in enumerate(("Mục test", "Min", "Max", "Đơn vị")):
            tk.Label(dlg, text=t0, bg=WHITE, fg=GREY_TX, font=FONT_B).grid(
                row=2, column=j, padx=10, pady=2, sticky="w")
        entries = {}
        r = 3
        for g, (lo, hi) in DEFAULT_TH.items():
            cur_lo, cur_hi = self.thresholds.get(g, (lo, hi))
            tk.Label(dlg, text=DISPLAY_NAME.get(g, g), bg=WHITE, fg=INK,
                     font=FONT).grid(row=r, column=0, padx=10, pady=3, sticky="w")
            e_lo = tk.Entry(dlg, width=9, font=FONT_MONO, justify="center")
            e_lo.insert(0, str(cur_lo)); e_lo.grid(row=r, column=1, padx=10)
            e_hi = tk.Entry(dlg, width=9, font=FONT_MONO, justify="center")
            e_hi.insert(0, str(cur_hi)); e_hi.grid(row=r, column=2, padx=10)
            tk.Label(dlg, text=TH_UNIT.get(g, ""), bg=WHITE, fg=GREY_TX,
                     font=FONT_SM).grid(row=r, column=3, padx=10, sticky="w")
            entries[g] = (e_lo, e_hi)
            r += 1

        # Ghi chú: ngưỡng Quạt là tốc độ ĐỌC VỀ
        tk.Label(dlg, text="* Ngưỡng Quạt 48V / Quạt 12V là giá trị tốc độ "
                 "ĐỌC VỀ (rpm) khi quạt đang chạy.", bg=WHITE, fg=HDR_SUB,
                 font=FONT_SM).grid(row=r, column=0, columnspan=4, sticky="w",
                                    padx=12, pady=(2, 8))
        r += 1
        # Tốc độ (%) khi BẬT Quạt 48V
        tk.Label(dlg, text="Tốc độ bật Quạt 48V", bg=WHITE, fg=INK,
                 font=FONT_B).grid(row=r, column=0, padx=10, pady=3, sticky="w")
        e_fan = tk.Entry(dlg, width=9, font=FONT_MONO, justify="center")
        e_fan.insert(0, str(getattr(self, "fan_speed", 80)))
        e_fan.grid(row=r, column=1, padx=10)
        tk.Label(dlg, text="% (0–100)", bg=WHITE, fg=GREY_TX,
                 font=FONT_SM).grid(row=r, column=3, padx=10, sticky="w")
        r += 1

        def _save():
            for g, (e_lo, e_hi) in entries.items():
                try:
                    a = float(e_lo.get().strip()); b2 = float(e_hi.get().strip())
                    self.thresholds[g] = (a, b2) if a <= b2 else (b2, a)
                except Exception:
                    self.thresholds[g] = DEFAULT_TH[g]
            try:
                self.fan_speed = max(0, min(100, int(float(e_fan.get().strip()))))
            except Exception:
                self.fan_speed = 80
            self.status.set(f"Đã lưu ngưỡng PASS. Tốc độ bật Quạt 48V = "
                            f"{self.fan_speed}%.")
            dlg.destroy()

        def _reset():
            for g, (e_lo, e_hi) in entries.items():
                lo0, hi0 = DEFAULT_TH[g]
                e_lo.delete(0, "end"); e_lo.insert(0, str(lo0))
                e_hi.delete(0, "end"); e_hi.insert(0, str(hi0))

        bar = tk.Frame(dlg, bg=WHITE)
        bar.grid(row=r, column=0, columnspan=4, sticky="ew", padx=16, pady=14)
        flat_btn(bar, "Lưu", ON_BG, ON_HOV, padx=16,
                 command=_save).pack(side="right", padx=4)
        sec_btn(bar, "↺ Mặc định", padx=10,
                command=_reset).pack(side="right", padx=4)
        dlg.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width() - dlg.winfo_reqwidth()) // 2
        y = self.winfo_rooty() + 80
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")
        dlg.grab_set()

    # ----- Panel test tu dong -----
    def _build_auto_panel(self, parent):
        card, b = self._card(parent, "Run test",
                             "PASS/FAIL theo ngưỡng")
        card.pack(fill="both", expand=True)

        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        # Bang 10 buoc test = khu KTV thao tac chinh -> chu LON, dong cao, noi bat
        style.configure("OP.Treeview", rowheight=36, font=("Segoe UI", 12))
        style.configure("OP.Treeview.Heading", font=("Segoe UI", 10, "bold"))
        style.configure("OP.Horizontal.TProgressbar", troughcolor=GREY_LT,
                        background=PRIMARY, bordercolor=CARD_BD,
                        lightcolor=PRIMARY, darkcolor=PRIMARY)

        # --- Serial board (chuyển từ toolbar xuống panel Run test) ---
        srow = tk.Frame(b, bg=WHITE)
        srow.pack(fill="x", pady=(0, 8))
        tk.Label(srow, text="Serial board:", bg=WHITE, fg=INK,
                 font=FONT_B).pack(side="left")
        self.ent_serial = ttk.Entry(srow, font=FONT)
        self.ent_serial.pack(side="left", fill="x", expand=True, padx=8)

        # --- Progress bar + dem PASS/FAIL (theo spec) ---
        prog = tk.Frame(b, bg=WHITE)
        prog.pack(fill="x", pady=(0, 8))
        self.lbl_fail = tk.Label(prog, text="FAIL 0", bg=WHITE, fg=FAIL_FG, font=FONT_B,
                                 width=8, anchor="e")
        self.lbl_fail.pack(side="right")
        self.lbl_pass = tk.Label(prog, text="PASS 0", bg=WHITE, fg=PASS_FG, font=FONT_B,
                                 width=8, anchor="e")
        self.lbl_pass.pack(side="right")
        self.lbl_prog_n = tk.Label(prog, text="0 / 0", bg=WHITE, fg=INK, font=FONT_B,
                                   width=9, anchor="e")
        self.lbl_prog_n.pack(side="right", padx=(8, 6))
        self.pbar = ttk.Progressbar(prog, style="OP.Horizontal.TProgressbar",
                                    mode="determinate", maximum=100)
        self.pbar.pack(side="left", fill="x", expand=True)

        # --- Điều khiển Run test: nav từng bước + Reset (Ở TRÊN bảng) ---
        nav = tk.Frame(b, bg=WHITE)
        nav.pack(fill="x", pady=(0, 8))
        flat_btn(nav, "◀ Trước", SEC_BG, SEC_HOV, fg=SEC_TX, padx=12, pady=4,
                 command=self._step_prev).pack(side="left")
        self.lbl_stepnav = tk.Label(nav, text="Bước 0/0", bg=WHITE, fg=INK,
                                    font=FONT_B, width=11, anchor="center")
        self.lbl_stepnav.pack(side="left", padx=4)
        flat_btn(nav, "Sau ▶", SEC_BG, SEC_HOV, fg=SEC_TX, padx=12, pady=4,
                 command=self._step_nextsel).pack(side="left")
        flat_btn(nav, "▶ Run test", PRIMARY, PRIMARY_HOV,
                 font=("Segoe UI", 13, "bold"), padx=16, pady=5,
                 command=self._step_run_current).pack(side="left", padx=(10, 0))
        sec_btn(nav, "↺ Reset", padx=12, pady=4,
                command=self.reset_results).pack(side="right")

        # --- Bang ket qua ---
        tree_box = tk.Frame(b, bg=WHITE)
        tree_box.pack(fill="both", expand=True)
        cols = ("res", "val")
        self.tree = ttk.Treeview(tree_box, columns=cols, show="tree headings",
                                 style="OP.Treeview", selectmode="browse", height=12)
        self.tree.heading("#0", text="Bước test")
        self.tree.heading("res", text="KQ")
        self.tree.heading("val", text="Giá trị")
        self.tree.column("#0", width=250)
        self.tree.column("res", width=60, anchor="center")
        self.tree.column("val", width=90, anchor="center")
        self.tree.tag_configure("pass", foreground=PASS_FG)
        self.tree.tag_configure("fail", foreground=FAIL_FG)
        self.tree.tag_configure("running", foreground=RUN_FG)   # dang chay
        self.tree.tag_configure("pending", foreground=DIS_FG)   # cho chay
        self.tree.tag_configure("group", font=("Segoe UI", 9, "bold"))
        vsb = ttk.Scrollbar(tree_box, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", self._on_tree_dbl)
        self.tree.bind("<Button-3>", self._tree_context_menu)   # chuột phải: đánh dấu thủ công
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)

        # --- Ghi chú ---
        tk.Label(b, text="💡 “Run test” = chạy bước đang chọn rồi sang bước kế. "
                 "Hoặc nhấn đúp một dòng để test bước đó.",
                 bg=WHITE, fg=BLUE_ACC, font=FONT_SM,
                 anchor="w", justify="left").pack(fill="x", pady=(4, 2))

        # --- Terminal (duoi) kieu Smart PDU ---
        self._build_terminal(b)

    # ----- Terminal (theo app Smart PDU) -----
    def _build_terminal(self, parent):
        tm = tk.Frame(parent, bg=WHITE)
        tm.pack(fill="both", expand=False, pady=(8, 0))

        hdr = tk.Frame(tm, bg=WHITE)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Terminal", bg=WHITE, fg=BLUE_ACC,
                 font=("Segoe UI", 10, "bold")).pack(side="left")
        tk.Button(hdr, text="Xóa Log", bg=WHITE, fg=GREY,
                  activebackground=GREY_LT, relief="solid", bd=1,
                  font=("Segoe UI", 8), padx=6, pady=1,
                  command=self._clear_log).pack(side="right")

        # Input bar (pack bottom truoc khi pack txt de khong bi day ra ngoai)
        inp = tk.Frame(tm, bg=TERM_BG, padx=8, pady=6)
        inp.pack(side="bottom", fill="x")
        tk.Label(inp, text=">", bg=TERM_BG, fg=TERM_TX,
                 font=("Consolas", 11, "bold")).pack(side="left", padx=(0, 6))
        self.ent_cmd = tk.Entry(inp, font=("Consolas", 10), bg=TERM_BG,
                                fg=TERM_FG, insertbackground="white",
                                relief="flat", bd=0, state="disabled")
        self.ent_cmd.pack(side="left", fill="x", expand=True, ipady=4)
        self.ent_cmd.bind("<Return>", self._send_manual)
        self.ent_cmd.bind("<Up>", self._cmd_hist_up)
        self.ent_cmd.bind("<Down>", self._cmd_hist_down)
        self.btn_send = flat_btn(inp, "Gửi", READ_BG, READ_HOV, padx=10,
                                 state="disabled", command=self._send_manual)
        self.btn_send.pack(side="left", padx=(8, 0))

        # Quick commands (khop lenh CLI firmware OPMS 1.6)
        qf = tk.Frame(tm, bg=TERM_BG, padx=6, pady=3)
        qf.pack(side="bottom", fill="x")
        tk.Label(qf, text="Quick:", bg=TERM_BG, fg=GREY_SB,
                 font=("Consolas", 8)).pack(side="left", padx=(0, 4))
        self._qcmd_btns = []
        # Tap trung cac phan chinh: (nhan hien thi, lenh CLI gui xuong)
        _qcmds = [("help", "help"), ("gpi", "gpi"), ("gpo", "gpoi"),
                  ("temp", "ntc 1"), ("hum", "hum 1"), ("rs485", "rs485 1")]
        for _lbl, _cmd in _qcmds:
            _b = tk.Button(qf, text=_lbl, bg=NAVY, fg=WHITE,
                           activebackground=NAVY_DK, activeforeground="white",
                           font=("Consolas", 9), relief="flat", bd=0,
                           padx=14, pady=3, state="disabled",
                           command=lambda c=_cmd: self._quick_send(c))
            _b.pack(side="left", padx=3, pady=1)
            self._qcmd_btns.append(_b)

        # Text log
        self.txt = tk.Text(tm, bg=TERM_BG, fg=TERM_FG, insertbackground="white",
                           font=("Consolas", 9), state="disabled", wrap="none",
                           padx=8, pady=6, width=40, height=9)
        self.txt.tag_configure("tx", foreground=TERM_TX)     # echo lenh (INFO)
        self.txt.tag_configure("rx", foreground=TERM_FG)     # phan hoi
        self.txt.tag_configure("err", foreground=FAIL_FG)    # loi
        self.txt.tag_configure("ok", foreground=PASS_FG)     # ok
        self.txt.tag_configure("warn", foreground=WARN_FG)   # canh bao
        sb_txt = ttk.Scrollbar(tm, command=self.txt.yview)
        self.txt.configure(yscrollcommand=sb_txt.set)
        sb_txt.pack(side="right", fill="y")
        self.txt.pack(fill="both", expand=True)

    # ---------------- enable/disable controls ----------------
    def _set_controls_enabled(self, en):
        st = "normal" if en else "disabled"
        widgets = []
        widgets += [t.btn for t in getattr(self, "gpo_toggles", [])]
        widgets += [t.btn for t in getattr(self, "fan_toggles", [])]
        widgets += [t.btn for t in getattr(self, "dfan_toggles", [])]
        widgets += [t.btn for t in getattr(self, "humpwr_toggles", [])]
        if hasattr(self, "ac_toggle"):
            widgets.append(self.ac_toggle.btn)
        if hasattr(self, "rspwr_toggle"):
            widgets.append(self.rspwr_toggle.btn)
        for w in widgets:
            try:
                w.config(state=st)
            except Exception:
                pass

    # ---------------- Auto plan tree ----------------
    def _populate_auto_plan(self):
        self.step_items = []
        self._meta = {}
        if not self.test_groups:
            self.tree.insert("", "end", text="(Khong nap duoc app_config.json)")
            return
        # Run test: DANH SACH PHANG, danh so 1..N (bo nhom B1/B2/B3 cho do roi).
        ORDER = ["Flash", "Quạt nhỏ", "Quạt", "Đầu ra", "Đầu vào", "Nhiệt độ",
                 "Độ ẩm", "Cổng RS485", "Relay AC", "Giao tiếp Orange Pi"]
        # Nhom bo khoi Run test (gop vao test khac):
        #  - Feedback = doc PG (da check trong Cong RS485)
        #  - Nguon 5V do am -> gop voi Do am ; Nguon 12V rs485 -> gop voi Cong RS485
        HIDDEN = {"Feedback", "Nguồn 5V độ ẩm", "Nguồn 12V rs485"}
        gmap = {g.get("name"): g for g in self.test_groups}
        ordered = [n for n in ORDER if n in gmap]
        for g in self.test_groups:               # them nhom con lai (neu co), giu thu tu
            n = g.get("name")
            if n not in ordered and n not in HIDDEN:
                ordered.append(n)

        num = 0
        for gname in ordered:
            g = gmap[gname]
            num += 1
            disp = DISPLAY_NAME.get(gname, gname)
            pkind = PANEL_GROUPS.get(gname)
            iid = self.tree.insert("", "end", text=f"{num}. {disp}",
                                   values=("-", "-"), tags=("pending",))
            if pkind:                            # nhom co popup -> 1 dong
                self._meta[iid] = {
                    "group": gname, "des": PANEL_LABEL.get(pkind, disp),
                    "control": 1, "port": 0, "all": False,
                    "gpi_combined": (pkind == "gpi"), "panel_kind": pkind,
                    "skip": False, "value": "", "result": "",
                }
            else:                                # nhom chay inline (Flash / Orange Pi)
                steps = g.get("list_test", [])
                h = steps[0].get("handle", {}) if steps else {}
                self._meta[iid] = {
                    "group": gname, "des": disp,
                    "control": h.get("control"), "port": h.get("port"),
                    "all": h.get("all", False), "gpi_combined": False,
                    "panel_kind": None,
                    "skip": False, "value": "", "result": "",
                }
            self.step_items.append(iid)
        self._step_idx = 0
        self._step_select(0)            # con tro dieu huong ve buoc 1

    def _on_tree_dbl(self, ev):
        iid = self.tree.identify_row(ev.y)
        if iid in self._meta:
            self.run_step(iid)

    def _mark_selected(self, ok):
        """Đánh dấu kết quả THỦ CÔNG cho dòng đang chọn (ok: True/False/None)."""
        sel = self.tree.selection()
        if not sel or sel[0] not in self._meta:
            self.status.set("⚠ Hãy chọn 1 dòng trong bảng test trước.")
            return
        self._set_result(sel[0], "thủ công", ok)
        m = self._meta[sel[0]]
        kq = "PASS" if ok else ("FAIL" if ok is False else "BỎ QUA")
        self.status.set(f"Thủ công: {m['des']} → {kq}")

    def _run_selected(self):
        sel = self.tree.selection()
        if sel and sel[0] in self._meta:
            self.run_step(sel[0])
        else:
            self.status.set("⚠ Hãy chọn 1 dòng trong bảng test trước.")

    # ---------------- Điều hướng từng bước (Next/Back) ----------------
    def _step_select(self, idx):
        """Chọn bước thứ idx (0-based) trong cây + cập nhật nhãn 'Bước N/M'."""
        if not self.step_items:
            if hasattr(self, "lbl_stepnav"):
                self.lbl_stepnav.config(text="Bước 0/0")
            return
        self._step_idx = max(0, min(idx, len(self.step_items) - 1))
        iid = self.step_items[self._step_idx]
        self.tree.selection_set(iid)
        self.tree.see(iid)
        if hasattr(self, "lbl_stepnav"):
            self.lbl_stepnav.config(
                text=f"Bước {self._step_idx + 1}/{len(self.step_items)}")

    def _on_tree_select(self, ev=None):
        sel = self.tree.selection()
        if sel and sel[0] in self.step_items:
            self._step_idx = self.step_items.index(sel[0])
            if hasattr(self, "lbl_stepnav"):
                self.lbl_stepnav.config(
                    text=f"Bước {self._step_idx + 1}/{len(self.step_items)}")

    def _step_prev(self):
        self._step_select(getattr(self, "_step_idx", 0) - 1)

    def _step_nextsel(self):
        self._step_select(getattr(self, "_step_idx", 0) + 1)

    def _step_run_current(self):
        """Chạy bước đang chọn (mở popup nếu có) rồi tự tiến tới bước kế."""
        if not self._need_conn():
            return
        if not self.step_items:
            return
        idx = max(0, min(getattr(self, "_step_idx", 0), len(self.step_items) - 1))
        self._step_select(idx)
        self.run_step(self.step_items[idx])
        if idx < len(self.step_items) - 1:        # tự sang bước kế tiếp
            self._step_select(idx + 1)

    def _tree_context_menu(self, ev):
        """Chuột phải vào 1 bước -> đánh dấu kết quả THỦ CÔNG (Đạt/Không đạt/Bỏ qua)."""
        iid = self.tree.identify_row(ev.y)
        if iid not in self._meta:
            return
        self.tree.selection_set(iid)
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="✓  Đạt (PASS)",
                         command=lambda: self._set_result(iid, "thủ công", True))
        menu.add_command(label="✗  Không đạt (FAIL)",
                         command=lambda: self._set_result(iid, "thủ công", False))
        menu.add_command(label="–  Bỏ qua",
                         command=lambda: self._set_result(iid, "bỏ qua", None))
        menu.add_separator()
        menu.add_command(label="▶  Chạy bước này (tự động)",
                         command=lambda: self.run_step(iid))
        try:
            menu.tk_popup(ev.x_root, ev.y_root)
        finally:
            menu.grab_release()

    def _beep(self):
        """Phát 1 tiếng còi trên thiết bị (nếu còi đang BẬT) để báo có bước test."""
        if getattr(self, "_buzzer_on", False) and self.ser and self.ser.is_open:
            try:
                self._send("beep")
            except Exception:
                pass

    def _ui_beep(self):
        """Còi khi thao tác dashboard (giống auto test), chống dội: tối đa 1 lần / 0.4s
        nên 1 lần bật/tắt (kèm đọc tự động) chỉ kêu 1 tiếng."""
        now = time.time()
        if now - getattr(self, "_last_ui_beep", 0.0) < 0.4:
            return
        self._last_ui_beep = now
        self._beep()

    def _center_dialog(self, dlg):
        """Đặt cửa sổ nổi cố định, canh giữa theo cửa sổ chính (không nhảy góc)."""
        dlg.update_idletasks()
        w = dlg.winfo_reqwidth()
        h = dlg.winfo_reqheight()
        px = self.winfo_rootx() + (self.winfo_width() - w) // 2
        py = self.winfo_rooty() + (self.winfo_height() - h) // 3
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        px = max(10, min(px, sw - w - 20))
        py = max(10, min(py, sh - h - 20))
        dlg.geometry(f"+{px}+{py}")

    def _panel_header(self, dlg, step_iid, title_text):
        """Tiêu đề lớn + đậm + số thứ tự bước cho popup Run test."""
        try:
            num = self.step_items.index(step_iid) + 1
        except (ValueError, AttributeError):
            num = 0
        total = len(getattr(self, "step_items", []) or [])
        txt = (f"Bước {num}/{total}  ·  {title_text}" if num else title_text)
        bar = tk.Frame(dlg, bg=HDR_BG)
        bar.pack(fill="x")
        tk.Label(bar, text=txt, bg=HDR_BG, fg="white",
                 font=("Segoe UI", 16, "bold"), anchor="w",
                 padx=16, pady=10).pack(fill="x")

    def _toggle_buzzer(self):
        """Bật/tắt còi báo (test im lặng) — tham khảo app Smart PDU."""
        self._buzzer_on = not self._buzzer_on
        self.btn_buzz.config(text="🔊 Còi: BẬT" if self._buzzer_on
                             else "🔇 Còi: TẮT")
        self.status.set("Còi BẬT — báo mỗi bước test." if self._buzzer_on
                        else "Còi TẮT — test im lặng.")

    def _show_manual_guide(self):
        messagebox.showinfo(
            "Test thủ công",
            "CÁCH TEST THỦ CÔNG (đơn giản):\n\n"
            "1) Thao tác trên các thẻ điều khiển bên trái "
            "(Bật/Tắt, ⚡ Đo dòng, 🔍 Đọc...) và quan sát kết quả.\n\n"
            "2) Ở bảng “Test tự động” bên phải: CHỌN một dòng (bấm 1 lần), "
            "rồi bấm nút ở ngay dưới bảng:\n"
            "      ✓ Đạt   ·   ✗ Không đạt   ·   – Bỏ qua\n"
            "   (▶ Chạy dòng = tự chạy riêng bước đang chọn.)\n\n"
            "3) Nhập Serial rồi bấm “💾 Lưu Excel” để ghi kết quả.")

    # ---- Đồng bộ DASHBOARD khi auto-test (tham khảo Smart PDU) ----
    def _sync_dashboard(self, m):
        """Khi auto-test gửi lệnh điều khiển -> cập nhật nút Bật/Tắt trên thẻ."""
        g = m["group"]
        port = m["port"] if m["port"] is not None else 0
        on = (m["control"] == 1)
        try:
            if g == "Đầu ra" and port < len(self.gpo_toggles):
                self.gpo_toggles[port].set(on)
            elif g == "Quạt" and port < len(self.fan_toggles):
                self.fan_toggles[port].set(on)
            elif g == "Relay AC" and hasattr(self, "ac_toggle"):
                self.ac_toggle.set(on)
            elif g == "Quạt nhỏ":
                for t in self.dfan_toggles:
                    t.set(on)
            elif g == "Nguồn 5V độ ẩm" and port < len(self.humpwr_toggles):
                self.humpwr_toggles[port].set(on)
            elif g == "Nguồn 12V rs485" and hasattr(self, "rspwr_toggle"):
                self.rspwr_toggle.set(on)
            self.update_idletasks()
        except Exception:
            pass

    def _sync_value_label(self, m, val):
        """Cập nhật ô giá trị (mA/rps) trên thẻ theo kết quả vừa đọc."""
        g = m["group"]
        port = m["port"] if m["port"] is not None else 0
        try:
            if g == "Đầu ra":
                self.lbl_gpo_cur.config(text=f"{val} mA")
            elif g == "Quạt" and port < len(self.fan_rps_lbl):
                self.fan_rps_lbl[port].config(text=f"{val}")
            elif g == "Relay AC":
                self.lbl_ac_cur.config(text=f"{val} mA")
            self.update_idletasks()
        except Exception:
            pass

    def _sync_multi_labels(self, m, vals):
        """Cập nhật ô NTC/Độ ẩm trên thẻ khi đọc nhiều kênh."""
        g = m["group"]
        try:
            labels = None
            suff = ""
            if g == "Nhiệt độ":
                labels = self.ntc_lbl
            elif g == "Độ ẩm":
                labels = self.hum_lbl
            elif g == "Quạt nhỏ":
                labels = self.dfan_rps_lbl
                suff = ""
            if labels:
                for i, v in enumerate(vals):
                    if i < len(labels):
                        labels[i].config(text=f"{v}{suff}")
            self.update_idletasks()
        except Exception:
            pass

    # ---------------- Serial ----------------
    def _list_ports(self):
        if serial is None:
            return []
        out = []
        for p in serial.tools.list_ports.comports():
            desc = (p.description or "").strip()
            if desc and desc.lower() != "n/a":
                out.append(f"{p.device}  —  {desc}")   # COMx — ten thiet bi
            else:
                out.append(p.device)
        return out

    def _reload_ports(self):
        self.cmb_port["values"] = self._list_ports()

    def _set_conn_dot(self, color, text):
        self._conn_dot.itemconfig(self._conn_oid, fill=color)
        s = " " + text
        if len(s) > 23:                  # cat bot de nhan khong tran -> nut co dinh
            s = s[:22] + "…"
        self.lbl_conn.config(text=s)

    def toggle_connect(self):
        if serial is None:
            messagebox.showerror("Thiếu thư viện", "Chưa cài pyserial.")
            return
        # Dang ket noi -> nut dong vai tro "Huy"
        if self._connecting:
            self._cancel_connect()
            return
        # Dang ket noi thanh cong -> ngat
        if self.ser and self.ser.is_open:
            self._disconnect()
            return
        # Bat dau ket noi moi
        # Lay COMx tu chuoi hien thi "COMx  —  ten thiet bi"
        port = self.cmb_port.get().strip().split("  —  ")[0].strip()
        if not port:
            messagebox.showwarning("Chưa chọn COM", "Hãy chọn cổng COM.")
            return
        baud = int(self.cmb_baud.get())
        self._connecting = True
        self._connect_cancelled = False
        self.cmb_port.config(state="disabled")
        self.cmb_baud.config(state="disabled")
        self.btn_conn.config(text="✖ Hủy kết nối", bg=GREY)
        self._set_conn_dot(GREY, f"Đang mở {port}...")
        self.status.set(f"Đang mở {port} @ {baud}...")
        self._log(f"[Đang mở {port} @ {baud} — bấm Hủy nếu treo]\n", "tx")

        def _do_open():     # mo cong chay nen (tranh treo UI voi cong ao)
            try:
                ser = serial.Serial(port, baud, timeout=0.2)
            except Exception as e:
                self.after(0, lambda: self._on_open_result(None, str(e), port, baud))
                return
            self.after(0, lambda: self._on_open_result(ser, None, port, baud))

        threading.Thread(target=_do_open, daemon=True).start()
        self._connect_timeout_id = self.after(4000, self._on_connect_slow)

    def _on_connect_slow(self):
        self._connect_timeout_id = None
        if self._connecting and not self._connect_cancelled:
            self.status.set("Mở cổng quá lâu — có thể là cổng COM ảo bị treo. "
                            "Bấm 'Hủy kết nối' để chọn cổng khác.")
            self._log("[Cảnh báo: mở cổng quá lâu — bấm Hủy để chọn cổng khác]\n", "err")

    def _reset_conn_ui(self):
        self._connecting = False
        self.cmb_port.config(state="normal")
        self.cmb_baud.config(state="normal")
        self.btn_conn.config(text="Kết nối", bg=READ_BG)

    def _cancel_connect(self):
        """Nguoi dung bam Huy trong luc mo cong / xac thuc ID."""
        self._connect_cancelled = True
        if self._connect_timeout_id is not None:
            self.after_cancel(self._connect_timeout_id)
            self._connect_timeout_id = None
        try:
            if self.ser:
                self.ser.close()
        except Exception:
            pass
        self.ser = None
        self._reset_conn_ui()
        self._set_conn_dot(GREY, "Đã hủy kết nối")
        self.status.set("Đã hủy kết nối — chọn cổng COM khác.")
        self._log("[Đã hủy kết nối]\n", "err")

    def _disconnect(self):
        try:
            self.ser.close()
        except Exception:
            pass
        self.ser = None
        self.btn_conn.config(text="Kết nối", bg=READ_BG)
        self._set_conn_dot(GREY, "Đã ngắt kết nối")
        self._set_controls_enabled(False)
        self._term_enabled(False)
        self._log("[Đã ngắt kết nối]\n", "err")

    def _on_open_result(self, ser, err, port, baud):
        if self._connect_timeout_id is not None:
            self.after_cancel(self._connect_timeout_id)
            self._connect_timeout_id = None
        if self._connect_cancelled:           # da Huy truoc khi mo xong
            if ser is not None:
                try:
                    ser.close()
                except Exception:
                    pass
            return
        if err is not None:                   # mo cong loi
            self._reset_conn_ui()
            self.ser = None
            self._set_conn_dot(FAIL_FG, "Không mở được cổng")
            self.status.set(f"Không mở được {port}.")
            self._log(f"[Không mở được {port}: {err}]\n", "err")
            messagebox.showerror("Lỗi mở cổng", f"Không mở được {port}:\n{err}")
            return
        # Mo cong OK -> xac thuc ID
        self.ser = ser
        self._set_conn_dot(GREY, f"{port} • đang kiểm tra ID...")
        self._log(f"[Mở {port} OK — kiểm tra ID...]\n", "tx")
        self.update()
        time.sleep(0.2)
        resp = self._send_recv("id", 0.6)
        self._on_verify_result(DEVICE_ID in resp, port, baud)

    def _on_verify_result(self, ok, port, baud):
        if self._connect_cancelled:
            return
        if not (self.ser and self.ser.is_open):
            self._reset_conn_ui()
            return
        if not ok:
            # SAI thiet bi / nham cong / sai baud -> dong cong + canh bao
            try:
                self.ser.close()
            except Exception:
                pass
            self.ser = None
            self._reset_conn_ui()
            self._set_conn_dot(FAIL_FG, f"{port} • SAI thiết bị")
            self.status.set(f"{port} không nhận đúng OPMS 1.6 — đã ngắt.")
            self._log(f"[{port} @ {baud} không đúng ID — đã ngắt]\n", "err")
            messagebox.showwarning(
                "Sai thiết bị",
                f"Cổng {port} @ {baud} KHÔNG nhận đúng thiết bị "
                f"{DEVICE_ID}!\n\nNguyên nhân có thể:\n"
                "  • Chọn nhầm cổng COM\n"
                "  • Sai baud (Console phải 115200)\n"
                "  • Chưa nạp đúng firmware test\n\n"
                "Kiểm tra lại cổng COM/baud rồi kết nối lại.")
            return
        # DUNG thiet bi -> hoan tat
        self._connecting = False
        self.cmb_port.config(state="normal")
        self.cmb_baud.config(state="normal")
        self.btn_conn.config(text="Ngắt kết nối", bg=GREY)
        self._set_conn_dot(DOT_ON, f"{port} • {DEVICE_ID}")
        self.status.set(f"Đã kết nối {port}.")
        self._log(f"[Xác thực OK — {DEVICE_ID}]\n", "ok")
        ver = self._send_recv("ver", 0.5)
        mver = re.search(r"FW\s+([0-9.]+)", ver)
        self._fw_ver = mver.group(1) if mver else ""
        self._term_enabled(True)
        self._set_controls_enabled(True)

    def _need_conn(self):
        if not (self.ser and self.ser.is_open):
            self.status.set("Chưa kết nối — hãy mở cổng COM trước.")
            return False
        return True

    def _send(self, cmd):
        if self.ser and self.ser.is_open:
            self.ser.write((cmd + "\r\n").encode())
            self._log(f"> {cmd}\n", "tx")

    def _send_recv(self, cmd, timeout=0.7):
        # 'timeout' = thoi gian CHO TOI DA cho phan hoi. Tra ve SOM khi da nhan
        # mot dong KEY=VALUE hoan chinh (co '=' va xuong dong) roi im >= GAP,
        # => UI khong bi treo tron timeout moi lan doc.
        # LUU Y: firmware echo '\r\n' NGAY khi nhan lenh, roi moi tinh toan (vd
        # 'aci' chay EmonLib ~150ms) moi in AC_CUR=... Vi vay KHONG duoc thoat
        # som chi vi thay '\r\n' echo -> phai cho den khi co dong '=' that su.
        if not (self.ser and self.ser.is_open):
            return ""
        GAP = 0.06            # khoang lang xac nhan het phan hoi (giay)
        self._busy = True
        try:
            self.ser.reset_input_buffer()
            self.ser.write((cmd + "\r\n").encode())
            self._log(f"> {cmd}\n", "tx")
            t0 = time.time()
            last_rx = None
            buf = b""
            while time.time() - t0 < timeout:
                n = self.ser.in_waiting
                if n:
                    buf += self.ser.read(n)
                    last_rx = time.time()
                elif (last_rx is not None and (time.time() - last_rx) >= GAP
                      and b"=" in buf and (b"\n" in buf or b"\r" in buf)):
                    break            # da co dong KEY=VALUE hoan chinh -> xong
                else:
                    time.sleep(0.01)
            resp = buf.decode(errors="ignore")
            if resp.strip():
                self._log(resp if resp.endswith("\n") else resp + "\n", "rx")
            return resp
        finally:
            self._busy = False

    # ---------------- Terminal ----------------
    def _log(self, text, tag=""):
        if not hasattr(self, "txt"):
            return
        self.txt.config(state="normal")
        self.txt.insert("end", text, tag)
        self.txt.see("end")
        self.txt.config(state="disabled")

    def _clear_log(self):
        self.txt.config(state="normal")
        self.txt.delete("1.0", "end")
        self.txt.config(state="disabled")

    def _poll_idle(self):
        """Doc du lieu serial khong yeu cau (boot banner, echo...) khi khong ban."""
        try:
            if (self.ser and self.ser.is_open and not self._busy
                    and self.ser.in_waiting):
                data = self.ser.read(self.ser.in_waiting).decode(errors="ignore")
                if data:
                    self._log(data, "rx")
        except Exception:
            pass
        self.after(150, self._poll_idle)

    def _term_enabled(self, en):
        st = "normal" if en else "disabled"
        for w in (getattr(self, "ent_cmd", None), getattr(self, "btn_send", None)):
            if w:
                try:
                    w.config(state=st)
                except Exception:
                    pass
        for b in getattr(self, "_qcmd_btns", []):
            b.config(state=st)

    def _push_history(self, cmd):
        if not self._cmd_history or self._cmd_history[-1] != cmd:
            self._cmd_history.append(cmd)
            if len(self._cmd_history) > 50:
                self._cmd_history.pop(0)
        self._cmd_hist_idx = -1

    def _send_manual(self, _evt=None):
        if not self._need_conn():
            messagebox.showwarning(APP_TITLE, "Chưa kết nối COM.")
            return
        cmd = self.ent_cmd.get().strip()
        if not cmd:
            return
        self._send(cmd)
        self._push_history(cmd)
        self.ent_cmd.delete(0, "end")

    def _quick_send(self, cmd):
        if not self._need_conn():
            return
        self._send(cmd)
        self._push_history(cmd)

    def _cmd_hist_up(self, _evt=None):
        if not self._cmd_history:
            return "break"
        if self._cmd_hist_idx < len(self._cmd_history) - 1:
            self._cmd_hist_idx += 1
        self.ent_cmd.delete(0, "end")
        self.ent_cmd.insert(0, self._cmd_history[-(self._cmd_hist_idx + 1)])
        return "break"

    def _cmd_hist_down(self, _evt=None):
        if self._cmd_hist_idx > 0:
            self._cmd_hist_idx -= 1
            self.ent_cmd.delete(0, "end")
            self.ent_cmd.insert(0, self._cmd_history[-(self._cmd_hist_idx + 1)])
        else:
            self._cmd_hist_idx = -1
            self.ent_cmd.delete(0, "end")
        return "break"

    # ---------------- Control callbacks ----------------
    def _on_gpo(self, ch, state):
        if not self._need_conn():
            return
        self._ui_beep()
        self._send(f"gpo {ch} {'on' if state else 'off'}")
        self.status.set(f"GPO{ch} = {'BẬT' if state else 'TẮT'}")
        self.after(300, self._read_gpo_cur)

    def _read_gpo_cur(self):
        if not self._need_conn():
            return
        self._ui_beep()
        v = None
        for _ in range(2):                       # thu lai 1 lan neu truot
            v = parse_kv(self._send_recv("gpoi", 1.2), "GPO_CUR")
            if v is not None:
                break
        self.lbl_gpo_cur.config(text=f"{v} mA" if v is not None else "—")

    def _on_fan(self, ch, state):
        if not self._need_conn():
            return
        self._ui_beep()
        # token: gạt lại (on/off) sẽ huỷ chuỗi đọc lặp cũ
        self._fan_tok = getattr(self, "_fan_tok", {})
        self._fan_tok[ch] = self._fan_tok.get(ch, 0) + 1
        if state:
            self._send(f"fan {ch} on {getattr(self, 'fan_speed', 80)}")
            # quạt tăng tốc dần -> đọc lặp ~8 lần (mỗi 700ms) để cập nhật tốc độ ổn định
            self.after(700, lambda t=self._fan_tok[ch]: self._poll_fan_rps(ch, 8, t))
        else:
            self._send(f"fan {ch} off")
            self.fan_rps_lbl[ch - 1].config(text="—")
        self.status.set(f"Fan{ch} = {'BẬT ' + str(getattr(self, 'fan_speed', 80)) + '%' if state else 'TẮT'}")

    def _poll_fan_rps(self, ch, n, tok):
        """Đọc lặp rpm trong lúc quạt tăng tốc; dừng khi hết lượt, mất kết nối,
        hoặc người dùng gạt lại quạt (token đổi)."""
        if (n <= 0 or not (self.ser and self.ser.is_open)
                or tok != getattr(self, "_fan_tok", {}).get(ch)):
            return
        self._read_fan_quiet(ch)
        self.after(700, lambda: self._poll_fan_rps(ch, n - 1, tok))

    def _read_fan_quiet(self, ch):
        v = parse_kv(self._send_recv(f"fanr {ch}", 1.2), f"FAN{ch}_RPS")
        self.fan_rps_lbl[ch - 1].config(text=f"{v}" if v is not None else "—")

    def _read_fan(self, ch):
        if not self._need_conn():
            return
        self._ui_beep()
        self._read_fan_quiet(ch)

    def _on_ac(self, state):
        if not self._need_conn():
            return
        self._ui_beep()
        self._send(f"ac {'on' if state else 'off'}")
        self.status.set(f"Relay AC = {'BẬT' if state else 'TẮT'}")
        self.after(300, self._read_ac)

    def _read_ac(self):
        if not self._need_conn():
            return
        self._ui_beep()
        v = None
        for _ in range(2):                       # thu lai 1 lan neu truot
            v = parse_kv(self._send_recv("aci", 1.2), "AC_CUR")
            if v is not None:
                break
        self.lbl_ac_cur.config(text=f"{v} mA" if v is not None else "—")

    def _on_dfan(self, ch, state):
        if not self._need_conn():
            return
        self._ui_beep()
        self._send(f"dfan {ch} {'on' if state else 'off'}")
        self.status.set(f"Quạt 12V – FC{ch} = {'BẬT' if state else 'TẮT'}")
        self.after(600, lambda: self._read_dfan_rps(ch))

    def _read_dfan_rps(self, ch):
        if not self._need_conn():
            return
        self._ui_beep()
        v = parse_kv(self._send_recv(f"dfanr {ch}", 1.2), f"DFAN{ch}_RPS")
        self.dfan_rps_lbl[ch - 1].config(text=f"{v}" if v is not None else "—")

    def _read_gpi(self):
        if not self._need_conn():
            return
        self._ui_beep()
        v = parse_kv(self._send_recv("gpi"), "GPI")
        if v is None:
            self.lbl_gpi.config(text="GPI = —")
            return
        self.lbl_gpi.config(text=f"GPI = {v:08b}b")   # nhị phân (bit7..bit0)
        for i in range(8):
            dc, oid = self.gpi_dots[i]
            dc.itemconfig(oid, fill=DOT_ON if (v >> i) & 1 else DOT_OFF)

    def _read_flash(self):
        if not self._need_conn():
            return
        self._ui_beep()
        r = self._send_recv("flash")
        m = re.search(r"FLASH_ID=([0-9A-Fa-fx ]+)", r)
        ok = "FLASH=OK" in r
        txt = (m.group(1).strip() if m else "?") + ("  OK" if ok else "  FAIL")
        self.lbl_flash.config(text=txt, fg=PASS_FG if ok else FAIL_FG)

    # ---- Cảm biến / nguồn / RS485 callbacks ----
    def _read_ntc(self):
        if not self._need_conn():
            return
        self._ui_beep()
        for ch in range(1, 5):
            v = parse_kv(self._send_recv(f"ntc {ch}"), f"NTC{ch}")
            self.ntc_lbl[ch - 1].config(text=str(v) if v is not None else "—")

    def _read_hum(self):
        if not self._need_conn():
            return
        self._ui_beep()
        for ch in range(1, 3):
            v = parse_kv(self._send_recv(f"hum {ch}"), f"HUM{ch}")
            self.hum_lbl[ch - 1].config(text=str(v) if v is not None else "—")

    def _on_humpwr(self, ch, state):
        if not self._need_conn():
            return
        self._ui_beep()
        self._send(f"humpwr {ch} {'on' if state else 'off'}")
        self.status.set(f"5V H{ch} = {'BẬT' if state else 'TẮT'}")

    def _on_rspwr(self, state):
        if not self._need_conn():
            return
        self._ui_beep()
        self._send(f"rspwr {'on' if state else 'off'}")
        self.status.set(f"12V RS485 = {'BẬT' if state else 'TẮT'}")
        self.after(300, self._read_pg)

    def _read_pg(self):
        if not self._need_conn():
            return
        v = parse_kv(self._send_recv("rspg"), "RS485_PG")
        dc, oid = self.pg_dot
        dc.itemconfig(oid, fill=DOT_ON if v else DOT_OFF)

    def _probe_rs485(self, p):
        if not self._need_conn():
            return
        self._ui_beep()
        ok = False
        for _ in range(2):                       # thu lai 1 lan neu truot
            if rs485_pass(self._send_recv(f"rs485 {p}", 1.0)):
                ok = True
                break
        lb = self.rs485_lbl[p - 1]
        lb.config(text=("OK" if ok else "FAIL"),
                  fg=PASS_FG if ok else FAIL_FG)

    # ---------------- Auto test ----------------
    def toggle_pause(self):
        self._paused = not self._paused
        self.btn_pause.config(text=("▶ Tiếp tục" if self._paused else "⏸ Tạm dừng"))
        self.status.set("Đã tạm dừng." if self._paused else "Tiếp tục chạy...")

    def cancel_test(self):
        self._cancel = True
        self._paused = False
        self.status.set("Đang hủy...")

    def _modal(self, title, msg, buttons):
        """Hien the thong bao modal. buttons: list (label, value). Tra ve value."""
        dlg = tk.Toplevel(self)
        dlg.title(title)
        dlg.configure(bg=WHITE)
        dlg.transient(self)
        dlg.resizable(False, False)
        result = {"v": None}
        tk.Label(dlg, text=title, bg=WHITE, fg=BLUE_ACC,
                 font=("Segoe UI", 12, "bold"), wraplength=420,
                 justify="left").pack(anchor="w", padx=18, pady=(16, 4))
        tk.Label(dlg, text=msg, bg=WHITE, fg=GREY_TX, font=FONT,
                 wraplength=420, justify="left").pack(anchor="w", padx=18, pady=(0, 14))
        bar = tk.Frame(dlg, bg=WHITE)
        bar.pack(fill="x", padx=18, pady=(0, 16))

        def _choose(val):
            result["v"] = val
            dlg.destroy()
        for label, val, bg, hov in buttons:
            fgc = SEC_TX if bg == SEC_BG else "white"   # nut nen sang -> chu toi
            flat_btn(bar, label, bg, hov, fg=fgc, padx=14,
                     command=lambda v=val: _choose(v)).pack(side="right", padx=4)
        dlg.update_idletasks()
        w, h = dlg.winfo_reqwidth(), dlg.winfo_reqheight()
        x = self.winfo_rootx() + (self.winfo_width() - w) // 2
        y = self.winfo_rooty() + (self.winfo_height() - h) // 3
        dlg.geometry(f"+{max(0,x)}+{max(0,y)}")
        dlg.grab_set()
        self.wait_window(dlg)
        return result["v"]

    def _prompt_continue(self, group):
        msg = PREP_PROMPT.get(group, "Chuẩn bị xong, bấm Tiếp tục.")
        return self._modal(
            f"Chuẩn bị: {DISPLAY_NAME.get(group, group)}", msg,
            [("Tiếp tục", "ok", ON_BG, ON_HOV),
             ("Bỏ qua nhóm", "skip", OFF_BG, OFF_HOV),
             ("Hủy test", "cancel", FAIL_FG, "#b91c1c")])

    def _prompt_passfail(self, title, msg):
        return self._modal(
            title, msg,
            [("Đạt (PASS)", True, ON_BG, ON_HOV),
             ("Không đạt (FAIL)", False, FAIL_FG, "#b91c1c"),
             ("Bỏ qua", None, OFF_BG, OFF_HOV)])

    def run_all(self):
        if not self._need_conn():
            messagebox.showinfo("Chưa kết nối", "Hãy mở cổng COM trước.")
            return
        if hasattr(self, "ent_serial") and not self.ent_serial.get().strip():
            if not messagebox.askyesno("Chưa nhập Serial",
                                       "Chưa nhập số Serial của board.\n"
                                       "Vẫn tiếp tục? (report sẽ dùng mã tạm)"):
                return
        self._cancel = False
        self._paused = False
        self._gpi_done_groups = set()    # nhom GPI da test gop trong 1 bang
        self.btn_runall.config(state="disabled")
        self.btn_pause.config(state="normal", text="⏸ Tạm dừng")
        self.btn_cancel.config(state="normal")
        skipped_groups = set()
        cur_group = None
        try:
            for iid in self.step_items:
                if self._cancel:
                    break
                while self._paused and not self._cancel:   # chờ khi tạm dừng
                    self.update()
                    time.sleep(0.1)
                if self._cancel:
                    break
                m = self._meta[iid]
                g = m["group"]
                # Bỏ popup "chuẩn bị thiết bị" — vào thẳng popup test của chức năng
                # (popup test đã có sẵn hướng dẫn cắm/đấu thiết bị bên trong).
                if g in skipped_groups or m["skip"]:
                    continue
                self.run_step(iid, batch=True)
                self.update()
        finally:
            cancelled = self._cancel
            self._cancel = False
            self._paused = False
            self.btn_runall.config(state="normal")
            self.btn_pause.config(state="disabled", text="⏸ Tạm dừng")
            self.btn_cancel.config(state="disabled")
        self.status.set("Đã hủy test." if cancelled else "Run test xong.")
        self._show_summary(cancelled)

    def _show_summary(self, cancelled=False):
        total = len(self.step_items)
        done = sum(1 for i in self.step_items if self._meta[i]["result"])
        npass = sum(1 for i in self.step_items if self._meta[i]["result"] == "PASS")
        nfail = sum(1 for i in self.step_items if self._meta[i]["result"] == "FAIL")
        final = "DỪNG GIỮA CHỪNG" if cancelled else ("PASS" if nfail == 0 else "FAIL")
        serial_no = self.ent_serial.get().strip() or "(chưa nhập serial)"
        head = ("ĐÃ HỦY TEST\n\n" if cancelled else
                ("✓ KẾT QUẢ CUỐI: PASS\n\n" if nfail == 0 else "✗ KẾT QUẢ CUỐI: FAIL\n\n"))
        msg = (head + f"Serial: {serial_no}\n"
               f"Đã chạy: {done}/{total}\n"
               f"PASS: {npass}    FAIL: {nfail}\n\n"
               f"Ghi kết quả vào file {REPORT_NAME}?")
        # Cuoi test: hoi co ghi vao Excel khong (giong Smart PDU)
        if messagebox.askyesno("Kết quả test — " + final, msg):
            self._save_report(stopped=cancelled, silent=False)
            # Xoa serial de nhap board ke tiep (se ghi noi tiep)
            try:
                self.ent_serial.delete(0, "end")
                self.ent_serial.focus_set()
            except Exception:
                pass
        else:
            self._log("[Đã bỏ qua lưu report (người dùng chọn)]\n", "tx")

    def run_step(self, iid, batch=False):
        m = self._meta[iid]
        if not self._need_conn():
            return
        # Bước bán tự động: mở popup tương ứng (đã gộp cả nhóm thành 1 bảng)
        pk = m.get("panel_kind")
        if pk:
            if pk == "gpi":
                self._run_gpi_panel(iid, m, batch)
            elif pk in ("fan", "dfan", "gpo", "rs485"):
                self._run_port_panel(iid, m, pk, batch)
            elif pk in ("ntc", "hum"):
                self._run_sensor_panel(iid, m, pk, batch)
            elif pk == "ac":
                self._run_relay_panel(iid, m, batch)
            return
        plan = resolve_plan(m, self._current_thresholds(),
                            getattr(self, "fan_speed", 80))
        if plan.get("mode") == "manual":
            self._set_result(iid, "thủ công", None)
            return
        if plan.get("mode") == "skip":           # bỏ qua trong Auto Test
            self._set_result(iid, "bỏ qua", None)
            if not batch:
                self.status.set(f"{m['des']}: bỏ qua")
            return
        # Thông báo chuẩn bị (cắm / không cắm...) trước khi đọc
        if plan.get("prompt"):
            if not messagebox.askokcancel(
                    f"{m['group']} – {m['des']}",
                    f"{plan['prompt']}\n\nChuẩn bị xong → bấm OK để đọc & kiểm tra."
                    "\n(Cancel = bỏ qua bước này)"):
                self._set_result(iid, "bỏ qua", None)
                return
        self._beep()                             # còi báo có bước test (nếu bật)
        self.status.set(f"Đang chạy: {m['group']} – {m['des']}")
        self.tree.item(iid, tags=("running",))   # xanh = dang chay
        self.update()
        for cmd in plan.get("cmds", []):
            self._send(cmd)
            time.sleep(0.05)
        self._sync_dashboard(m)          # dashboard chạy theo (giống Smart PDU)
        time.sleep(plan.get("settle", 0.3))

        mode = plan["mode"]
        if mode == "control_manual":     # gui lenh xong -> do tay (Volt ke) -> xac nhan
            hint = plan.get("hint", "")
            ans = self._prompt_passfail(f"{m['group']} – {m['des']}",
                                        f"{hint}\n\nKết quả đo có đạt không?")
            if ans is None:
                self._set_result(iid, hint, None)   # bo qua xac nhan -> MANUAL
            else:
                self._set_result(iid, hint, ans)
            return
        if mode == "system":             # buoc: kiem Flash W25Q80
            r_flash = self._send_recv("flash", 1.2)
            ok_flash = "FLASH=OK" in r_flash
            val = f"Flash:{'OK' if ok_flash else 'X'}"
            self._set_result(iid, val, ok_flash)
            if not batch:
                self.status.set(f"{m['des']}: {'PASS' if ok_flash else 'FAIL'} ({val})")
            return
        if mode == "pi":                 # buoc: giao tiep Orange Pi (loopback UART3)
            r = self._send_recv("pi", 1.5)
            ok = "PI=OK" in r
            self._set_result(iid, "OK" if ok else "FAIL", ok)
            if not batch:
                self.status.set(f"{m['des']}: {'PASS' if ok else 'FAIL'}")
            return
        if mode == "multi":              # doc nhieu kenh, tat ca trong nguong
            lo, hi = plan["expect"]
            vals = []
            ok_all = True
            for cmd, key in plan["reads"]:
                v = parse_kv(self._send_recv(cmd, 0.8), key)
                vals.append("?" if v is None else str(v))
                if v is None or not (lo <= v <= hi):
                    ok_all = False
            for cmd in plan.get("post_cmds", []):   # vd: tat nguon 5V sau khi doc
                self._send(cmd)
                time.sleep(0.05)
            self._sync_multi_labels(m, vals)
            self._set_result(iid, ",".join(vals), ok_all)
            if not batch:
                self.status.set(f"{m['des']}: {'PASS' if ok_all else 'FAIL'}")
            return

        resp = self._send_recv(plan["read"], 1.2) if plan.get("read") else ""
        ok, val = self._evaluate(plan, resp)
        self._sync_value_label(m, val)
        self._set_result(iid, val, ok)
        if not batch:
            self.status.set(f"{m['des']}: {'PASS' if ok else 'FAIL'} ({val})")

    def _run_gpi_panel(self, step_iid, m, batch=False):
        """Bảng test 8 cổng đầu vào GPI trong 1 cửa sổ: kích lần lượt I1–I8,
        ô chuyển XANH khi nhận tín hiệu; xong bấm 'Xác nhận'. Cổng chưa kích -> FAIL.
        Hiện ngay tại vị trí con trỏ chuột; kết quả ghi gộp vào 1 dòng."""
        nbits = 8
        self._beep()
        dlg = tk.Toplevel(self)
        dlg.title("Đầu vào GPI — kích 8 cổng I1–I8")
        dlg.configure(bg=WHITE)
        dlg.transient(self)
        dlg.withdraw()
        self._panel_header(dlg, step_iid, "Đầu vào GPI (I1–I8)")
        tk.Label(dlg, text="Bước 1 (tự động): KHÔNG cắm cổng nào → kiểm tra GPI=255.\n"
                 "Bước 2: kích lần lượt 8 cổng (cắm 48V vào I1…I8); ô XANH khi nhận.\n"
                 "Xong bấm “Xác nhận”.",
                 bg=WHITE, fg=INK, font=FONT, justify="left").pack(padx=18,
                                                                   pady=(14, 8))
        grid = tk.Frame(dlg, bg=WHITE)
        grid.pack(padx=18, pady=4)
        dots = {}
        detected = {b: False for b in range(nbits)}
        for b in range(nbits):
            cell = tk.Frame(grid, bg=WHITE)
            cell.grid(row=0, column=b, padx=6)
            dc, oid = make_dot(cell, 22, DOT_OFF)
            dc.pack()
            tk.Label(cell, text=f"I{b+1}", bg=WHITE, fg=GREY_TX,
                     font=FONT_SM).pack()
            dots[b] = (dc, oid)
        lbl_np = tk.Label(dlg, text="Chưa cắm: (đang đọc…)", bg=WHITE,
                          fg=HDR_SUB, font=FONT_SM)
        lbl_np.pack(pady=(2, 0))
        prog = tk.Label(dlg, text="Đã nhận: 0 / 8", bg=WHITE, fg=BLUE_ACC,
                        font=FONT_B)
        prog.pack(pady=(4, 4))
        state = {"done": False, "cancel": False}
        btnf = tk.Frame(dlg, bg=WHITE)
        btnf.pack(pady=(6, 14))
        flat_btn(btnf, "✓ Xác nhận xong", PRIMARY, PRIMARY_HOV, padx=12, pady=4,
                 command=lambda: state.update(done=True)).pack(side="left", padx=6)
        sec_btn(btnf, "Dừng", padx=12, pady=4,
                command=lambda: state.update(cancel=True)).pack(side="left", padx=6)
        dlg.protocol("WM_DELETE_WINDOW", lambda: state.update(done=True))
        self._center_dialog(dlg)
        dlg.deiconify()
        dlg.grab_set()
        v0 = parse_kv(self._send_recv("gpi", 0.4), "GPI")   # bước "không cắm"
        noplug_ok = (v0 == 255)
        lbl_np.config(text=("Chưa cắm: GPI=255 ✓" if noplug_ok
                            else f"Chưa cắm: GPI={v0} ✗ (hãy rút hết cổng!)"),
                      fg=PASS_FG if noplug_ok else FAIL_FG)
        while not state["done"] and not state["cancel"]:
            v = parse_kv(self._send_recv("gpi", 0.3), "GPI")
            if v is not None:
                for b in range(nbits):
                    if not detected[b] and ((v >> b) & 1) == 0:
                        detected[b] = True
                        dc, oid = dots[b]
                        dc.itemconfig(oid, fill=PASS_FG)
                ndet = sum(detected.values())
                if ndet >= nbits:
                    prog.config(text="Đã nhận đủ 8 / 8 ✓", fg=PASS_FG)
                else:
                    prog.config(text=f"Đã nhận: {ndet} / 8", fg=BLUE_ACC)
            try:
                dlg.update()
            except tk.TclError:
                break
            time.sleep(0.15)
        try:
            dlg.grab_release()
            dlg.destroy()
        except Exception:
            pass
        # Ghi kết quả GỘP vào 1 dòng (PASS nếu đủ 8 cổng, FAIL kèm cổng thiếu)
        npass = sum(1 for v in detected.values() if v)
        fails = [f"I{b+1}" for b in range(nbits) if not detected.get(b)]
        ok_all = noplug_ok and (npass == nbits)
        if not noplug_ok:
            val = "lỗi chưa-cắm (GPI≠255)"
        elif npass == nbits:
            val = "8/8 OK"
        else:
            val = "thiếu: " + ",".join(fails)
        self._set_result(step_iid, val, ok_all)
        self.status.set(f"GPI: {npass}/{nbits} cổng đạt"
                        + (" — đã DỪNG" if state["cancel"] else ""))
        if state["cancel"]:
            self._cancel = True

    def _run_port_panel(self, step_iid, m, kind, batch=False):
        """Bảng kiểm tra 4 cổng (Quạt 48V / Đầu ra GPO) trong 1 cửa sổ — giống GPI.
        Cắm/đấu lần lượt từng cổng; ô chuyển XANH khi giá trị đạt ngưỡng."""
        th = self._current_thresholds()
        if kind == "fan":
            ports = [1, 2, 3, 4]
            prefix, unit = "Fan", "rpm"
            lo, hi = th.get("Quạt", DEFAULT_TH["Quạt"])
            title = "Quạt 48V — cắm & kiểm tra 4 quạt"
            hint = ("Cắm lần lượt từng quạt 48V vào jig.\n"
                    "Ô chuyển XANH khi tốc độ đạt ngưỡng. Xong bấm “Xác nhận”.")
            independent = True
            toggles = getattr(self, "fan_toggles", None)

            def on_cmd(p):
                self._send(f"fan {p} on {getattr(self, 'fan_speed', 80)}")

            def off_cmd(p):
                self._send(f"fan {p} off")

            def read_val(p):
                return parse_kv(self._send_recv(f"fanr {p}", 1.0), f"FAN{p}_RPS")
        elif kind == "dfan":
            ports = [1, 2]
            prefix, unit = "FC", "rpm"
            lo, hi = th.get("Quạt nhỏ", DEFAULT_TH["Quạt nhỏ"])
            title = "Quạt 12V — bấm Test từng quạt"
            hint = ""
            independent = True
            toggles = getattr(self, "dfan_toggles", None)

            def on_cmd(p):
                self._send(f"dfan {p} on")

            def off_cmd(p):
                self._send(f"dfan {p} off")

            def read_val(p):
                return parse_kv(self._send_recv(f"dfanr {p}", 1.0), f"DFAN{p}_RPS")
        else:
            ports = [1, 2, 3, 4]
            prefix, unit = "O", "mA"
            lo, hi = th.get("Đầu ra", DEFAULT_TH["Đầu ra"])
            title = "Đầu ra GPO — kiểm tra 4 cổng"
            hint = ("Đấu tải vào từng cổng O1–O4 (đo dòng tổng).\n"
                    "Hệ thống bật lần lượt từng cổng; ô XANH khi dòng đạt ngưỡng.")
            independent = False
            toggles = getattr(self, "gpo_toggles", None)

            def on_cmd(p):
                self._send(f"gpo {p} on")

            def off_cmd(p):
                self._send(f"gpo {p} off")

            def read_val(p):
                return parse_kv(self._send_recv("gpoi", 0.8), "GPO_CUR")

        if kind == "rs485":                      # ghi đè: probe gửi & nhận lại
            ports = [1, 2, 3, 4]
            prefix, unit = "RS", "byte"
            lo, hi = 1, 9999
            title = "Cổng RS485 — bấm Test từng cổng (gửi & nhận lại)"
            toggles = None

            def on_cmd(p):
                pass

            def off_cmd(p):
                pass

            def read_val(p):
                # gui & nhan lai; thu toi 2 lan de bot truot (chap chon duong ve)
                s = ""
                for _ in range(2):
                    s = rs485_rx_str(self._send_recv(f"rs485 {p}", 1.0))
                    if RS485_EXPECT in s.upper():
                        break
                return s

        n = len(ports)
        settle = 0.1 if kind == "rs485" else (0.5 if kind == "gpo" else 1.4)
        self._beep()
        dlg = tk.Toplevel(self)
        dlg.title(title)
        dlg.configure(bg=WHITE)
        dlg.transient(self)
        dlg.withdraw()
        self._panel_header(dlg, step_iid, title)
        tk.Label(dlg, text="Cắm/đấu cổng cần kiểm tra rồi bấm “Test” từng cổng "
                 "(không tự động).", bg=WHITE, fg=INK, font=FONT,
                 justify="left").pack(padx=18, pady=(14, 4))
        th_txt = ('Đạt khi nhận đúng chuỗi xác nhận từ AK MCU KIT (chứa "OK")'
                  if kind == "rs485" else f"Ngưỡng đạt: {lo:g}–{hi:g} {unit}")
        tk.Label(dlg, text=th_txt, bg=WHITE,
                 fg=HDR_SUB, font=FONT_SM).pack(padx=18, pady=(0, 6))
        grid = tk.Frame(dlg, bg=WHITE)
        grid.pack(padx=18, pady=4)
        cells = {}
        passed = {p: None for p in ports}        # None=chưa test, True/False

        def _tog(p, on):
            if toggles and 0 <= p - 1 < len(toggles):
                try:
                    toggles[p - 1].set(on)
                except Exception:
                    pass

        def _refresh_prog():
            nd = sum(1 for x in passed.values() if x)
            ndone = sum(1 for x in passed.values() if x is not None)
            prog.config(text=f"Đạt {nd}/{n}  ·  đã đo {ndone}/{n}",
                        fg=PASS_FG if nd == n else BLUE_ACC)

        def do_test(p):
            dc, oid, vlb, btn = cells[p]
            btn.config(state="disabled", text="…")
            dc.itemconfig(oid, fill=RUN_FG)
            vlb.config(text="đang đo", fg=RUN_FG)
            on_cmd(p); _tog(p, True)
            dlg.update()
            time.sleep(settle)
            v = read_val(p)
            off_cmd(p); _tog(p, False)
            if kind == "rs485":                  # so chuoi AK MCU tra ve
                ok = RS485_EXPECT in str(v).upper()
                disp = ("OK" if ok else ("—" if not v else "FAIL"))
            else:
                ok = (v is not None and lo <= v <= hi)
                disp = (f"{v} {unit}" if v is not None else "—")
            passed[p] = ok
            dc.itemconfig(oid, fill=PASS_FG if ok else FAIL_FG)
            vlb.config(text=disp, fg=PASS_FG if ok else FAIL_FG)
            btn.config(state="normal", text="Test lại")
            _refresh_prog()

        for i, p in enumerate(ports):
            cell = tk.Frame(grid, bg=WHITE)
            cell.grid(row=0, column=i, padx=10)
            dc, oid = make_dot(cell, 22, DOT_OFF)
            dc.pack()
            tk.Label(cell, text=f"{prefix}{p}", bg=WHITE, fg=GREY_TX,
                     font=FONT_SM).pack()
            vlb = tk.Label(cell, text="—", bg=WHITE, fg=BLUE_ACC, font=FONT_B)
            vlb.pack()
            btn = flat_btn(cell, "Test", READ_BG, READ_HOV, padx=12, pady=2,
                           command=lambda pp=p: do_test(pp))
            btn.pack(pady=(3, 0))
            cells[p] = (dc, oid, vlb, btn)
        prog = tk.Label(dlg, text=f"Đạt 0/{n}  ·  đã đo 0/{n}", bg=WHITE,
                        fg=BLUE_ACC, font=FONT_B)
        prog.pack(pady=(8, 4))
        state = {"cancel": False}
        btnf = tk.Frame(dlg, bg=WHITE)
        btnf.pack(pady=(6, 14))
        flat_btn(btnf, "✓ Xác nhận xong", PRIMARY, PRIMARY_HOV, padx=12, pady=4,
                 command=dlg.destroy).pack(side="left", padx=6)

        def _cancel():
            state["cancel"] = True
            dlg.destroy()

        sec_btn(btnf, "Dừng", padx=12, pady=4,
                command=_cancel).pack(side="left", padx=6)
        dlg.protocol("WM_DELETE_WINDOW", dlg.destroy)
        self._center_dialog(dlg)
        dlg.deiconify()
        dlg.grab_set()
        dlg.wait_window()
        for p in ports:                          # tắt hết cho an toàn
            try:
                off_cmd(p); _tog(p, False)
            except Exception:
                pass
        npass = sum(1 for p in ports if passed[p])
        fails = [f"{prefix}{p}" for p in ports if not passed[p]]
        ok_all = (npass == n)
        val = f"{npass}/{n} OK" if ok_all else ("lỗi: " + ",".join(fails))
        self._set_result(step_iid, val, ok_all)
        self.status.set(f"{title}: {npass}/{n} đạt"
                        + (" — đã DỪNG" if state["cancel"] else ""))
        if state["cancel"]:
            self._cancel = True

    def _run_sensor_panel(self, step_iid, m, kind, batch=False):
        """Cảm biến Nhiệt độ/Độ ẩm bán tự động: 'Test KHÔNG cắm' (hở ~255) và
        'Test CÓ cắm' (trong ngưỡng) — giống ý tưởng GPI."""
        th = self._current_thresholds()
        if kind == "ntc":
            ports = [1, 2, 3, 4]; prefix = "NTC"
            lo, hi = th.get("Nhiệt độ", DEFAULT_TH["Nhiệt độ"])
            title = "Nhiệt độ (NTC1–4) — bán tự động"
            powered = False

            def read_val(p):
                return parse_kv(self._send_recv(f"ntc {p}", 0.6), f"NTC{p}")

            def open_ok(v):                       # NTC hở -> node treo lên 3V3 (~255)
                return v is not None and v >= 240
            open_txt = "≥240 (hở ~255)"
        else:
            ports = [1, 2]; prefix = "H"
            lo, hi = th.get("Độ ẩm", DEFAULT_TH["Độ ẩm"])
            title = "Độ ẩm (H1–2) — bán tự động"
            powered = True

            def read_val(p):
                return parse_kv(self._send_recv(f"hum {p}", 0.6), f"HUM{p}")

            def open_ok(v):                       # HUM chưa cắm -> đã xả tụ, thả nổi (~0)
                return v is not None and v <= 20
            open_txt = "≤20 (đã xả tụ ~0)"

        self._beep()
        dlg = tk.Toplevel(self)
        dlg.title(title)
        dlg.configure(bg=WHITE)
        dlg.transient(self)
        dlg.withdraw()
        self._panel_header(dlg, step_iid, title)
        tk.Label(dlg, text="Bước 1: CHƯA cắm cảm biến → bấm “Test KHÔNG cắm”.\n"
                 "Bước 2: cắm cảm biến → bấm “Test CÓ cắm”.",
                 bg=WHITE, fg=INK, font=FONT, justify="left").pack(padx=18,
                                                                   pady=(14, 4))
        tk.Label(dlg, text=f"Không cắm: {open_txt}   ·   Có cắm: {lo:g}–{hi:g}",
                 bg=WHITE, fg=HDR_SUB, font=FONT_SM).pack(padx=18, pady=(0, 6))
        grid = tk.Frame(dlg, bg=WHITE)
        grid.pack(padx=18, pady=4)
        vlbs = {}
        for i, p in enumerate(ports):
            cell = tk.Frame(grid, bg=WHITE)
            cell.grid(row=0, column=i, padx=12)
            tk.Label(cell, text=f"{prefix}{p}", bg=WHITE, fg=GREY_TX,
                     font=FONT_SM).pack()
            vlb = tk.Label(cell, text="—", bg=WHITE, fg=BLUE_ACC, font=FONT_B)
            vlb.pack()
            vlbs[p] = vlb
        res = {"noplug": None, "plug": None}
        st_np = tk.Label(dlg, text="Không cắm: chưa test", bg=WHITE, fg=HDR_SUB,
                         font=FONT_B)
        st_np.pack(pady=(8, 0))
        st_pl = tk.Label(dlg, text="Có cắm: chưa test", bg=WHITE, fg=HDR_SUB,
                         font=FONT_B)
        st_pl.pack()

        def read_all():
            vals = {}
            for p in ports:
                v = read_val(p)
                vals[p] = v
                vlbs[p].config(text=str(v) if v is not None else "—")
                dlg.update()
            return vals

        def test_noplug():
            vals = read_all()
            ok = all(open_ok(v) for v in vals.values())
            res["noplug"] = ok
            st_np.config(text=("Không cắm: ĐẠT ✓" if ok
                               else f"Không cắm: LỖI ✗ (cần {open_txt})"),
                         fg=PASS_FG if ok else FAIL_FG)

        def test_plug():
            if powered:
                for p in ports:
                    self._send(f"humpwr {p} on")
                dlg.update()
                time.sleep(0.4)
            vals = read_all()
            if powered:
                for p in ports:
                    self._send(f"humpwr {p} off")
            ok = all((v is not None and lo <= v <= hi) for v in vals.values())
            res["plug"] = ok
            st_pl.config(text=("Có cắm: ĐẠT ✓" if ok
                               else "Có cắm: LỖI ✗ (ngoài ngưỡng)"),
                         fg=PASS_FG if ok else FAIL_FG)

        bf = tk.Frame(dlg, bg=WHITE)
        bf.pack(pady=(8, 4))
        flat_btn(bf, "Test KHÔNG cắm", PRIMARY, PRIMARY_HOV, padx=10, pady=4,
                 command=test_noplug).pack(side="left", padx=4)
        flat_btn(bf, "Test CÓ cắm", PRIMARY, PRIMARY_HOV, padx=10, pady=4,
                 command=test_plug).pack(side="left", padx=4)
        state = {"cancel": False}
        bf2 = tk.Frame(dlg, bg=WHITE)
        bf2.pack(pady=(2, 14))
        flat_btn(bf2, "✓ Xác nhận xong", PRIMARY, PRIMARY_HOV, padx=12, pady=4,
                 command=dlg.destroy).pack(side="left", padx=6)

        def _cancel():
            state["cancel"] = True
            dlg.destroy()

        sec_btn(bf2, "Dừng", padx=12, pady=4,
                command=_cancel).pack(side="left", padx=6)
        dlg.protocol("WM_DELETE_WINDOW", dlg.destroy)
        self._center_dialog(dlg)
        dlg.deiconify()
        dlg.grab_set()
        dlg.wait_window()
        ok_all = bool(res["noplug"]) and bool(res["plug"])

        def _mk(x):
            return "OK" if x else ("X" if x is False else "-")
        val = f"hở:{_mk(res['noplug'])} cắm:{_mk(res['plug'])}"
        self._set_result(step_iid, val, ok_all)
        self.status.set(f"{title}: {'ĐẠT' if ok_all else 'CHƯA ĐẠT'}"
                        + (" — đã DỪNG" if state["cancel"] else ""))
        if state["cancel"]:
            self._cancel = True

    def _run_relay_panel(self, step_iid, m, batch=False):
        """Relay AC bán tự động: 'Test BẬT' (đóng relay, dòng trong ngưỡng) và
        'Test TẮT' (mở relay, dòng ~0)."""
        th = self._current_thresholds()
        lo, hi = th.get("Relay AC", DEFAULT_TH["Relay AC"])
        OFF_MAX = 50
        self._beep()
        dlg = tk.Toplevel(self)
        dlg.title("Relay AC — bán tự động")
        dlg.configure(bg=WHITE)
        dlg.transient(self)
        dlg.withdraw()
        self._panel_header(dlg, step_iid, "Relay AC")
        tk.Label(dlg, text="Đấu tải AC vào relay. Bấm “Test BẬT” rồi “Test TẮT”.",
                 bg=WHITE, fg=INK, font=FONT, justify="left").pack(padx=18,
                                                                   pady=(14, 4))
        tk.Label(dlg, text=f"Bật: {lo:g}–{hi:g} mA   ·   Tắt: ≤{OFF_MAX} mA",
                 bg=WHITE, fg=HDR_SUB, font=FONT_SM).pack(padx=18, pady=(0, 6))
        res = {"on": None, "off": None}
        toggle = getattr(self, "ac_toggle", None)
        rowf = tk.Frame(dlg, bg=WHITE)
        rowf.pack(padx=18, pady=6)
        st_on = tk.Label(rowf, text="BẬT: chưa test", bg=WHITE, fg=HDR_SUB,
                         font=FONT_B, width=24, anchor="w")
        st_on.grid(row=0, column=0, sticky="w", pady=2)
        st_off = tk.Label(rowf, text="TẮT: chưa test", bg=WHITE, fg=HDR_SUB,
                          font=FONT_B, width=24, anchor="w")
        st_off.grid(row=1, column=0, sticky="w", pady=2)

        def _tog(on):
            if toggle:
                try:
                    toggle.set(on)
                except Exception:
                    pass

        def test_on():
            self._send("ac on"); _tog(True)
            dlg.update()
            time.sleep(0.5)
            v = parse_kv(self._send_recv("aci", 0.8), "AC_CUR")
            ok = (v is not None and lo <= v <= hi)
            res["on"] = ok
            st_on.config(text=f"BẬT: {v} mA {'✓' if ok else '✗'}",
                         fg=PASS_FG if ok else FAIL_FG)

        def test_off():
            self._send("ac off"); _tog(False)
            dlg.update()
            time.sleep(0.4)
            v = parse_kv(self._send_recv("aci", 0.8), "AC_CUR")
            ok = (v is not None and v <= OFF_MAX)
            res["off"] = ok
            st_off.config(text=f"TẮT: {v} mA {'✓' if ok else '✗'}",
                          fg=PASS_FG if ok else FAIL_FG)

        bf = tk.Frame(dlg, bg=WHITE)
        bf.pack(pady=(4, 4))
        flat_btn(bf, "Test BẬT", PRIMARY, PRIMARY_HOV, padx=12, pady=4,
                 command=test_on).pack(side="left", padx=4)
        flat_btn(bf, "Test TẮT", PRIMARY, PRIMARY_HOV, padx=12, pady=4,
                 command=test_off).pack(side="left", padx=4)
        state = {"cancel": False}
        bf2 = tk.Frame(dlg, bg=WHITE)
        bf2.pack(pady=(4, 14))
        flat_btn(bf2, "✓ Xác nhận xong", PRIMARY, PRIMARY_HOV, padx=12, pady=4,
                 command=dlg.destroy).pack(side="left", padx=6)

        def _cancel():
            state["cancel"] = True
            dlg.destroy()

        sec_btn(bf2, "Dừng", padx=12, pady=4,
                command=_cancel).pack(side="left", padx=6)
        dlg.protocol("WM_DELETE_WINDOW", dlg.destroy)
        self._center_dialog(dlg)
        dlg.deiconify()
        dlg.grab_set()
        dlg.wait_window()
        self._send("ac off"); _tog(False)

        def _mk(x):
            return "OK" if x else ("X" if x is False else "-")
        ok_all = bool(res["on"]) and bool(res["off"])
        val = f"bật:{_mk(res['on'])} tắt:{_mk(res['off'])}"
        self._set_result(step_iid, val, ok_all)
        self.status.set(f"Relay AC: {'ĐẠT' if ok_all else 'CHƯA ĐẠT'}"
                        + (" — đã DỪNG" if state["cancel"] else ""))
        if state["cancel"]:
            self._cancel = True

    def _evaluate(self, plan, resp):
        mode = plan["mode"]
        if mode == "numeric":
            v = parse_kv(resp, plan["key"])
            if v is None:
                return False, "no-resp"
            lo, hi = plan["expect"]
            return (lo <= v <= hi), v
        if mode == "gpi_all":
            v = parse_kv(resp, "GPI")
            return (v == 255), (v if v is not None else "no-resp")
        if mode == "gpi_bit":
            v = parse_kv(resp, "GPI")
            if v is None:
                return False, "no-resp"
            bit = (v >> plan["bit"]) & 1
            return (bit == plan["expect"]), f"bit{plan['bit']}={bit}"
        if mode == "flag":
            v = parse_kv(resp, plan["key"])
            return (v == plan["expect"]), (v if v is not None else "no-resp")
        if mode == "flash":
            ok = "FLASH=OK" in resp
            return ok, ("OK" if ok else "FAIL")
        if mode == "rs485":
            s = rs485_rx_str(resp)
            return rs485_pass(resp), (s if s else "no-resp")
        if mode == "read_ok":
            v = parse_kv(resp, plan["key"])
            return (v is not None), (v if v is not None else "no-resp")
        return False, "?"

    def _set_result(self, iid, value, ok):
        m = self._meta[iid]
        m["value"] = "" if value is None else str(value)
        if ok is None:
            m["result"] = "MANUAL"
            tag = ()
        else:
            m["result"] = "PASS" if ok else "FAIL"
            tag = ("pass",) if ok else ("fail",)
        self.tree.item(iid, values=(m["result"], m["value"] or "-"), tags=tag)
        self._refresh_progress()

    def reset_results(self):
        for iid, m in self._meta.items():
            m["value"] = ""
            m["result"] = ""
            self.tree.item(iid, values=("-", "-"), tags=("pending",))
        self.status.set("Đã reset kết quả.")
        self._refresh_progress()

    def _refresh_progress(self):
        total = len(self.step_items)
        npass = sum(1 for i in self.step_items if self._meta[i]["result"] == "PASS")
        nfail = sum(1 for i in self.step_items if self._meta[i]["result"] == "FAIL")
        done = sum(1 for i in self.step_items if self._meta[i]["result"])
        self.lbl_prog_n.config(text=f"{done} / {total}")
        self.lbl_pass.config(text=f"PASS {npass}")
        self.lbl_fail.config(text=f"FAIL {nfail}")
        self.pbar["value"] = (done / total * 100) if total else 0

    # ---------------- Export (1 thiet bi = 1 hang ngang, append) ----------------
    def _step_labels(self):
        """Nhan cot cho tung buoc test (theo thu tu step_items)."""
        labels, seen = [], {}
        for iid in self.step_items:
            m = self._meta[iid]
            base = f"{m['group']}: {m['des']}".strip()
            n = seen.get(base, 0)
            seen[base] = n + 1
            labels.append(base if n == 0 else f"{base} ({n+1})")
        return labels

    def _collect_cells(self):
        """Gia tri tung cot + dem PASS/FAIL. Cell = gia tri (KQ) hoac KQ."""
        cells, npass, nfail = [], 0, 0
        for iid in self.step_items:
            m = self._meta[iid]
            res = m["result"] or "-"
            val = m["value"]
            if res == "PASS":
                npass += 1
            elif res == "FAIL":
                nfail += 1
            if val and res in ("PASS", "FAIL"):
                cells.append(f"{val} ({res})")
            else:
                cells.append(val or res)
        return cells, npass, nfail

    def _save_report(self, stopped=False, silent=True):
        if openpyxl is None:
            if not silent:
                messagebox.showerror("Thiếu thư viện", "Chưa cài openpyxl.")
            return
        serial_no = self.ent_serial.get().strip() if hasattr(self, "ent_serial") else ""
        if not serial_no:
            serial_no = "NA-" + datetime.datetime.now().strftime("%H%M%S")
        cells, npass, nfail = self._collect_cells()
        ntotal = len(self.step_items)
        ket_luan = "DUNG GIUA" if stopped else ("OK" if nfail == 0 else "LOI")
        header = ["STT", "Serial", "Thoi gian", "COM", "Baud", "FW",
                  *self._step_labels(), "PASS/Tong", "Ket luan"]
        com = self.cmb_port.get().strip()
        baud = self.cmb_baud.get()
        path = report_path()
        try:
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
                ws = wb.active
                stt = ws.max_row   # tru dong header
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "OPMS Test"
                ws.append(header)
                stt = 1
            row = [stt, serial_no,
                   datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                   com, baud, self._fw_ver,
                   *cells, f"{npass}/{ntotal}", ket_luan]
            ws.append(row)
            wb.save(path)
            self.status.set(f"Đã ghi report ({serial_no}): {path}")
            self._log(f"[Đã ghi report: {serial_no} -> {ket_luan}]\n", "ok")
            if not silent:
                messagebox.showinfo("Xuất Excel",
                                    f"Đã ghi 1 dòng cho serial {serial_no} vào:\n{path}")
        except PermissionError:
            messagebox.showerror("Lỗi xuất Excel",
                                 f"Không ghi được {REPORT_NAME} — file đang mở trong "
                                 "Excel?\nĐóng file rồi thử lại.")
        except Exception as e:
            messagebox.showerror("Lỗi xuất Excel", str(e))

    def _on_close(self):
        try:
            if self.ser and self.ser.is_open:
                self.ser.close()
        except Exception:
            pass
        self.destroy()


def main():
    # 1 app duy nhat: co '--console' -> chay module test CH348 (cua so rieng).
    if "--console" in sys.argv:
        import ch348_test_app
        ch348_test_app.main()
        return
    App().mainloop()


if __name__ == "__main__":
    main()
