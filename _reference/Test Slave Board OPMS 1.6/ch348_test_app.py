# -*- coding: utf-8 -*-
"""
OPMS Console Board - CH348 8-COM Loopback Test App
==================================================
Kiem tra mach USB 8 cong COM (chip WCH CH348) bang LOOPBACK tung cong
(chap TX-RX moi cong): gui chuoi -> nhan lai dung tren cung cong = PASS.

- Tu do cong theo VID/PID CH348 (WCH, VID 0x1A86).
- Moi thiet bi ghi 1 hang ngang vao console_report.xlsx (append, khong tao file moi).
- Cung phong cach Industrial/Engineering voi app test slave board.

Mo tu app chinh (opms_test_app) qua nut "Test Console (CH348)" - phuong an A.
Chay truc tiep:  python ch348_test_app.py
Phu thuoc:       pip install pyserial openpyxl
"""

import os
import sys
import time
import datetime
import tkinter as tk
from tkinter import ttk, messagebox

try:
    import serial
    import serial.tools.list_ports
except ImportError:
    serial = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

APP_TITLE = "OPMS Console - CH348 Loopback Test (6 cổng)"
REPORT_NAME = "console_report.xlsx"
DEFAULT_BAUD = 115200
N_PORTS = 6                  # OPMS Console board: 6 cong console
CONSOLE_CHANNELS = "ABCDEF"  # chi test 6 kenh Ch A..F cua CH348 (bo G, H)
CH348_VID = 0x1A86           # WCH (CH34x/CH348)
# Chuoi test mac dinh (dai) - nguoi dung co the sua trong o "Chuoi test"
DEFAULT_TEST_STR = ("OPMS-1.6-CONSOLE-CH348-LOOPBACK-TEST-"
                    "0123456789-ABCDEFGHIJKLMNOPQRSTUVWXYZ-"
                    "abcdefghijklmnopqrstuvwxyz")

# ===== Theme dung chung (mau + font + nut + resource_path) tu opms_theme =====
from opms_theme import *   # noqa: F401,F403

# Header logos (rieng cua app Console)
HDR_LOGO_H = 52
OPMS_DEVICE_IMG = "USB_RS232_6PORT.png"
FPT_LOGO_IMG = "Vivoo_logo.png"


def report_path():
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, REPORT_NAME)


def list_ch348_ports():
    """Tra ve danh sach (device, kenh) CHI cua chip CH348 (USB nhieu cong, VID 1A86).

    Nhan dien (tranh quet nham cong COM khac, nhung van bat dung CH348):
      1. La cong WCH: VID = 0x1A86 HOAC ten/hwid co 'WCH' / '1A86'
         (vd Device Manager: "WCH USB-SERIAL Ch A").
      2. Uu tien cong co chu kenh 'Ch A'..'Ch H' hoac ten 'CH348/CH9344/Enhanced'
         -> dung kenh A..F (bo G, H).
      3. Neu khong co chu kenh: chi chap nhan khi co >=2 cong WCH cung luc
         (CH348 luon hien nhieu cong); 1 cong WCH don thuong la CH340 -> BO QUA.
    Khong con fallback 'tra ve tat ca cong COM' -> khong liet ke cong khong phai CH348.
    PID khong dung de loc (CH348 cau hinh PID qua EEPROM, khong co dinh).
    """
    if serial is None:
        return []
    import re

    def com_num(dev):
        m = re.search(r"(\d+)", dev or "")
        return int(m.group(1)) if m else 0

    wch = []   # (com_num, device, ch_or_None, named_ch348)
    for p in serial.tools.list_ports.comports():
        vid  = getattr(p, "vid", None)
        desc = (p.description or "")
        hwid = (getattr(p, "hwid", "") or "")
        u = (desc + " " + hwid).upper()
        is_wch = (vid == CH348_VID) or ("1A86" in u) or ("WCH" in u)
        if not is_wch:
            continue
        named = ("CH348" in u) or ("CH9344" in u) or ("ENHANCED" in u)
        m = re.search(r"\bCH(?:ANNEL)?\s*([A-H])\b", u)   # "Ch A".."Ch H" neu co
        wch.append((com_num(p.device), p.device, m.group(1) if m else None, named))

    if not wch:
        return []

    named_ports = [w for w in wch if w[3]]   # driver ten CH348/CH9344/Enhanced
    lettered    = [w for w in wch if w[2]]   # co chu kenh "Ch A".."Ch H"
    if named_ports:
        group = named_ports
    elif lettered:              # cong co chu kenh -> CH348 nhieu cong
        group = lettered
    elif len(wch) >= 2:         # nhieu cong WCH cung luc -> kha nang la CH348
        group = wch
    else:                       # 1 cong WCH don (CH340...) -> khong phai CH348
        return []

    group.sort(key=lambda x: x[0])                # theo so COM tang dan
    have_all = all(ch for _, _, ch, _ in group)   # driver co ghi du chu kenh?
    out = []
    if have_all:
        for _, dev, ch, _ in group:
            if ch in CONSOLE_CHANNELS:            # chi Ch A..F (bo G, H)
                out.append((dev, ch))
        out.sort(key=lambda x: x[1])
    else:                                         # gan kenh A,B,C... theo thu tu COM
        for i, (_, dev, _, _) in enumerate(group):
            out.append((dev, chr(ord("A") + i)))
    return out[:N_PORTS]


def loopback_test(dev, baud, msg, timeout=0.8):
    """Mo cong, gui chuoi 'msg' (bytes), doc lai (loopback TX-RX). True neu nhan dung.

    QUAN TRONG: dat ca write_timeout. Neu cong KHONG dung (vi du cong Bluetooth/
    modem dang giu RTS/CTS), write() khong co write_timeout se BLOCK vo han ->
    treo app. Co write_timeout, write() se bao loi sau 'timeout' thay vi treo.
    """
    try:
        s = serial.Serial(dev, baud, timeout=timeout, write_timeout=timeout)
    except Exception as e:
        return False, f"mo cong loi: {e}"
    try:
        s.reset_input_buffer()
        s.write(msg)
        t0 = time.time()
        buf = b""
        while time.time() - t0 < timeout:
            n = s.in_waiting
            if n:
                buf += s.read(n)
                if msg in buf:
                    break
            else:
                time.sleep(0.02)
        if msg in buf:
            return True, f"OK ({len(msg)} byte)"
        if not buf:
            return False, "khong nhan (chua chap TX-RX?)"
        return False, "sai du lieu"
    except serial.SerialTimeoutException:
        return False, "write timeout (cong khong dung / dang ban?)"
    except Exception as e:
        return False, f"loi doc/ghi: {e}"
    finally:
        try:
            s.close()
        except Exception:
            pass


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("980x720")
        self.configure(bg=MAIN_BG)
        self._set_window_icon()
        self.ports = []          # [(dev, ch)]
        self.rows = {}           # dev -> {"res":..., "val":..., "iid":...}
        self._step_idx = 0       # cong dang chon (dieu huong tung buoc - giong OPMS)
        self._build_ui()
        self.minsize(820, 560)
        self.after(200, self.scan_ports)

    # ---------------- UI ----------------
    def _set_window_icon(self):
        """Icon cua so don gian: o vuong bo goc xanh than + dau ">" console."""
        try:
            from PIL import Image as _Img, ImageDraw as _Dr, ImageTk as _ImgTk
            s = 64
            im = _Img.new("RGBA", (s, s), (0, 0, 0, 0))
            d = _Dr.Draw(im)
            navy = tuple(int(HDR_BG.lstrip("#")[i:i+2], 16) for i in (0, 2, 4))
            d.rounded_rectangle([4, 4, s - 4, s - 4], radius=12, fill=navy)
            # dau ">" (console prompt) mau trang
            d.line([(22, 20), (40, 32), (22, 44)], fill="white", width=6,
                   joint="curve")
            self._app_icon = _ImgTk.PhotoImage(im)
            self.iconphoto(True, self._app_icon)
        except Exception:
            pass

    def _load_hdr_image(self, name, h=HDR_LOGO_H):
        """Nap PNG, scale ve chieu cao h, dan len nen header (dong bo app chinh)."""
        try:
            from PIL import Image as _Img, ImageTk as _ImgTk
            im = _Img.open(resource_path(name)).convert("RGBA")
            w0, h0 = im.size
            nw = max(1, int(w0 * h / h0))
            im = im.resize((nw, h), _Img.LANCZOS)
            bgc = tuple(int(HDR_BG.lstrip("#")[i:i+2], 16) for i in (0, 2, 4))
            bg = _Img.new("RGB", (nw, h), bgc)
            bg.paste(im, mask=im.split()[3])
            return _ImgTk.PhotoImage(bg)
        except Exception:
            return None

    def _card(self, parent, title, subtitle=""):
        frm = tk.Frame(parent, bg=WHITE, highlightbackground=CARD_BD,
                       highlightthickness=1)
        hf = tk.Frame(frm, bg=WHITE, padx=16, pady=10)
        hf.pack(fill="x")
        tk.Label(hf, text=title, bg=WHITE, fg=HDR_BG, font=FONT_CARD).pack(side="left")
        if subtitle:
            tk.Label(hf, text=subtitle, bg=WHITE, fg=HDR_SUB,
                     font=FONT_SM).pack(side="right")
        tk.Frame(frm, bg=CARD_BD, height=1).pack(fill="x")
        body = tk.Frame(frm, bg=WHITE, padx=16, pady=12)
        body.pack(fill="both", expand=True)
        return frm, body

    def _build_ui(self):
        # ===== HEADER (dong bo app chinh: accent bar + anh thiet bi + logo) =====
        hdr = tk.Frame(self, bg=HDR_BG)
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=HDR_ACC, width=6).pack(side="left", fill="y")

        self._dev_img = self._load_hdr_image(OPMS_DEVICE_IMG)
        if self._dev_img:
            tk.Label(hdr, image=self._dev_img, bg=HDR_BG).pack(
                side="left", padx=(12, 10), pady=8)

        tit = tk.Frame(hdr, bg=HDR_BG, padx=4, pady=10)
        tit.pack(side="left")
        tk.Label(tit, text="OPMS CONSOLE BOARD", bg=HDR_BG, fg="white",
                 font=("Segoe UI", 15, "bold")).pack(anchor="w")
        tk.Label(tit, text="CH348 USB  •  6 cổng Console  •  Loopback (chập TX-RX từng cổng)",
                 bg=HDR_BG, fg=HDR_SUB, font=FONT_SM).pack(anchor="w")

        self._fpt_img = self._load_hdr_image(FPT_LOGO_IMG)
        if self._fpt_img:
            tk.Label(hdr, image=self._fpt_img, bg=HDR_BG).pack(
                side="right", padx=(8, 16), pady=8)

        # ===== TOOLBAR (2 hang de khong bi tran/mat chu) =====
        bar = tk.Frame(self, bg=CTRL_BG)
        bar.pack(fill="x")
        # -- hang 1: Serial + Baud + nut --
        bi = tk.Frame(bar, bg=CTRL_BG, padx=12, pady=6)
        bi.pack(fill="x")
        tk.Label(bi, text="Serial:", bg=CTRL_BG, fg=HDR_SUB,
                 font=FONT_SM).pack(side="left")
        self.ent_serial = ttk.Entry(bi, width=18)
        self.ent_serial.pack(side="left", padx=4)
        tk.Label(bi, text="  Baud:", bg=CTRL_BG, fg=HDR_SUB,
                 font=FONT_SM).pack(side="left")
        self.cmb_baud = ttk.Combobox(bi, width=8,
                                     values=[9600, 19200, 38400, 57600, 115200])
        self.cmb_baud.set(DEFAULT_BAUD)
        self.cmb_baud.pack(side="left", padx=4)
        sec_btn(bi, "🔄 Dò cổng CH348", padx=12, pady=4,
                command=self.scan_ports).pack(side="left", padx=(16, 4))
        self.lbl_dev = tk.Label(bi, text="", bg=CTRL_BG, fg=HDR_SUB, font=FONT_SM)
        self.lbl_dev.pack(side="right")
        # -- hang 2: Chuoi test (rong, khong che nut) --
        bi2 = tk.Frame(bar, bg=CTRL_BG, padx=12, pady=6)
        bi2.pack(fill="x")
        tk.Label(bi2, text="Chuỗi test:", bg=CTRL_BG, fg=HDR_SUB,
                 font=FONT_SM).pack(side="left")
        self.ent_msg = ttk.Entry(bi2)
        self.ent_msg.insert(0, DEFAULT_TEST_STR)
        self.ent_msg.pack(side="left", fill="x", expand=True, padx=4)

        # Body
        body = tk.Frame(self, bg=MAIN_BG, padx=12, pady=12)
        body.pack(fill="both", expand=True)

        # Progress + counters
        card, b = self._card(body, "Kết quả loopback 6 cổng Console",
                             "Cổng 1–6 ⟷ Ch A–F  •  PASS = nhận lại đúng")
        card.pack(fill="both", expand=True)
        prog = tk.Frame(b, bg=WHITE)
        prog.pack(fill="x", pady=(0, 8))
        self.lbl_fail = tk.Label(prog, text="FAIL 0", bg=WHITE, fg=FAIL_FG,
                                 font=FONT_B, width=8, anchor="e")
        self.lbl_fail.pack(side="right")
        self.lbl_pass = tk.Label(prog, text="PASS 0", bg=WHITE, fg=PASS_FG,
                                 font=FONT_B, width=8, anchor="e")
        self.lbl_pass.pack(side="right")
        self.lbl_prog = tk.Label(prog, text="0 / 0", bg=WHITE, fg=INK,
                                 font=FONT_B, width=9, anchor="e")
        self.lbl_prog.pack(side="right", padx=(8, 6))
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("CH.Treeview", rowheight=26, font=FONT_SM)
        style.configure("CH.Horizontal.TProgressbar", troughcolor=GREY_LT,
                        background=PRIMARY, bordercolor=CARD_BD,
                        lightcolor=PRIMARY, darkcolor=PRIMARY)
        self.pbar = ttk.Progressbar(prog, style="CH.Horizontal.TProgressbar",
                                    mode="determinate", maximum=100)
        self.pbar.pack(side="left", fill="x", expand=True)

        # --- Điều khiển Run test: nav từng cổng (đồng bộ với app OPMS) ---
        nav = tk.Frame(b, bg=WHITE)
        nav.pack(fill="x", pady=(0, 8))
        flat_btn(nav, "◀ Trước", SEC_BG, SEC_HOV, fg=SEC_TX, padx=12, pady=4,
                 command=self._step_prev).pack(side="left")
        self.lbl_stepnav = tk.Label(nav, text="Cổng 0/0", bg=WHITE, fg=INK,
                                    font=FONT_B, width=11, anchor="center")
        self.lbl_stepnav.pack(side="left", padx=4)
        flat_btn(nav, "Sau ▶", SEC_BG, SEC_HOV, fg=SEC_TX, padx=12, pady=4,
                 command=self._step_nextsel).pack(side="left")
        flat_btn(nav, "▶ Run test", PRIMARY, PRIMARY_HOV,
                 font=("Segoe UI", 13, "bold"), padx=16, pady=5,
                 command=self._step_run_current).pack(side="left", padx=(10, 0))
        sec_btn(nav, "💾 Lưu Excel", padx=12, pady=4,
                command=self._finish_and_prompt).pack(side="right", padx=(4, 0))
        sec_btn(nav, "↺ Reset", padx=12, pady=4,
                command=self.reset_results).pack(side="right")
        tk.Label(b, text="💡 Chập TX–RX cổng đang chọn → “Run test” (test rồi sang cổng kế). "
                         "Hoặc bấm đúp một dòng để test riêng.",
                 bg=WHITE, fg=RUN_FG, font=FONT_SM, anchor="w").pack(fill="x",
                                                                     pady=(0, 4))
        tbox = tk.Frame(b, bg=WHITE)
        tbox.pack(fill="both", expand=True)
        self.tree = ttk.Treeview(tbox, columns=("port", "res", "val"),
                                 show="headings", style="CH.Treeview",
                                 selectmode="browse", height=9, cursor="hand2")
        self.tree.heading("port", text="Cổng (1–6 · Ch A–F)")
        self.tree.heading("res", text="Kết quả")
        self.tree.heading("val", text="Chi tiết / Thao tác")
        self.tree.column("port", width=200, anchor="w")
        self.tree.column("res", width=90, anchor="center")
        self.tree.column("val", width=260, anchor="w")
        self.tree.tag_configure("pass", foreground=PASS_FG)
        self.tree.tag_configure("fail", foreground=FAIL_FG)
        self.tree.tag_configure("running", foreground=RUN_FG)
        self.tree.tag_configure("pending", foreground=DIS_FG)
        vsb = ttk.Scrollbar(tbox, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", self._on_dbl)

        # Terminal
        self._build_terminal(b)

        # Status bar
        self.status = tk.StringVar(value="Sẵn sàng. Bấm 'Dò cổng CH348'.")
        tk.Label(self, textvariable=self.status, bg=GREY_LT, fg=INK,
                 anchor="w", padx=10, pady=4, font=FONT_SM).pack(fill="x")

    def _build_terminal(self, parent):
        tm = tk.Frame(parent, bg=WHITE)
        tm.pack(fill="both", expand=False, pady=(8, 0))
        hdr = tk.Frame(tm, bg=WHITE)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Log", bg=WHITE, fg=HDR_BG,
                 font=FONT_B).pack(side="left")
        tk.Button(hdr, text="Xóa Log", bg=WHITE, fg=SEC_TX,
                  activebackground=GREY_LT, relief="solid", bd=1,
                  font=("Segoe UI", 8), padx=6, pady=1,
                  command=self._clear_log).pack(side="right")
        self.txt = tk.Text(tm, bg=TERM_BG, fg=TERM_FG, insertbackground="white",
                           font=("Consolas", 9), state="disabled", wrap="word",
                           padx=8, pady=6, height=7)
        self.txt.tag_configure("info", foreground=TERM_TX)
        self.txt.tag_configure("ok", foreground=PASS_FG)
        self.txt.tag_configure("err", foreground=FAIL_FG)
        self.txt.tag_configure("warn", foreground=WARN_FG)
        sb = ttk.Scrollbar(tm, command=self.txt.yview)
        self.txt.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.txt.pack(fill="both", expand=True)

    def _log(self, text, tag="info"):
        self.txt.config(state="normal")
        self.txt.insert("end", text, tag)
        self.txt.see("end")
        self.txt.config(state="disabled")

    def _clear_log(self):
        self.txt.config(state="normal")
        self.txt.delete("1.0", "end")
        self.txt.config(state="disabled")

    # ---------------- Logic ----------------
    def scan_ports(self):
        if serial is None:
            messagebox.showerror(APP_TITLE, "Chưa cài pyserial.")
            return
        self.ports = list_ch348_ports()
        self.tree.delete(*self.tree.get_children())
        self.rows = {}
        for idx, (dev, ch) in enumerate(self.ports, 1):
            disp = f"Cổng {idx}  ·  Ch {ch}  ·  {dev}"
            iid = self.tree.insert("", "end",
                                   values=(disp, "-", "▶ bấm đúp để test"),
                                   tags=("pending",))
            self.rows[dev] = {"iid": iid, "res": "", "val": "", "disp": disp}
        n = len(self.ports)
        self.lbl_dev.config(text=f"CH348: {n}/6 cổng")
        if n == 0:
            self.status.set("Không thấy cổng CH348 (Ch A–F). Kiểm tra cáp/driver WCH.")
            self._log("[Không phát hiện cổng CH348]\n", "warn")
        else:
            self.status.set(f"Phát hiện {n} cổng CH348.")
            self._log(f"[Phát hiện {n} cổng CH348: "
                      f"{', '.join(d for d, _ in self.ports)}]\n", "info")
            if n != N_PORTS:
                self._log(f"[Lưu ý: kỳ vọng {N_PORTS} cổng, thấy {n}]\n", "warn")
        self._step_select(0)            # con trỏ điều hướng về cổng 1
        self._refresh_progress()

    def _on_dbl(self, ev):
        iid = self.tree.identify_row(ev.y)
        for dev, r in self.rows.items():
            if r["iid"] == iid:
                self.test_one(dev)
                break

    # ---------------- Điều hướng từng cổng (đồng bộ app OPMS) ----------------
    def _step_select(self, idx):
        if not self.ports:
            self._step_idx = 0
            self.lbl_stepnav.config(text="Cổng 0/0")
            return
        idx = max(0, min(idx, len(self.ports) - 1))
        self._step_idx = idx
        iid = self.rows.get(self.ports[idx][0], {}).get("iid")
        if iid:
            self.tree.selection_set(iid)
            self.tree.see(iid)
        self.lbl_stepnav.config(text=f"Cổng {idx + 1}/{len(self.ports)}")

    def _step_prev(self):
        if self.ports:
            self._step_select(self._step_idx - 1)

    def _step_nextsel(self):
        if self.ports:
            self._step_select(self._step_idx + 1)

    def _step_run_current(self):
        if not self.ports:
            self.scan_ports()
            if not self.ports:
                return
        idx = min(self._step_idx, len(self.ports) - 1)
        self.test_one(self.ports[idx][0])        # test cổng đang chọn
        if idx < len(self.ports) - 1:            # tự sang cổng kế
            self._step_select(idx + 1)

    def _test_bytes(self):
        """Chuoi test nguoi dung nhap -> bytes (them CRLF). Mac dinh la chuoi dai."""
        s = self.ent_msg.get() if hasattr(self, "ent_msg") else DEFAULT_TEST_STR
        if not s.strip():
            s = DEFAULT_TEST_STR
        return (s + "\r\n").encode("utf-8", "replace")

    def test_one(self, dev, baud=None):
        baud = baud or int(self.cmb_baud.get())
        r = self.rows[dev]
        self.tree.item(r["iid"], values=(r["disp"], "...", "đang test"),
                       tags=("running",))
        self.update()
        msg = self._test_bytes()
        self._log(f"> {dev}: gửi {len(msg)} byte @ {baud}\n", "info")
        ok, detail = loopback_test(dev, baud, msg)
        r["res"] = "PASS" if ok else "FAIL"
        r["val"] = detail
        self.tree.item(r["iid"], values=(r["disp"], r["res"], detail),
                       tags=("pass" if ok else "fail",))
        self._log(f"  {dev}: {r['res']} ({detail})\n", "ok" if ok else "err")
        self._refresh_progress()
        return ok

    def run_all(self):
        """Autotest co huong dan: lan luot tung cong -> nhac nguoi dung CHAP TX-RX,
        kiem tra; PASS -> tu dong sang cong sau; FAIL -> Thu lai / Bo qua."""
        if not self.ports:
            self.scan_ports()
        if not self.ports:
            messagebox.showinfo(APP_TITLE,
                                "Chưa thấy cổng nào.\nBấm 'Dò cổng CH348'.")
            return
        if not self.ent_serial.get().strip():
            if not messagebox.askyesno(APP_TITLE,
                                       "Chưa nhập Serial thiết bị.\nVẫn tiếp tục?"):
                return
        self.reset_results()
        n = len(self.ports)
        stopped = False
        for idx, (dev, ch) in enumerate(self.ports, 1):
            disp = f"Cổng {idx}  ·  Ch {ch}  ·  {dev}"
            while True:
                ans = messagebox.askyesnocancel(
                    APP_TITLE + f" — Cổng {idx}/{n}",
                    f"{disp}\n(Cổng {idx} trên thiết bị = kênh Ch {ch})\n\n"
                    "→ CHẬP (nối tắt) chân TX–RX của cổng này, rồi chọn:\n\n"
                    "   • Yes (Có)    = Kiểm tra cổng\n"
                    "   • No (Không)  = Bỏ qua cổng này\n"
                    "   • Cancel      = Dừng autotest")
                if ans is None:                  # Dừng toàn bộ
                    stopped = True
                    break
                if ans is False:                 # Bỏ qua cổng này
                    r = self.rows[dev]
                    r["res"] = "SKIP"
                    r["val"] = "bỏ qua"
                    self.tree.item(r["iid"], values=(r["disp"], "SKIP", "bỏ qua"),
                                   tags=("pending",))
                    self._refresh_progress()
                    break
                # Yes -> kiểm tra cổng
                if self.test_one(dev):           # PASS -> tự động sang cổng sau
                    break
                # FAIL -> Thử lại / Bỏ qua
                if not messagebox.askretrycancel(
                        APP_TITLE,
                        f"Cổng {idx} ({dev}) CHƯA ĐẠT.\n\n"
                        "Retry = thử lại (chập lại rồi đo)\n"
                        "Cancel = bỏ qua cổng này (giữ FAIL)"):
                    break                        # giữ FAIL, sang cổng sau
            if stopped:
                self.status.set("Đã dừng autotest.")
                break
        self._finish_and_prompt()

    def _finish_and_prompt(self):
        """Tong ket + hoi luu Excel."""
        total = len(self.ports)
        npass = sum(1 for r in self.rows.values() if r["res"] == "PASS")
        nfail = sum(1 for r in self.rows.values() if r["res"] == "FAIL")
        nskip = sum(1 for r in self.rows.values() if r["res"] == "SKIP")
        final = "PASS" if (npass == total and total > 0) else "FAIL"
        self.status.set(f"Xong. PASS {npass} / FAIL {nfail} / SKIP {nskip}.")
        serial_no = self.ent_serial.get().strip() or "(chưa nhập)"
        if messagebox.askyesno(
                APP_TITLE + " — " + final,
                f"{'✓' if final == 'PASS' else '✗'} KẾT QUẢ: {final}\n\n"
                f"Serial: {serial_no}\n"
                f"PASS {npass} / FAIL {nfail} / SKIP {nskip} / Tổng {total}\n\n"
                f"Lưu kết quả vào {REPORT_NAME}?"):
            self._save_report(final, npass, nfail)
            self.ent_serial.delete(0, "end")
            self.ent_serial.focus_set()

    def reset_results(self):
        for dev, r in self.rows.items():
            r["res"] = ""
            r["val"] = ""
            self.tree.item(r["iid"], values=(r["disp"], "-", "▶ bấm đúp để test"),
                           tags=("pending",))
        self.status.set("Đã reset kết quả.")
        self._refresh_progress()

    def _refresh_progress(self):
        total = len(self.ports)
        npass = sum(1 for r in self.rows.values() if r["res"] == "PASS")
        nfail = sum(1 for r in self.rows.values() if r["res"] == "FAIL")
        done = sum(1 for r in self.rows.values() if r["res"])
        self.lbl_prog.config(text=f"{done} / {total}")
        self.lbl_pass.config(text=f"PASS {npass}")
        self.lbl_fail.config(text=f"FAIL {nfail}")
        self.pbar["value"] = (done / total * 100) if total else 0

    # ---------------- Report (1 thiet bi = 1 hang, append) ----------------
    def _save_report(self, final, npass, nfail):
        if openpyxl is None:
            messagebox.showerror(APP_TITLE, "Chưa cài openpyxl.")
            return
        serial_no = self.ent_serial.get().strip() or \
            ("NA-" + datetime.datetime.now().strftime("%H%M%S"))
        header = ["STT", "Serial", "Thoi gian", "Baud",
                  *[f"P{i+1}" for i in range(N_PORTS)], "PASS/Tong", "Ket luan"]
        # ket qua tung cong theo thu tu P1..P8 (du -> ghi '-')
        cells = []
        for i in range(N_PORTS):
            if i < len(self.ports):
                dev = self.ports[i][0]
                r = self.rows.get(dev, {})
                cells.append(f"{dev}:{r.get('res') or '-'}")
            else:
                cells.append("-")
        ket_luan = "OK" if final == "PASS" else "LOI"
        path = report_path()
        for attempt in range(5):     # retry neu file dang mo boi cua so khac
            try:
                if os.path.exists(path):
                    wb = openpyxl.load_workbook(path)
                    ws = wb.active
                    stt = ws.max_row
                else:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Console CH348"
                    ws.append(header)
                    stt = 1
                row = [stt, serial_no,
                       datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                       self.cmb_baud.get(), *cells,
                       f"{npass}/{len(self.ports)}", ket_luan]
                ws.append(row)
                wb.save(path)
                self.status.set(f"Đã ghi report ({serial_no}): {path}")
                self._log(f"[Đã ghi report: {serial_no} -> {ket_luan}]\n", "ok")
                return
            except PermissionError:
                time.sleep(0.4)      # file dang bi khoa -> thu lai
        messagebox.showerror(APP_TITLE,
                             f"Không ghi được {REPORT_NAME} (file đang mở?).\n"
                             "Đóng file Excel rồi thử lại.")


def main():
    if serial is None:
        root = tk.Tk(); root.withdraw()
        messagebox.showerror(APP_TITLE, "Chưa cài pyserial.\npip install pyserial")
        return
    App().mainloop()


if __name__ == "__main__":
    main()