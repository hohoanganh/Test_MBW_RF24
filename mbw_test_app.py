# -*- coding: utf-8 -*-
"""
mbw_test_app.py - App test MBW RF24 RS485 2.0 (MBW-RF-00-24)
==============================================================
GUI Python (tkinter) test TỪNG BOARD RIÊNG LẺ qua 1 cổng COM. Vì 2 board khi
truyền/nhận thật sẽ được cắm vào 2 MÁY TÍNH KHÁC NHAU (mỗi máy chạy 1 instance
app riêng, mỗi app chỉ thấy board của máy mình), app KHÔNG điều khiển đồng
thời 2 board như thiết kế trước — mỗi cửa sổ app chỉ kết nối 1 board.

Chức năng mặc định của board là TỰ ĐỘNG forward RS485 <-> Wireless ngay khi
cấp nguồn (bridge luôn bật, không cần lệnh kích hoạt). App chỉ CẮM VÀO ĐỂ QUAN
SÁT: mỗi khi board relay 1 khung dữ liệu (2 chiều), firmware in ra console 1
dòng "FWD RS485->RF: ..." / "FWD RF->RS485: ...", app đọc và hiển thị trực
quan lên màn hình (tab "Giám sát Forward") — đây là bằng chứng trực tiếp thiết
bị đang hoạt động đúng chức năng, xem được ngay trên từng máy mà không cần
đồng bộ giữa 2 máy tính.

Chức năng:
  - Kết nối 1 cổng COM, xác thực ID ("MBW_RF24_RS485").
  - Giám sát Forward: bảng log các khung đã relay (giờ, hướng, số byte,
    preview hex), bật/tắt bridge, bật/tắt log forward, xem thống kê, gửi thử
    RS485 để tự kiểm tra forward khi chưa có Modbus Master/Slave thật.
  - Nút "🔌 Modbus Poll Test (RS485 thật)" (tab Giám sát Forward) mở
    modbus_poll_app.py trong 1 tiến trình/cửa sổ riêng - đây là công cụ
    Modbus RTU Master THẬT (đọc/ghi thanh ghi FC3/FC6, log Excel, biểu đồ)
    kết nối qua cổng RS485 VẬT LÝ của board (khác cổng console CLI) - phép
    test thực tế nhất cho chức năng bridge: 1 máy chạy Modbus Master qua
    RS485 của board A, đầu kia là board B nối Modbus Slave thật.
  - Test tự động: id/dip/flash/rtc/rsl (loopback)/rf id trên board đang nối.
  - Terminal: gõ lệnh CLI trực tiếp.
  - Xuất báo cáo nối tiếp vào mbw_test_report.xlsx (hoặc .csv nếu thiếu openpyxl).

Cài đặt:  pip install pyserial openpyxl matplotlib
Chạy:     python mbw_test_app.py
Đóng gói: build_mbw_exe.bat -> dist/MBW_RF24_Test.exe
"""

import os
import re
import sys
import json
import time
import threading
import subprocess
import collections
from datetime import datetime

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    import serial
    import serial.tools.list_ports as list_ports
except ImportError:
    serial = None
    list_ports = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

from mbw_theme import (
    MAIN_BG, WHITE, CARD_BD, HDR_BG, HDR_SUB, PRIMARY, PRIMARY_HOV,
    SEC_BG, SEC_HOV, SEC_TX, INK, PASS_FG, FAIL_FG, RUN_FG, WARN_FG, DIS_FG,
    TERM_BG, TERM_FG, TERM_TX, FONT, FONT_B, FONT_SM, FONT_CARD, FONT_MONO,
    resource_path, flat_btn, sec_btn,
)

APP_TITLE = "MBW RF24 RS485 2.0 - Test App"
CONFIG_PATH = resource_path("app_config.json")
REPORT_PATH = os.path.join(
    os.path.dirname(os.path.abspath(sys.argv[0] if not hasattr(sys, "_MEIPASS") else sys.executable)),
    "mbw_test_report.xlsx")

FWD_RE = re.compile(r"^FWD (RS485->RF|RF->RS485): (\d+) bytes:(.*)$")

# ----- Heartbeat / link health (giong "link quality" cua bo telemetry) -----
# 2 dong tuc thoi khi link doi trang thai:
LINK_UP_RE = re.compile(r"^RF LINK: UP \(peer=(\d+)\)")
LINK_DOWN_RE = re.compile(r"^RF LINK: DOWN \(peer=(\d+), (\d+)ms")
# dong "rf stat" day du (doc dinh ky) chua ca so lieu dinh luong.
# 2026-07-06: firmware in "LAST_DEV_ID=" (truoc day la "PEER=") va co them
# "REPEATER=ON|OFF RELAY=<n>" sau HB_RX (khong anh huong regex - khong anchor
# cuoi dong). Chap nhan CA 2 ten truong de tuong thich firmware cu/moi.
RF_STAT_RE = re.compile(
    r"RF_LINK=(UP|DOWN)\s+(?:PEER|LAST_DEV_ID)=(\d+)\s+AGE_MS=(\d+)\s+LOSS_PROMILLE=(\d+)\s+"
    r"REDUND=(\d+)\s+HB_TX=(\d+)\s+HB_RX=(\d+)")
# dong dem goi RF ("rf stat" in truoc dong RF_LINK=):
RF_CNT_RE = re.compile(
    r"RF_TX=(\d+)\s+RF_RX_OK=(\d+)\s+RF_RX_DUP=(\d+)\s+RF_RX_CRCERR=(\d+)\s+"
    r"RF_RX_FRAGDROP=(\d+)")
# dong "RF_DEVS:" (breakdown tung dev_id trong "rf stat") - biet CON NAO ROT /
# CON NAO LINK KEM. Moi token dang "<id>:<UP|DOWN>(<age>s,<loss%o | ->)".
RF_DEVS_RE = re.compile(r"^RF_DEVS:")
DEV_TOKEN_RE = re.compile(r"(\d+):(UP|DOWN)\((\d+)s,(-|\d+)\)")
# dong log su kien tu Flash: "LOG: <idx> dd/mm hh:mm:ss dev=<id> UP|DOWN loss=<n%|->"
LOG_LINE_RE = re.compile(
    r"^LOG:\s+(\d+)\s+(\d\d/\d\d)\s+(\d\d:\d\d:\d\d)\s+dev=(\d+)\s+(UP|DOWN)\s+loss=(\S+)")
# dong "bridge stat" (dem khung forward + drop queue lien-task):
BR_CNT_RE = re.compile(
    r"BRIDGE=(?:ON|OFF)\s+LOG=(?:ON|OFF)\s+RS485_TO_RF=(\d+)\s+RF_TO_RS485=(\d+)\s+"
    r"DROP_RS485_TO_RF=(\d+)\s+DROP_RF_TO_RS485=(\d+)")


def modbus_crc16(data):
    """CRC16 Modbus RTU (poly 0xA001, init 0xFFFF). Tinh tren CA KHUNG (gom
    2 byte CRC cuoi) -> khung dung tra ve 0."""
    crc = 0xFFFF
    for b in data:
        crc ^= b
        for _ in range(8):
            crc = (crc >> 1) ^ 0xA001 if crc & 1 else crc >> 1
    return crc


def load_config():
    default = {
        "device_id": "MBW_RF24_RS485",
        "console_baud": 115200,
        "test_groups": [],
    }
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        default.update(cfg)
        return default
    except Exception:
        return default


# =========================================================
# PortLink: 1 ket noi serial + doc nen (background thread) + buffer dong
# =========================================================
class PortLink:
    def __init__(self):
        self.ser = None
        self.port = None
        self.baud = 115200
        self.lines = collections.deque(maxlen=2000)
        self.lock = threading.Lock()
        self.running = False
        self.thread = None
        self.on_line = None  # callback(text) - GOI TU THREAD DOC, phai tu after() ben GUI

    def is_open(self):
        return self.ser is not None and self.ser.is_open

    def connect(self, port, baud=115200):
        if serial is None:
            raise RuntimeError("Chua cai pyserial: pip install pyserial")
        self.disconnect()
        self.ser = serial.Serial(port, baud, timeout=0.1)
        self.port = port
        self.baud = baud
        self.running = True
        self.thread = threading.Thread(target=self._reader, daemon=True)
        self.thread.start()

    def disconnect(self):
        self.running = False
        if self.ser is not None:
            try:
                self.ser.close()
            except Exception:
                pass
        self.ser = None

    def _reader(self):
        buf = b""
        while self.running and self.ser is not None:
            try:
                data = self.ser.read(256)
            except Exception:
                break
            if data:
                buf += data
                while b"\n" in buf:
                    raw, buf = buf.split(b"\n", 1)
                    text = raw.decode(errors="replace").rstrip("\r")
                    with self.lock:
                        self.lines.append(text)
                    if self.on_line:
                        try:
                            self.on_line(text)
                        except Exception:
                            pass

    def send(self, cmd):
        if self.is_open():
            try:
                self.ser.write((cmd + "\n").encode())
            except Exception:
                pass

    def snapshot_len(self):
        with self.lock:
            return len(self.lines)

    def lines_since(self, idx):
        with self.lock:
            return list(self.lines)[idx:]

    def wait_for(self, substr, timeout_s=2.0):
        start_idx = self.snapshot_len()
        t0 = time.time()
        while time.time() - t0 < timeout_s:
            for ln in self.lines_since(start_idx):
                if substr in ln:
                    return ln
            time.sleep(0.02)
        return None

    def cmd_and_wait(self, cmd, expect_substr, timeout_s=2.0):
        self.send(cmd)
        if expect_substr is None:
            time.sleep(min(timeout_s, 0.3))
            return True, ""
        found = self.wait_for(expect_substr, timeout_s)
        return (found is not None), (found or "")


# =========================================================
# APP
# =========================================================
class MbwTestApp:
    def __init__(self, root):
        self.root = root
        self.cfg = load_config()
        self.link = PortLink()
        self.fwd_count = {"RS485->RF": 0, "RF->RS485": 0}

        root.title(APP_TITLE)
        root.configure(bg=MAIN_BG)
        root.geometry("1000x760")

        self._build_style()
        self._build_header()
        self._build_conn_bar()
        self._build_tabs()

        self.link.on_line = self._on_line

        self.test_running = False

    # ---------- STYLE (dong bo voi modbus_poll_app.py) ----------
    def _build_style(self):
        """Style ttk dung chung voi modbus_poll_app.py: cung bang mau/font
        Industrial Flat (mbw_theme.py) cho Notebook (tab)/Treeview, kem fix
        loi tab "bi lech/nhay" khi bam vao (theme clam mac dinh ve them 1
        'focus ring' rieng ben trong tab, gay lech tam thoi luc bam)."""
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure(".", background=MAIN_BG, foreground=INK, font=FONT)
        style.configure("TFrame", background=MAIN_BG)
        style.configure("TLabel", background=MAIN_BG, foreground=INK, font=FONT)

        style.configure("TNotebook", background=MAIN_BG, borderwidth=0, tabmargins=(2, 4, 2, 0))
        style.configure("TNotebook.Tab", background=SEC_BG, foreground=SEC_TX,
                        font=FONT, padding=(14, 6), borderwidth=0)
        style.map("TNotebook.Tab",
                  background=[("selected", PRIMARY)],
                  foreground=[("selected", WHITE)],
                  expand=[("selected", (1, 1, 1, 0))])

        # Bo phan tu "Notebook.focus" trong layout mac dinh cua theme "clam" -
        # day la nguyen nhan gay tab bi "nhay"/lech vi tri ngay luc bam vao.
        style.layout("TNotebook.Tab", [
            ("Notebook.tab", {
                "sticky": "nswe",
                "children": [
                    ("Notebook.padding", {
                        "side": "top",
                        "sticky": "nswe",
                        "children": [
                            ("Notebook.label", {"side": "top", "sticky": ""})
                        ],
                    })
                ],
            })
        ])

        style.configure("Treeview", rowheight=24, font=FONT, background=WHITE,
                        fieldbackground=WHITE, foreground=INK, borderwidth=0)
        style.configure("Treeview.Heading", font=FONT_CARD, background=HDR_BG,
                        foreground=WHITE, relief="flat")
        style.map("Treeview.Heading", background=[("active", HDR_BG)])
        style.map("Treeview", background=[("selected", PRIMARY)],
                  foreground=[("selected", WHITE)])

    # ---------- HEADER ----------
    def _build_header(self):
        hdr = tk.Frame(self.root, bg=HDR_BG, height=56)
        hdr.pack(fill="x", side="top")
        tk.Label(hdr, text="MBW RF24 RS485 2.0 - Test App", bg=HDR_BG, fg=WHITE,
                  font=("Segoe UI", 14, "bold")).pack(side="left", padx=16, pady=10)
        tk.Label(hdr, text="MBW-RF-00-24  |  1 board / 1 máy tính  |  Bridge mặc định luôn BẬT",
                  bg=HDR_BG, fg=HDR_SUB, font=FONT_SM).pack(side="left", padx=4)

    # ---------- CONNECTION BAR (1 board) ----------
    def _build_conn_bar(self):
        card = tk.Frame(self.root, bg=WHITE, highlightbackground=CARD_BD, highlightthickness=1)
        card.pack(fill="x", padx=12, pady=8)

        tk.Label(card, text="Kết nối board", bg=WHITE, fg=INK, font=FONT_CARD).grid(
            row=0, column=0, columnspan=5, sticky="w", padx=10, pady=(8, 2))

        self.port_var = tk.StringVar()
        # rong 42 ky tu + kem MO TA thiet bi (USB-SERIAL CH340...) de chon
        # dung cong khi may co nhieu COM
        combo = ttk.Combobox(card, textvariable=self.port_var, width=30, state="readonly")
        combo.grid(row=1, column=0, padx=(10, 4), pady=(0, 10))
        self.combo = combo

        def refresh():
            items = []
            if list_ports:
                for p in list_ports.comports():
                    desc = (p.description or "").strip()
                    items.append("%s - %s" % (p.device, desc) if desc else p.device)
            combo["values"] = items
            if items and not self.port_var.get():
                self.port_var.set(items[0])

        refresh()
        sec_btn(card, "↻", command=refresh, width=3).grid(row=1, column=1, pady=(0, 10))

        # Nut Connect dat TRUOC label trang thai + label co width CO DINH:
        # truoc day label doi do dai ("Dang xac thuc..." -> "Da ket noi...")
        # lam nut Connect bi xo day qua lai
        self.conn_btn = flat_btn(card, "Connect", PRIMARY, PRIMARY_HOV, command=self._toggle_conn)
        self.conn_btn.grid(row=1, column=2, padx=10, pady=(0, 10))

        self.status_lbl = tk.Label(card, text="● Chưa kết nối", bg=WHITE, fg=DIS_FG,
                                    font=FONT_SM, width=30, anchor="w")
        self.status_lbl.grid(row=1, column=3, padx=8, pady=(0, 10), sticky="w")

        card.grid_columnconfigure(4, weight=1)
        # Nut mo cong cu Modbus Master that - dat o thanh ket noi (truoc day
        # o tab Forward bi cac o so day ra ngoai man hinh)
        flat_btn(card, "🔌 Modbus Poll Test (RS485 thật)", PRIMARY, PRIMARY_HOV,
                 command=self._open_modbus_poll).grid(row=1, column=5, padx=10,
                                                      pady=(0, 10), sticky="e")

    def _toggle_conn(self):
        if self.link.is_open():
            self.link.disconnect()
            self.conn_btn.config(text="Connect", bg=PRIMARY, activebackground=PRIMARY_HOV)
            self.status_lbl.config(text="● Chưa kết nối", fg=DIS_FG)
            if hasattr(self, "lbl_rf_link"):
                self.lbl_rf_link.config(text="● Chưa rõ", fg=DIS_FG)
                self.lbl_rf_link_detail.config(text="")
            return
        # combo hien "COM9 - USB-SERIAL CH340" -> lay phan ten cong truoc " - "
        port = self.port_var.get().split(" - ")[0].strip()
        if not port:
            messagebox.showwarning(APP_TITLE, "Chưa chọn cổng COM")
            return
        try:
            self.link.connect(port, self.cfg.get("console_baud", 115200))
        except Exception as e:
            messagebox.showerror(APP_TITLE, "Không mở được cổng: %s" % e)
            return
        self.status_lbl.config(text="● Đang kiểm tra ID...", fg=WARN_FG)
        self.root.after(300, self._verify_id)

    def _verify_id(self):
        """Xac thuc ID o THREAD NEN, thu 3 lan x 2.5s. Truoc day chi cho 1.5s
        tren GUI thread: khi console dang ngap log FWD (Modbus chay), phan hoi
        'id' ve tre -> bao 'Sai thiet bi' OAN du cong da mo va app van chay
        binh thuong."""
        self.status_lbl.config(text="● Đang xác thực ID...", fg=RUN_FG)
        self.conn_btn.config(text="Disconnect", bg=SEC_BG, fg=SEC_TX, activebackground=SEC_HOV)
        threading.Thread(target=self._verify_id_worker, daemon=True).start()

    def _verify_id_worker(self):
        dev_id = self.cfg.get("device_id", "MBW_RF24_RS485")
        line = None
        for _ in range(3):
            if not self.link.is_open():
                return
            self.link.send("id")
            line = self.link.wait_for(dev_id, 2.5)
            if line:
                break

        def apply():
            if not self.link.is_open():
                return
            if line:
                self.status_lbl.config(text="● Đã kết nối (%s)" % self.link.port, fg=PASS_FG)
            else:
                # cong da mo, du lieu van chay - chi la chua doc duoc chuoi ID
                # (console ban hoac cam nham thiet bi khac)
                self.status_lbl.config(
                    text="● %s - CHƯA XÁC NHẬN ID" % self.link.port,
                    fg=WARN_FG)
            self._refresh_bridge_stat()
            self._schedule_rf_link_poll()
            self._read_net_id()  # dien san NET_ID hien tai vao o cau hinh
            self._read_rtc()     # hien gio RTC hien tai cua board

        self.root.after(0, apply)

    # ---------- LINE ROUTER (dung chung cho moi tab dang lang nghe) ----------
    def _on_line(self, text):
        m = FWD_RE.match(text)
        m_up = LINK_UP_RE.match(text)
        m_down = LINK_DOWN_RE.match(text)
        m_devs = RF_DEVS_RE.match(text)

        def apply():
            if hasattr(self, "raw_log"):
                self.raw_log.insert("end", text + "\n")
                self.raw_log.see("end")
            if hasattr(self, "term_log"):
                self.term_log.insert("end", text + "\n")
                self.term_log.see("end")
            if m:
                self._add_fwd_row(m.group(1), int(m.group(2)), m.group(3).strip())
            # Cap nhat NGAY khi board tu bao doi trang thai link (khong doi
            # vong poll "rf stat" tiep theo) - giong canh bao "mat lien lac"
            # tuc thi cua GCS/telemetry.
            if m_up:
                self._apply_link_ui(True, m_up.group(1))
            elif m_down:
                self._apply_link_ui(False, m_down.group(1))
            if m_devs:
                self._seen_rf_devs = True  # firmware moi co in RF_DEVS
                self._update_dev_panel(text)

        self.root.after(0, apply)

    def _update_dev_panel(self, line):
        """Cap nhat 16 O CO DINH (dev_id 1..16) tu dong RF_DEVS: chi doi mau +
        text moi o, KHONG tao/xoa widget (nen luon hien, khong bi trong).
        Do = DOWN / loss >10%, cam = kha 2-10%, xanh = tot <2%, xam = chua nghe thay."""
        if not hasattr(self, "dev_cells"):
            return
        toks = DEV_TOKEN_RE.findall(line)
        seen = {int(d): (u, a, l) for d, u, a, l in toks if int(d) <= self.DEV_MAX}
        # Tieu de: dem so con k/n + so con mat link
        if hasattr(self, "lbl_dev_hdr"):
            n = len(toks)
            n_down = sum(1 for _d, u, _a, _l in toks if u != "UP")
            extra = (" — %d DOWN" % n_down) if n_down else ""
            self.lbl_dev_hdr.config(
                text="Thiết bị kết nối: %d con%s  (số %% = tỷ lệ mất gói, càng nhỏ càng tốt)"
                     % (n, extra))
        for dev_id, (cell, lid, lsub) in self.dev_cells.items():
            if dev_id not in seen:
                bg, fg, sub = self.DEV_NEUTRAL, DIS_FG, "—"  # chua nghe thay
            else:
                updown, age, loss = seen[dev_id]
                if updown != "UP":
                    bg, fg, sub = FAIL_FG, WHITE, "DOWN"
                elif loss == "-":
                    bg, fg, sub = "#E9EDF1", INK, "-- %"  # chua du du lieu loss
                else:
                    lp = int(loss) / 10.0  # %o -> %
                    if lp < 2:
                        bg, fg = PASS_FG, WHITE
                    elif lp <= 10:
                        bg, fg = WARN_FG, INK
                    else:
                        bg, fg = FAIL_FG, WHITE
                    sub = "%.1f%%" % lp
            cell.config(bg=bg)
            lid.config(bg=bg, fg=fg)
            lsub.config(bg=bg, fg=fg, text=sub)

    # ---------- TABS ----------
    def _build_tabs(self):
        nb = ttk.Notebook(self.root)
        nb.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        self.tab_fwd = tk.Frame(nb, bg=MAIN_BG)
        self.tab_test = tk.Frame(nb, bg=MAIN_BG)
        self.tab_stress = tk.Frame(nb, bg=MAIN_BG)
        self.tab_term = tk.Frame(nb, bg=MAIN_BG)
        # Thu tu tab theo tan suat su dung khi LAP DAT/van hanh: giam sat +
        # danh gia RF truoc, Factory Test (chi dung o xuong) de sau cung.
        nb.add(self.tab_fwd, text="  Giám sát Forward  ")
        nb.add(self.tab_stress, text="  Stress Test  ")
        nb.add(self.tab_term, text="  Terminal  ")
        nb.add(self.tab_test, text="  Factory Test  ")

        self._build_forward_tab(self.tab_fwd)
        self._build_test_tab(self.tab_test)
        self._build_stress_tab(self.tab_stress)
        self._build_terminal_tab(self.tab_term)

    # ---------- TAB: GIAM SAT FORWARD ----------
    def _build_forward_tab(self, parent):
        top = tk.Frame(parent, bg=MAIN_BG)
        top.pack(fill="x", pady=(6, 4))

        self.btn_bridge = flat_btn(top, "Bridge: ?", PRIMARY, PRIMARY_HOV,
                                    command=lambda: self._bridge_toggle())
        self.btn_bridge.pack(side="left")

        self.btn_bridge_log = sec_btn(top, "Log: ?", command=lambda: self._bridge_log_toggle())
        self.btn_bridge_log.pack(side="left", padx=6)

        sec_btn(top, "Đọc thống kê (bridge stat)", command=self._refresh_bridge_stat).pack(
            side="left", padx=6)

        # width co dinh + anchor trai: so thay doi khong lam xo day ca hang
        self.lbl_stat = tk.Label(top, text="RS485→RF: 0    RF→RS485: 0", bg=MAIN_BG, fg=INK,
                                  font=FONT_B, width=32, anchor="w")
        self.lbl_stat.pack(side="left", padx=(16, 0))

        # Hieu so 2 chieu quy ra % LOI GIAO DICH (request khong co response):
        # tong tu luc mo app + rieng 60 giay gan nhat (phan anh "hien tai").
        # Nguong danh gia: <=1%% xanh (TOT) / 1-2%% vang (CHAP NHAN) / >2%% do (KEM).
        self.lbl_err = tk.Label(top, text="LỖI: --", bg=MAIN_BG, fg=DIS_FG,
                                 font=("Segoe UI", 12, "bold"), width=34, anchor="w")
        self.lbl_err.pack(side="left", padx=(4, 0))
        self._fwd_hist = collections.deque(maxlen=1200)  # (t, dem RS485->RF, dem RF->RS485)

        # --- Cau hinh board: NET ID (chung 1 mang, luu Flash - giu qua mat nguon) ---
        cfgrow = tk.Frame(parent, bg=MAIN_BG)
        cfgrow.pack(fill="x", pady=(0, 4))
        tk.Label(cfgrow, text="Net ID (chung cho cả mạng):", bg=MAIN_BG, fg=SEC_TX,
                 font=FONT_SM).pack(side="left")
        self.net_id_var = tk.StringVar(value="")
        tk.Spinbox(cfgrow, from_=0, to=63, width=4, textvariable=self.net_id_var,
                   font=FONT_SM).pack(side="left", padx=(4, 4))
        sec_btn(cfgrow, "Đọc ID", command=self._read_net_id).pack(side="left", padx=(0, 4))
        sec_btn(cfgrow, "Set & Lưu Flash", command=self._set_net_id).pack(side="left")
        self.lbl_net_id = tk.Label(cfgrow, text="", bg=MAIN_BG, fg=SEC_TX, font=FONT_SM)
        self.lbl_net_id.pack(side="left", padx=(8, 0))
        # Log su kien mat/khoi phuc link luu Flash
        sec_btn(cfgrow, "Xóa log", command=self._clear_flash_log).pack(side="right")
        sec_btn(cfgrow, "📋 Đọc log sự kiện (Flash)",
                command=self._read_flash_log).pack(side="right", padx=(0, 6))

        # --- Dong ho RTC (moc thoi gian cho log; can lap pin CR1220 de giu gio) ---
        cfgrow2 = tk.Frame(parent, bg=MAIN_BG)
        cfgrow2.pack(fill="x", pady=(0, 4))
        tk.Label(cfgrow2, text="Đồng hồ RTC (mốc thời gian cho log):", bg=MAIN_BG,
                 fg=SEC_TX, font=FONT_SM).pack(side="left")
        sec_btn(cfgrow2, "⏱ Đồng bộ giờ máy tính",
                command=self._sync_rtc_pc).pack(side="left", padx=(6, 4))
        sec_btn(cfgrow2, "Đọc giờ", command=self._read_rtc).pack(side="left")
        self.lbl_rtc = tk.Label(cfgrow2, text="", bg=MAIN_BG, fg=SEC_TX, font=FONT_SM)
        self.lbl_rtc.pack(side="left", padx=(8, 0))

        # (nut "Modbus Poll Test" da chuyen len thanh "Ket noi board" - truoc
        # day dat o day bi cac o so width co dinh day ra ngoai man hinh)

        # --- PANEL CHAT LUONG RF (dung khi LAP DAT): 5 o KPI lon, mau theo
        # nguong de tho lap dat nhin 1 giay biet dat/khong. LUU Y: nRF24L01
        # KHONG co RSSI (chi co bit RPD >-64dBm, gan nhu vo dung de danh gia)
        # -> danh gia bang TI LE MAT heartbeat + MUC DU PHONG tu dong (REDUND)
        # - trung thuc hon RSSI vi do truc tiep "gui co toi noi khong". ---
        qcard = tk.Frame(parent, bg=WHITE, highlightbackground=CARD_BD, highlightthickness=1)
        qcard.pack(fill="x", pady=(0, 6))
        qrow = tk.Frame(qcard, bg=WHITE)
        qrow.pack(fill="x", padx=10, pady=(6, 8))

        FONT_KPI = ("Segoe UI", 13, "bold")

        def _kpi(col, caption, w):
            # width CO DINH (don vi ky tu) de gia tri thay doi khong lam xo
            # day cac o ben canh
            f = tk.Frame(qrow, bg=WHITE)
            f.grid(row=0, column=col, sticky="w", padx=(0, 8))
            tk.Label(f, text=caption, bg=WHITE, fg=SEC_TX, font=FONT_SM,
                     width=w, anchor="w").pack(anchor="w")
            v = tk.Label(f, text="--", bg=WHITE, fg=DIS_FG, font=FONT_KPI,
                         width=w, anchor="w")
            v.pack(anchor="w")
            return v

        self.lbl_rf_link = _kpi(0, "RF LINK", 15)
        self.lbl_q_loss = _kpi(1, "Mất gói (heartbeat)", 15)
        self.lbl_q_redund = _kpi(2, "Dự phòng (tự động)", 14)
        self.lbl_q_hb = _kpi(3, "Heartbeat TX / RX", 13)
        self.lbl_q_rate = _kpi(4, "ĐÁNH GIÁ LẮP ĐẶT", 18)

        qbot = tk.Frame(qcard, bg=WHITE)
        qbot.pack(fill="x", padx=10, pady=(0, 6))
        self.lbl_rf_link_detail = tk.Label(
            qbot, text="(đánh giá theo % mất heartbeat + mức dự phòng)",
            bg=WHITE, fg=SEC_TX, font=FONT_SM)
        self.lbl_rf_link_detail.pack(side="left")
        sec_btn(qbot, "Đọc ngay (rf stat)", command=self._refresh_rf_link_stat).pack(side="right")
        sec_btn(qbot, "Reset % mất gói (rf reset)",
                command=self._reset_rf_stats).pack(side="right", padx=(0, 6))

        # --- Breakdown TUNG dev_id: con nao ROT (DOWN) / con nao LINK KEM (loss cao) ---
        # Dong RF_LINK= toan cuc luon UP neu con nghe duoc BAT KY ai; panel nay
        # chi ro tung thiet bi. Mau: xanh=tot(<2%), cam=kha(2-10%), do=kem(>10%)/DOWN.
        qdev = tk.Frame(qcard, bg=WHITE)
        qdev.pack(fill="x", padx=10, pady=(0, 10))
        self.lbl_dev_hdr = tk.Label(
            qdev, text="Thiết bị kết nối (số % = tỷ lệ mất gói, càng nhỏ càng tốt):",
            bg=WHITE, fg=SEC_TX, font=FONT_SM)
        self.lbl_dev_hdr.pack(anchor="w")
        self.dev_panel = tk.Frame(qdev, bg=WHITE)
        self.dev_panel.pack(fill="x", pady=(5, 0))
        # 16 O CO DINH (dev_id 1..16) tao SAN 1 lan, luon hien thi. Cap nhat chi
        # doi mau + text (khong tao/xoa dong) - chac chan hien, khong bi trong.
        self.DEV_MAX = 16
        self.DEV_NEUTRAL = "#EEF1F4"
        self.dev_cells = {}
        _cols = 8
        _rowf = None
        for _idx in range(self.DEV_MAX):
            _dev_id = _idx + 1
            if _idx % _cols == 0:
                _rowf = tk.Frame(self.dev_panel, bg=WHITE)
                _rowf.pack(anchor="w")
            _cell = tk.Frame(_rowf, bg=self.DEV_NEUTRAL, bd=1, relief="solid")
            _cell.pack(side="left", padx=3, pady=3)
            _lid = tk.Label(_cell, text="#%d" % _dev_id, bg=self.DEV_NEUTRAL, fg=DIS_FG,
                            font=("Segoe UI", 15, "bold"), width=3, padx=6)
            _lid.pack(pady=(4, 0))
            _lsub = tk.Label(_cell, text="—", bg=self.DEV_NEUTRAL, fg=DIS_FG,
                             font=("Segoe UI", 10, "bold"), width=6)
            _lsub.pack(pady=(0, 5))
            self.dev_cells[_dev_id] = (_cell, _lid, _lsub)

        # --- gui thu RS485 de tu kiem tra forward khi chua co Modbus that ---
        send_row = tk.Frame(parent, bg=MAIN_BG)
        send_row.pack(fill="x", pady=(0, 6))
        tk.Label(send_row, text="Gửi thử lên RS485 của board này:", bg=MAIN_BG, fg=INK,
                 font=FONT_SM).pack(side="left")
        self.entry_rs485 = tk.Entry(send_row, font=FONT_MONO, width=30)
        self.entry_rs485.insert(0, "MBW TEST")
        self.entry_rs485.pack(side="left", padx=6)
        flat_btn(send_row, "Gửi (rs485 <text>)", PRIMARY, PRIMARY_HOV,
                  command=self._send_rs485_test).pack(side="left")

        # dong tong hop KET QUA KIEM TRA KHUNG (CRC16 + exception Modbus)
        chk_row = tk.Frame(parent, bg=MAIN_BG)
        chk_row.pack(fill="x", pady=(0, 2))
        tk.Label(chk_row, text="Kiểm tra khung Modbus:", bg=MAIN_BG, fg=INK,
                 font=FONT_SM).pack(side="left")
        # nguong do dai TUY CHINH: khung < "Ngan" hoac > "Dai" bi coi la LOI
        # (VD manh vo "02 00 44" 3 byte tach ra tu response bi mat dau).
        # Modbus RTU hop le: exception = 5B, request FC3 = 8B, toi da 256B.
        tk.Label(chk_row, text="Ngắn <", bg=MAIN_BG, fg=SEC_TX, font=FONT_SM).pack(side="left", padx=(12, 2))
        self.ent_chk_min = tk.Entry(chk_row, width=4, font=FONT_SM, justify="center")
        self.ent_chk_min.insert(0, str(self.cfg.get("chk_min_len", 5)))
        self.ent_chk_min.pack(side="left")
        tk.Label(chk_row, text="byte, Dài >", bg=MAIN_BG, fg=SEC_TX, font=FONT_SM).pack(side="left", padx=2)
        self.ent_chk_max = tk.Entry(chk_row, width=5, font=FONT_SM, justify="center")
        self.ent_chk_max.insert(0, str(self.cfg.get("chk_max_len", 256)))
        self.ent_chk_max.pack(side="left")
        tk.Label(chk_row, text="byte", bg=MAIN_BG, fg=SEC_TX, font=FONT_SM).pack(side="left", padx=(2, 8))
        self.lbl_chk = tk.Label(chk_row, text="(chưa có khung)", bg=MAIN_BG, fg=DIS_FG,
                                 font=FONT_B, width=58, anchor="w")
        self.lbl_chk.pack(side="left", padx=6)
        self.chk_counts = {}

        # --- ghi log Excel: moi khung forward 1 dong (gio, huong, so byte,
        # ket qua kiem tra, hex). Ghi theo LO moi 5s de khong gia GUI khi
        # traffic cao (openpyxl luu lai ca file moi lan save). Nut dat o hang
        # "Gui thu RS485" (send_row) vi hang kiem tra khung da chat cho. ---
        self._xlog_on = False
        self.btn_xlog = sec_btn(send_row, "⏺ Ghi Excel", command=self._fwd_log_toggle)
        self.btn_xlog.pack(side="right", padx=(0, 4))
        self.lbl_xlog = tk.Label(send_row, text="", bg=MAIN_BG, fg=SEC_TX,
                                  font=FONT_SM, width=30, anchor="e")
        self.lbl_xlog.pack(side="right", padx=6)

        cols = ("time", "dir", "len", "check", "preview")
        self.fwd_tree = ttk.Treeview(parent, columns=cols, show="headings", height=14)
        self.fwd_tree.heading("time", text="Giờ")
        self.fwd_tree.heading("dir", text="Hướng")
        self.fwd_tree.heading("len", text="Số byte")
        self.fwd_tree.heading("check", text="Kiểm tra")
        self.fwd_tree.heading("preview", text="Preview (hex)")
        self.fwd_tree.column("time", width=85, anchor="center")
        self.fwd_tree.column("dir", width=110, anchor="center")
        self.fwd_tree.column("len", width=60, anchor="center")
        self.fwd_tree.column("check", width=95, anchor="center")
        self.fwd_tree.column("preview", width=480)
        self.fwd_tree.pack(fill="both", expand=True, pady=(0, 6))
        self.fwd_tree.tag_configure("rs485_rf", foreground=PRIMARY)
        self.fwd_tree.tag_configure("rf_rs485", foreground=WARN_FG)
        self.fwd_tree.tag_configure("bad", foreground=FAIL_FG)  # khung loi: chu do

        tk.Label(parent, text="Console thô (mọi dòng board in ra):", bg=MAIN_BG, fg=INK,
                 font=FONT_SM).pack(anchor="w")
        self.raw_log = tk.Text(parent, height=8, bg=TERM_BG, fg=TERM_FG, font=FONT_MONO,
                                insertbackground=WHITE)
        self.raw_log.pack(fill="x")

    def _chk_limits(self):
        """Nguong do dai khung tu 2 o nhap (tuy chinh duoc khi dang chay)."""
        try:
            lo = int(self.ent_chk_min.get())
        except (ValueError, AttributeError):
            lo = 5
        try:
            hi = int(self.ent_chk_max.get())
        except (ValueError, AttributeError):
            hi = 256
        return max(1, lo), max(lo, hi)

    def _analyze_fwd_frame(self, length, preview):
        """Kiem tra 1 khung tu dong FWD: do dai (nguong tuy chinh) + CRC16 +
        exception Modbus. Tra ve (nhan hien thi, True neu la khung LOI).
        Gioi han: firmware chi in toi da 16 byte dau (FWD_PREVIEW_MAX) - khung
        dai hon nguong nay khong du du lieu tinh CRC -> '(>16B)', bo qua."""
        lo, hi = self._chk_limits()
        if length < lo:   # manh vo kieu "02 00 44" (response mat dau)
            return "NGẮN %dB" % length, True
        if length > hi:   # dai bat thuong (dinh khung / rac bus)
            return "QUÁ DÀI %dB" % length, True
        if "..." in preview:
            return "(>16B)", False
        try:
            data = bytes(int(x, 16) for x in preview.split())
        except ValueError:
            return "?", False
        if len(data) != length:
            return "(>16B)", False
        if modbus_crc16(data) != 0:
            return "CRC LỖI", True
        if data[1] & 0x80:  # slave tra loi exception (FC | 0x80, byte 3 = ma loi)
            return "EXC %02X" % data[2], True
        return "OK", False

    def _add_fwd_row(self, direction, length, preview):
        self.fwd_count[direction] = self.fwd_count.get(direction, 0) + 1
        tag = "rs485_rf" if direction == "RS485->RF" else "rf_rs485"
        arrow = "RS485 → RF" if direction == "RS485->RF" else "RF → RS485"

        verdict, bad = self._analyze_fwd_frame(length, preview)
        k = ("crc" if verdict.startswith("CRC") else
             "exc" if verdict.startswith("EXC") else
             "short" if verdict.startswith("NGẮN") else
             "long" if verdict.startswith("QUÁ DÀI") else
             "ok" if verdict == "OK" else "skip")
        self.chk_counts[k] = self.chk_counts.get(k, 0) + 1
        c = self.chk_counts
        n_bad = (c.get("crc", 0) + c.get("exc", 0) + c.get("short", 0) +
                 c.get("long", 0))
        self.lbl_chk.config(
            text="OK %d  |  CRC lỗi %d  |  Exception %d  |  ngắn %d  |  quá dài %d  |  bỏ qua (>16B) %d" %
                 (c.get("ok", 0), c.get("crc", 0), c.get("exc", 0),
                  c.get("short", 0), c.get("long", 0), c.get("skip", 0)),
            fg=FAIL_FG if n_bad else PASS_FG)

        now = datetime.now()
        self.fwd_tree.insert("", 0, values=(now.strftime("%H:%M:%S"), arrow, length,
                                            verdict, preview),
                              tags=(("bad",) if bad else (tag,)))
        if self._xlog_on:
            self._xlog_buf.append((now.strftime("%d/%m/%Y"),
                                   now.strftime("%H:%M:%S.%f")[:-3],
                                   arrow, length, verdict, preview))
        a = self.fwd_count.get("RS485->RF", 0)
        b = self.fwd_count.get("RF->RS485", 0)
        self.lbl_stat.config(text="RS485→RF: %d    RF→RS485: %d" % (a, b))
        self._update_err_label(a, b)

    def _update_err_label(self, a, b):
        """% loi giao dich = |hieu so 2 chieu| / chieu lon hon. Hien ca TONG
        (tu luc mo app) va 60S GAN NHAT - so 60s moi phan anh chat luong
        HIEN TAI (tong bi "keo" boi qua khu). Mau theo so 60s neu co, khong
        thi theo tong."""
        now = time.time()
        self._fwd_hist.append((now, a, b))
        total = max(a, b)
        err_cum = (abs(a - b) / total * 100.0) if total > 0 else 0.0

        # cua so 60 giay gan nhat
        err_60 = None
        for (t0, a0, b0) in self._fwd_hist:
            if now - t0 <= 60.0:
                da, db = a - a0, b - b0
                dmax = max(da, db)
                if dmax >= 20:  # du mau moi tinh, tranh nhieu thong ke
                    err_60 = abs(da - db) / dmax * 100.0
                break

        self._last_err = (err_cum, err_60)  # cho log Excel dinh ky
        basis = err_60 if err_60 is not None else err_cum
        fg = PASS_FG if basis <= 1.0 else (WARN_FG if basis <= 2.0 else FAIL_FG)
        rate = "TỐT" if basis <= 1.0 else ("CHẤP NHẬN" if basis <= 2.0 else "KÉM")
        if err_60 is not None:
            txt = "LỖI: %.1f%% (60s: %.1f%%) - %s" % (err_cum, err_60, rate)
        else:
            txt = "LỖI: %.1f%% - %s" % (err_cum, rate)
        self.lbl_err.config(text=txt, fg=fg)

    # ---------- GHI LOG FORWARD RA EXCEL ----------
    def _fwd_log_toggle(self):
        if self._xlog_on:
            self._xlog_on = False
            self._fwd_log_flush(final=True)
            self.btn_xlog.config(text="⏺ Ghi Excel")
            return
        use_xlsx = openpyxl is not None
        default_name = "fwd_log_%s.%s" % (datetime.now().strftime("%Y%m%d_%H%M%S"),
                                          "xlsx" if use_xlsx else "csv")
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx" if use_xlsx else ".csv",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
            initialfile=default_name)
        if not path:
            return
        header = ("Ngày", "Giờ", "Hướng", "Số byte", "Kiểm tra", "Hex")
        # sheet 2: snapshot TOAN BO chi so tab Forward moi 5s - de truy vet
        # mat frame/mat goi theo thoi gian (doi chieu voi sheet khung & log
        # Modbus Poll bang timestamp)
        header2 = ("Ngày", "Giờ", "RS485→RF", "RF→RS485", "Hiệu số",
                   "Lỗi tổng %", "Lỗi 60s %", "RF LINK", "Mất gói %",
                   "Dự phòng", "HB_TX", "HB_RX",
                   "OK", "CRC lỗi", "Exception", "Ngắn", "Quá dài", "Bỏ qua >16B")
        self._xlog_csv = path.lower().endswith(".csv") or not use_xlsx
        if self._xlog_csv and not path.lower().endswith(".csv"):
            path = path.rsplit(".", 1)[0] + ".csv"
            messagebox.showinfo(APP_TITLE, "Thiếu openpyxl - ghi CSV thay thế:\n%s" % path)
        try:
            if self._xlog_csv:
                with open(path, "w", encoding="utf-8-sig") as f:
                    f.write(",".join(header) + "\n")
                self._xlog_path2 = path.rsplit(".", 1)[0] + "_link.csv"
                with open(self._xlog_path2, "w", encoding="utf-8-sig") as f:
                    f.write(",".join(header2) + "\n")
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "FWD log"
                ws.append(header)
                ws2 = wb.create_sheet("Link 5s")
                ws2.append(header2)
                wb.save(path)
                self._xlog_wb, self._xlog_ws, self._xlog_ws2 = wb, ws, ws2
        except PermissionError:
            messagebox.showerror(APP_TITLE, "Không ghi được file (đang mở trong Excel?)")
            return
        self._xlog_path = path
        self._xlog_buf = []
        self._xlog_rows = 0
        self._xlog_on = True
        self.btn_xlog.config(text="⏹ Dừng ghi")
        self.lbl_xlog.config(text="%s | 0 dòng" % os.path.basename(path))
        self.root.after(5000, self._fwd_log_tick)

    def _fwd_log_tick(self):
        if not self._xlog_on:
            return
        self._fwd_log_flush()
        self.root.after(5000, self._fwd_log_tick)

    def _link_snapshot_row(self):
        """1 dong tong hop TOAN BO chi so tab Forward tai thoi diem hien tai."""
        now = datetime.now()
        a = self.fwd_count.get("RS485->RF", 0)
        b = self.fwd_count.get("RF->RS485", 0)
        err_cum, err_60 = getattr(self, "_last_err", (0.0, None))
        rf = getattr(self, "_last_rf", None)
        c = self.chk_counts
        return (now.strftime("%d/%m/%Y"), now.strftime("%H:%M:%S.%f")[:-3],
                a, b, abs(a - b),
                "%.2f" % err_cum,
                ("%.2f" % err_60) if err_60 is not None else "",
                (("UP" if rf["up"] else "DOWN") if rf else ""),
                ("%.1f" % rf["loss_pct"]) if rf else "",
                rf["redund"] if rf else "", rf["hb_tx"] if rf else "",
                rf["hb_rx"] if rf else "",
                c.get("ok", 0), c.get("crc", 0), c.get("exc", 0),
                c.get("short", 0), c.get("long", 0), c.get("skip", 0))

    def _fwd_log_flush(self, final=False):
        buf = getattr(self, "_xlog_buf", [])
        snap = self._link_snapshot_row()
        if buf or snap:
            self._xlog_buf = []
            try:
                if self._xlog_csv:
                    if buf:
                        with open(self._xlog_path, "a", encoding="utf-8-sig") as f:
                            for r in buf:
                                f.write(",".join(str(x) for x in r) + "\n")
                    with open(self._xlog_path2, "a", encoding="utf-8-sig") as f:
                        f.write(",".join(str(x) for x in snap) + "\n")
                else:
                    for r in buf:
                        self._xlog_ws.append(r)
                    self._xlog_ws2.append(snap)
                    self._xlog_wb.save(self._xlog_path)
                self._xlog_rows += len(buf)
            except PermissionError:
                # file dang mo trong Excel - giu lai buffer, thu lai lan sau
                self._xlog_buf = buf + self._xlog_buf
                self.lbl_xlog.config(text="LỖI: file đang mở trong Excel!", fg=FAIL_FG)
                return
        self.lbl_xlog.config(
            text="%s | %d dòng%s" % (os.path.basename(self._xlog_path),
                                     self._xlog_rows,
                                     " (đã dừng)" if final else ""),
            fg=SEC_TX)

    def _require_conn(self):
        if not self.link.is_open():
            messagebox.showwarning(APP_TITLE, "Chưa kết nối board.")
            return False
        return True

    def _bridge_toggle(self):
        if not self._require_conn():
            return
        self.link.send("bridge stat")
        line = self.link.wait_for("BRIDGE=", 1.0) or ""
        want_off = "BRIDGE=ON" in line
        self.link.send("bridge off" if want_off else "bridge on")
        self.root.after(200, self._refresh_bridge_stat)

    def _bridge_log_toggle(self):
        if not self._require_conn():
            return
        self.link.send("bridge stat")
        line = self.link.wait_for("BRIDGE=", 1.0) or ""
        want_off = "LOG=ON" in line
        self.link.send("bridge log off" if want_off else "bridge log on")
        self.root.after(200, self._refresh_bridge_stat)

    def _refresh_bridge_stat(self):
        if not self.link.is_open():
            return
        self.link.send("bridge stat")
        line = self.link.wait_for("BRIDGE=", 1.0)
        if not line:
            return
        on = "BRIDGE=ON" in line
        log_on = "LOG=ON" in line
        self.btn_bridge.config(text="Bridge: BẬT" if on else "Bridge: TẮT",
                                bg=PASS_FG if on else FAIL_FG,
                                activebackground=PASS_FG if on else FAIL_FG)
        self.btn_bridge_log.config(text="Log: BẬT" if log_on else "Log: TẮT")

    # ---------- RF LINK QUALITY (heartbeat, giong dong ho telemetry) ----------
    def _refresh_rf_link_stat(self):
        """Doc 'rf stat' o THREAD NEN roi cap nhat KPI qua after(). Truoc day
        doc ngay tren GUI thread voi timeout 1s: (1) GUI dong bang moi lan
        doi, (2) khi Modbus chay console ngap dong FWD lam phan hoi tre >1s
        -> KPI mai la '--'."""
        if not self.link.is_open() or getattr(self, "_rf_q_busy", False):
            return
        self._rf_q_busy = True
        threading.Thread(target=self._rf_link_query_worker, daemon=True).start()

    # ---------- NET ID (cau hinh chung 1 mang, luu Flash) ----------
    def _set_net_id(self):
        """Gui 'net id <n>' - firmware luu Flash + ap dung ngay (khong can reset)."""
        if not self.link.is_open():
            self.lbl_net_id.config(text="chưa kết nối board", fg=FAIL_FG)
            return
        try:
            n = int(self.net_id_var.get().strip())
        except ValueError:
            self.lbl_net_id.config(text="ID không hợp lệ", fg=FAIL_FG)
            return
        if not (0 <= n <= 63):
            self.lbl_net_id.config(text="ID phải 0–63", fg=FAIL_FG)
            return
        self.lbl_net_id.config(text="đang lưu...", fg=SEC_TX)
        threading.Thread(target=self._net_id_worker, args=("net id %d" % n, n), daemon=True).start()

    def _read_net_id(self):
        """Doc NET_ID hien tai tu board (khi vua ket noi) de dien san vao o."""
        if not self.link.is_open():
            return
        threading.Thread(target=self._net_id_worker, args=("net id", None), daemon=True).start()

    def _net_id_worker(self, cmd, set_val):
        self.link.send(cmd)
        line = self.link.wait_for("NET_ID", 2.5)
        m = re.search(r"NET_ID=(\d+)", line) if line else None
        def apply():
            if set_val is not None:  # lenh SET
                ok = m is not None and int(m.group(1)) == set_val
                self.lbl_net_id.config(
                    text=("✓ NET_ID=%d (đã lưu Flash)" % set_val) if ok else "lỗi / không phản hồi",
                    fg=PASS_FG if ok else FAIL_FG)
            elif m:                  # lenh DOC
                self.net_id_var.set(m.group(1))
                self.lbl_net_id.config(text="hiện tại: %s" % m.group(1), fg=SEC_TX)
        self.root.after(0, apply)

    # ---------- DONG HO RTC (moc thoi gian cho log) ----------
    def _sync_rtc_pc(self):
        """Gui 'rtc set hh:mm:ss' theo gio may tinh -> board luu vao chip RTC
        PCF85063 (co pin CR1220 nuoi thi giu qua mat nguon)."""
        if not self.link.is_open():
            self.lbl_rtc.config(text="chưa kết nối board", fg=FAIL_FG)
            return
        now = datetime.now()
        self.link.send("rtc set %02d/%02d/%02d %02d:%02d:%02d" % (
            now.day, now.month, now.year % 100, now.hour, now.minute, now.second))
        self.lbl_rtc.config(text="đã đồng bộ %02d/%02d %02d:%02d:%02d..." % (
            now.day, now.month, now.hour, now.minute, now.second), fg=SEC_TX)
        self.root.after(700, self._read_rtc)  # doc lai xac nhan

    def _read_rtc(self):
        if not self.link.is_open():
            return
        threading.Thread(target=self._read_rtc_worker, daemon=True).start()

    def _read_rtc_worker(self):
        self.link.send("rtc")
        line = self.link.wait_for("RTC:", 2.5)
        m = re.search(r"RTC:\s*(\d\d/\d\d/\d\d)\s+(\d\d:\d\d:\d\d)", line) if line else None
        def apply():
            if m:
                self.lbl_rtc.config(text="giờ board: %s %s ✓" % (m.group(1), m.group(2)), fg=PASS_FG)
            elif line:  # co dong RTC nhung la "--/--/-- --:--:--" (chua dat / mat pin)
                self.lbl_rtc.config(text="giờ board: chưa đặt (lắp pin + đồng bộ)", fg=WARN_FG)
        self.root.after(0, apply)

    # ---------- LOG SU KIEN mat/khoi phuc link (luu Flash) ----------
    def _read_flash_log(self):
        if not self.link.is_open():
            return
        threading.Thread(target=self._read_flash_log_worker, daemon=True).start()

    def _read_flash_log_worker(self):
        start_idx = self.link.snapshot_len()
        self.link.send("log")
        t0 = time.time()
        done = False
        while time.time() - t0 < 5.0 and not done:
            for ln in self.link.lines_since(start_idx):
                if ln.startswith("LOG_END"):
                    done = True
                    break
            time.sleep(0.05)
        rows = []
        for ln in self.link.lines_since(start_idx):
            if ln.startswith("LOG_END"):
                break
            m = LOG_LINE_RE.match(ln)
            if m:
                rows.append(m.groups())  # (idx, time, dev, evt, loss)
        self.root.after(0, lambda: self._show_log_window(rows))

    def _show_log_window(self, rows):
        win = tk.Toplevel(self.root)
        win.title("Log sự kiện mất link RF (lưu Flash)")
        win.configure(bg=WHITE)
        win.geometry("640x480")
        topbar = tk.Frame(win, bg=WHITE)
        topbar.pack(fill="x", padx=10, pady=(8, 4))
        tk.Label(topbar, text="DOWN = mất link. Ngày/giờ theo RTC của board. Tổng: %d sự kiện" % len(rows),
                 bg=WHITE, fg=SEC_TX, font=FONT_SM).pack(side="left")
        sec_btn(topbar, "💾 Lưu ra file (.log/.csv)",
                command=lambda: self._save_log_file(rows)).pack(side="right")
        cols = ("idx", "date", "time", "dev", "evt", "loss")
        tv = ttk.Treeview(win, columns=cols, show="headings")
        for _c, _t, _w in [("idx", "#", 45), ("date", "Ngày", 80), ("time", "Giờ (RTC)", 90),
                           ("dev", "dev_id", 70), ("evt", "Sự kiện", 110), ("loss", "Loss lúc đó", 90)]:
            tv.heading(_c, text=_t)
            tv.column(_c, width=_w, anchor="center")
        tv.tag_configure("down", foreground=FAIL_FG)
        tv.tag_configure("up", foreground=PASS_FG)
        for idx, dt, tm, dev, evt, loss in rows:
            tv.insert("", "end",
                      values=(idx, dt, tm, dev, "MẤT LINK" if evt == "DOWN" else "Khôi phục", loss),
                      tags=("down" if evt == "DOWN" else "up",))
        tv.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        if not rows:
            tk.Label(win, text="(chưa có sự kiện nào trong Flash)",
                     bg=WHITE, fg=DIS_FG, font=FONT_SM).pack(pady=6)

    def _save_log_file(self, rows):
        """Luu log ra file .log (text) hoac .csv (mo Excel duoc)."""
        path = filedialog.asksaveasfilename(
            defaultextension=".log",
            filetypes=[("Log text", "*.log"), ("CSV (Excel)", "*.csv"), ("Text", "*.txt")],
            initialfile="mbw_rf24_log")
        if not path:
            return
        try:
            is_csv = path.lower().endswith(".csv")
            with open(path, "w", encoding="utf-8-sig") as f:
                if is_csv:
                    f.write("STT,Ngay,Gio,dev_id,Su kien,Loss\n")
                    for idx, dt, tm, dev, evt, loss in rows:
                        f.write("%s,%s,%s,%s,%s,%s\n" % (
                            idx, dt, tm, dev, "MAT LINK" if evt == "DOWN" else "Khoi phuc", loss))
                else:
                    f.write("# MBW RF24 - Log su kien mat link RF (Flash), %d su kien\n" % len(rows))
                    for idx, dt, tm, dev, evt, loss in rows:
                        f.write("[%s] %s %s  dev=%s  %s  loss=%s\n" % (
                            idx, dt, tm, dev, "MAT LINK" if evt == "DOWN" else "Khoi phuc", loss))
            messagebox.showinfo(APP_TITLE, "Đã lưu %d sự kiện:\n%s" % (len(rows), path))
        except Exception as e:
            messagebox.showerror(APP_TITLE, "Không lưu được file:\n%s" % e)

    def _clear_flash_log(self):
        if not self.link.is_open():
            return
        if not messagebox.askyesno(APP_TITLE, "Xóa TOÀN BỘ log sự kiện trong Flash?\n(không khôi phục được)"):
            return
        self.link.send("log clear")

    def _reset_rf_stats(self):
        """Gui 'rf reset' - xoa bo dem % mat goi de bat dau CUA SO DO MOI cho
        test (moi thiet bi phat 1 heartbeat/giay; cho >=10s roi doc lai)."""
        if not self.link.is_open():
            return
        self.link.send("rf reset")
        self._seen_rf_devs = False  # cho firmware in lai RF_DEVS voi so lieu moi
        if hasattr(self, "dev_cells"):
            for _cell, _lid, _lsub in self.dev_cells.values():
                _cell.config(bg=self.DEV_NEUTRAL)
                _lid.config(bg=self.DEV_NEUTRAL, fg=DIS_FG)
                _lsub.config(bg=self.DEV_NEUTRAL, fg=DIS_FG, text="…")
        if hasattr(self, "lbl_q_loss"):
            self.lbl_q_loss.config(text="--", fg=DIS_FG)
        self.root.after(1200, self._refresh_rf_link_stat)  # doc lai sau 1 nhip

    def _rf_link_query_worker(self):
        try:
            self.link.send("rf stat")
            line = self.link.wait_for("RF_LINK=", 2.5)
        finally:
            self._rf_q_busy = False
        if not line:
            return
        self.root.after(0, self._rf_link_apply_stat, line)

    def _rf_link_apply_stat(self, line):
        m = RF_STAT_RE.search(line)
        if not m:
            return
        up_s, peer, age_ms, loss_pm, redund_s, hb_tx, hb_rx = m.groups()
        up = up_s == "UP"
        loss_pct = int(loss_pm) / 10.0  # phan nghin -> %
        redund = int(redund_s)
        # luu snapshot cho log Excel dinh ky (sheet "Link 5s")
        self._last_rf = dict(up=up, peer=peer, age_ms=age_ms, loss_pct=loss_pct,
                             redund=redund, hb_tx=hb_tx, hb_rx=hb_rx)

        self._apply_link_ui(up, peer)

        # FALLBACK: neu firmware CU (chua in dong "RF_DEVS:") thi panel dev_id se
        # trong. Tam dung so lieu toan cuc tu "rf stat" de it nhat hien THIET BI
        # DANG NGHE (peer) - nap firmware moi se co breakdown day du tung dev_id.
        if not getattr(self, "_seen_rf_devs", False):
            synth = "RF_DEVS: %s:%s(%ss,%s)" % (
                peer, "UP" if up else "DOWN", int(age_ms) // 1000, loss_pm)
            self._update_dev_panel(synth)

        # Mat goi: <2%% xanh (TOT) / 2-10%% vang (KHA) / >10%% do (KEM)
        loss_fg = PASS_FG if loss_pct < 2 else (WARN_FG if loss_pct < 10 else FAIL_FG)
        self.lbl_q_loss.config(text="%.1f %%" % loss_pct, fg=loss_fg if up else DIS_FG)

        # Du phong dang phai dung: 2-3 binh thuong / 4-5 dang gong / 6+ te
        red_fg = PASS_FG if redund <= 3 else (WARN_FG if redund <= 5 else FAIL_FG)
        self.lbl_q_redund.config(text="%dx" % redund, fg=red_fg if up else DIS_FG)

        self.lbl_q_hb.config(text="%s / %s" % (hb_tx, hb_rx), fg=INK if up else DIS_FG)

        if not up:
            rate, rate_fg = "MẤT LINK", FAIL_FG
        elif loss_pct < 2 and redund <= 3:
            rate, rate_fg = "TỐT", PASS_FG
        elif loss_pct < 10 and redund <= 5:
            rate, rate_fg = "KHÁ", WARN_FG
        else:
            rate, rate_fg = "KÉM", FAIL_FG
        self.lbl_q_rate.config(text=rate, fg=rate_fg)

        self.lbl_rf_link_detail.config(
            text="peer=%s | heartbeat cuối %sms trước" % (peer, age_ms))

    def _apply_link_ui(self, up, peer):
        self.lbl_rf_link.config(
            text="● UP (peer=%s)" % peer if up else "● DOWN",
            fg=PASS_FG if up else FAIL_FG)

    def _schedule_rf_link_poll(self):
        """Tu dong doc 'rf stat' dinh ky (giong man hinh 'link quality' cua
        GCS/telemetry) - chi chay khi con dang ket noi board."""
        if self.link.is_open():
            self._refresh_rf_link_stat()
            self.root.after(3000, self._schedule_rf_link_poll)

    def _send_rs485_test(self):
        if not self._require_conn():
            return
        text = self.entry_rs485.get().strip() or "MBW TEST"
        self.link.send("rs485 %s" % text)

    def _open_modbus_poll(self):
        """Mo cong cu Modbus Poll Test (modbus_poll_app.py) trong 1 tien trinh/
        cua so rieng - dung de bom Modbus RTU that qua cong RS485 vat ly cua
        board (khac voi cong COM console CLI dang ket noi o tren). Chay tach
        tien trinh de khong block app chinh va de dung dong thoi ca 2 (VD: 1
        may vua giam sat Forward vua chay Modbus Master qua RS485)."""
        try:
            if getattr(sys, "frozen", False):
                subprocess.Popen([sys.executable, "--modbus"])
            else:
                script_dir = os.path.dirname(os.path.abspath(__file__))
                subprocess.Popen([sys.executable,
                                   os.path.join(script_dir, "mbw_test_app.py"), "--modbus"])
        except Exception as e:
            messagebox.showerror(APP_TITLE, "Không mở được Modbus Poll Test:\n%s" % e)

    # ---------- TAB: TEST TU DONG ----------
    def _build_test_tab(self, parent):
        top = tk.Frame(parent, bg=MAIN_BG)
        top.pack(fill="x", pady=(6, 4))

        self.btn_run = flat_btn(top, "▶ Run Test", PRIMARY, PRIMARY_HOV,
                                 command=self.run_test_async)
        self.btn_run.pack(side="left")

        sec_btn(top, "Export Report", command=self.export_report).pack(side="left", padx=8)

        self.progress = ttk.Progressbar(top, mode="determinate")
        self.progress.pack(side="left", fill="x", expand=True, padx=12)

        cols = ("name", "status", "detail")
        self.tree = ttk.Treeview(parent, columns=cols, show="headings", height=10)
        self.tree.heading("name", text="Bước test")
        self.tree.heading("status", text="Kết quả")
        self.tree.heading("detail", text="Chi tiết")
        self.tree.column("name", width=380)
        self.tree.column("status", width=90, anchor="center")
        self.tree.column("detail", width=420)
        self.tree.pack(fill="both", expand=True, pady=6)

        self.tree.tag_configure("pass", foreground=PASS_FG)
        self.tree.tag_configure("fail", foreground=FAIL_FG)
        self.tree.tag_configure("run", foreground=RUN_FG)
        self.tree.tag_configure("pend", foreground=DIS_FG)

        for g in self.cfg.get("test_groups", []):
            self.tree.insert("", "end", iid=g["key"], values=(g["name"], "chờ", ""), tags=("pend",))

        # Double-click 1 dong = chay RIENG buoc do (debug/kiem lai 1 muc
        # khong can chay ca chuoi)
        self.tree.bind("<Double-1>", self._run_single_test)
        tk.Label(parent, text="(Double-click một dòng để chạy riêng bước đó)",
                 bg=MAIN_BG, fg=SEC_TX, font=FONT_SM).pack(anchor="w")

        self.log = tk.Text(parent, height=6, bg=TERM_BG, fg=TERM_FG, font=FONT_MONO,
                            insertbackground=WHITE)
        self.log.pack(fill="x", pady=(4, 0))

    def _log(self, msg):
        self.log.insert("end", "[%s] %s\n" % (datetime.now().strftime("%H:%M:%S"), msg))
        self.log.see("end")

    def _set_row(self, key, status, detail=""):
        tagmap = {"PASS": "pass", "FAIL": "fail", "RUN": "run", "PEND": "pend"}
        vals = self.tree.item(key, "values")
        self.tree.item(key, values=(vals[0], status, detail), tags=(tagmap.get(status, "pend"),))
        self.root.update_idletasks()

    def run_test_async(self):
        if self.test_running:
            return
        if not self.link.is_open():
            messagebox.showwarning(APP_TITLE, "Cần kết nối board trước khi Run Test.")
            return
        t = threading.Thread(target=self._run_test_seq, daemon=True)
        t.start()

    def _run_test_seq(self):
        self.test_running = True
        groups = self.cfg.get("test_groups", [])
        self.progress["maximum"] = max(1, len(groups))
        self.progress["value"] = 0
        pass_count = 0

        # TAT log forward trong luc chay test: neu Modbus dang chay qua bridge,
        # console bi ngap dong "FWD ..." lam phan hoi lenh CLI tre/chim -> moi
        # buoc deu FAIL oan mac du board van tot. Xong chuoi test se BAT lai.
        self.link.send("bridge log off")
        time.sleep(0.5)
        self._log("(đã tạm tắt log FWD trong lúc test - sẽ bật lại khi xong)")

        for g in groups:
            key = g["key"]
            self._set_row(key, "RUN")
            self._log("Chạy: %s" % g["name"])

            prereq = g.get("manual_prereq")
            if prereq:
                self._log("  (*) Yêu cầu: %s" % prereq)

            ok, detail = self._exec_test_step(g)

            self._set_row(key, "PASS" if ok else "FAIL", detail)
            self._log("  -> %s %s" % ("PASS" if ok else "FAIL", detail))
            if ok:
                pass_count += 1
            self.progress["value"] += 1

        self.link.send("bridge log on")  # bat lai log FWD (da tat o dau chuoi test)
        self._log("Hoàn tất: %d/%d bước PASS" % (pass_count, len(groups)))
        self.test_running = False
        self._last_result_summary = (pass_count, len(groups))
        self._append_report_row()

    def _exec_test_step(self, g):
        """Chay 1 buoc test - dung chung cho Run Test (ca chuoi) va
        double-click (chay rieng)."""
        try:
            check = g.get("check")
            if check in ("rf_hb", "rf_loss"):
                # Buoc test RF link can PARSE SO LIEU (khong chi tim chuoi):
                # heartbeat 2 chieu / ti le khung mat so voi nguong.
                return self._test_rf_link_check(check, g)
            return self.link.cmd_and_wait(g["cmd"], g.get("expect_contains"),
                                          timeout_s=3.0)
        except Exception as e:
            return False, "Lỗi: %s" % e

    def _run_single_test(self, event=None):
        """Double-click 1 dong trong bang Factory Test -> chay rieng buoc do."""
        if self.test_running or not self._require_conn():
            return
        iid = self.tree.focus()
        g = next((x for x in self.cfg.get("test_groups", []) if x["key"] == iid), None)
        if not g:
            return
        self.test_running = True
        threading.Thread(target=self._run_single_seq, args=(g,), daemon=True).start()

    def _run_single_seq(self, g):
        self._set_row(g["key"], "RUN")
        self._log("Chạy riêng: %s" % g["name"])
        prereq = g.get("manual_prereq")
        if prereq:
            self._log("  (*) Yêu cầu: %s" % prereq)
        self.link.send("bridge log off")  # tranh console ngap FWD lam timeout oan
        time.sleep(0.4)
        ok, detail = self._exec_test_step(g)
        self.link.send("bridge log on")
        self._set_row(g["key"], "PASS" if ok else "FAIL", detail)
        self._log("  -> %s %s" % ("PASS" if ok else "FAIL", detail))
        self.test_running = False

    def _test_rf_link_check(self, check, g):
        """2 buoc Factory Test cho RF link (can board doi dien dang bat):
          - rf_hb  : heartbeat 2 chieu - LINK=UP va HB_RX > 0 (nghe duoc peer).
          - rf_loss: ti le khung mat LOSS_PROMILLE <= nguong (mac dinh 100‰).
        Doc 'rf stat' va parse RF_STAT_RE (khong dung expect_contains vi phai
        so sanh SO)."""
        self.link.send("rf stat")
        line = self.link.wait_for("RF_LINK=", 2.0)
        if not line:
            return False, "Không đọc được rf stat"
        m = RF_STAT_RE.search(line)
        if not m:
            return False, "Không parse được: %s" % line
        up = m.group(1) == "UP"
        peer, age = m.group(2), m.group(3)
        loss, redund = int(m.group(4)), m.group(5)
        hb_tx, hb_rx = int(m.group(6)), int(m.group(7))
        if check == "rf_hb":
            ok = up and hb_rx > 0
            detail = "LINK=%s PEER=%s HB tx/rx=%d/%d (%sms trước)" % (
                "UP" if up else "DOWN", peer, hb_tx, hb_rx, age)
            return ok, detail
        # rf_loss
        loss_max = int(g.get("loss_max_promille", 100))
        ok = up and loss <= loss_max
        detail = "LOSS=%d‰ (ngưỡng ≤%d‰) REDUND=%sx" % (loss, loss_max, redund)
        if not up:
            detail = "LINK DOWN - " + detail
        return ok, detail

    def _append_report_row(self):
        pass_count, total = getattr(self, "_last_result_summary", (0, 0))
        row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), self.link.port, pass_count, total,
               "PASS" if pass_count == total else "FAIL"]
        for g in self.cfg.get("test_groups", []):
            vals = self.tree.item(g["key"], "values")
            row.append(vals[1])

        if openpyxl is not None:
            try:
                if os.path.exists(REPORT_PATH):
                    wb = openpyxl.load_workbook(REPORT_PATH)
                    ws = wb.active
                else:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    header = ["Thời gian", "Cổng COM", "Số bước PASS", "Tổng số bước",
                               "Kết quả chung"] + [g["name"] for g in self.cfg.get("test_groups", [])]
                    ws.append(header)
                ws.append(row)
                wb.save(REPORT_PATH)
                self._log("Đã ghi báo cáo: %s" % REPORT_PATH)
            except Exception as e:
                self._log("Lỗi ghi report xlsx: %s" % e)
        else:
            csv_path = REPORT_PATH.replace(".xlsx", ".csv")
            new_file = not os.path.exists(csv_path)
            with open(csv_path, "a", encoding="utf-8") as f:
                if new_file:
                    f.write(",".join(["Thoi gian", "Cong COM", "So buoc PASS", "Tong so buoc",
                                        "Ket qua chung"] +
                                       [g["name"] for g in self.cfg.get("test_groups", [])]) + "\n")
                f.write(",".join(str(x) for x in row) + "\n")
            self._log("Đã ghi báo cáo (CSV, thiếu openpyxl): %s" % csv_path)

    def export_report(self):
        self._append_report_row()

    # ---------- TAB: TERMINAL ----------
    # ---------- TAB: STRESS TEST RF LINK ----------
    # Muc dich: danh gia DO ON DINH link RF de forward Modbus RS485. Cach lam:
    #   - Bom khung "rf tx" (data that qua RF, co ACK/redundant nhu khung
    #     forward) hoac "rs485" (ra bus RS485 vat ly) theo chu ky cai duoc.
    #   - Dinh ky doc "rf stat" + "bridge stat", tinh delta tung mau: LOSS,
    #     REDUND, HB, CRC error, FRAGDROP, DROP queue.
    #   - Ket thuc cham DAT/KHONG DAT theo nguong (xem _stress_verdict).
    def _build_stress_tab(self, parent):
        self.stress_running = False
        self.stress_stop_evt = threading.Event()
        self.stress_rows = []  # (t_s, link, loss, redund, d_hbtx, d_hbrx, d_crc, d_frag, d_drop, pump_ok, pump_fail)

        # --- hang cau hinh ---
        card = tk.Frame(parent, bg=WHITE, highlightbackground=CARD_BD, highlightthickness=1)
        card.pack(fill="x", padx=10, pady=(10, 6))
        row = tk.Frame(card, bg=WHITE)
        row.pack(fill="x", padx=10, pady=(8, 2))

        def _lbl(parent_, text):
            tk.Label(parent_, text=text, bg=WHITE, fg=INK, font=FONT).pack(side="left")

        def _ent(parent_, width, default):
            e = tk.Entry(parent_, width=width, font=FONT, justify="center")
            e.insert(0, default)
            e.pack(side="left", padx=(4, 14))
            return e

        _lbl(row, "Thời lượng (phút, 0=chạy mãi):")
        self.ent_st_dur = _ent(row, 5, "30")
        _lbl(row, "Chu kỳ mẫu (s):")
        self.ent_st_intv = _ent(row, 4, "5")
        _lbl(row, "Bơm khung:")
        self.cbo_st_pump = ttk.Combobox(row, state="readonly", width=20, font=FONT,
                                        values=["Tắt (chỉ heartbeat)",
                                                "rf tx (qua RF)",
                                                "rs485 (ra bus RS485)"])
        self.cbo_st_pump.current(1)
        self.cbo_st_pump.pack(side="left", padx=(4, 14))
        _lbl(row, "mỗi (ms):")
        self.ent_st_pms = _ent(row, 5, "200")
        _lbl(row, "size (byte, ≤63):")
        self.ent_st_plen = _ent(row, 4, "32")

        row2 = tk.Frame(card, bg=WHITE)
        row2.pack(fill="x", padx=10, pady=(2, 8))
        self.btn_st_start = flat_btn(row2, "▶ Bắt đầu", PRIMARY, PRIMARY_HOV,
                                     command=self.stress_start)
        self.btn_st_start.pack(side="left")
        self.btn_st_stop = sec_btn(row2, "■ Dừng", command=self.stress_stop_click,
                                   state="disabled")
        self.btn_st_stop.pack(side="left", padx=8)
        sec_btn(row2, "Xuất CSV", command=self.stress_export_csv).pack(side="left")
        self.lbl_st_verdict = tk.Label(row2, text="", bg=WHITE, font=FONT_CARD)
        self.lbl_st_verdict.pack(side="right", padx=6)

        # --- so lieu song (live) ---
        card2 = tk.Frame(parent, bg=WHITE, highlightbackground=CARD_BD, highlightthickness=1)
        card2.pack(fill="x", padx=10, pady=(0, 6))
        self.lbl_st_live = tk.Label(card2, text="(chưa chạy)", bg=WHITE, fg=INK,
                                    font=FONT_MONO, justify="left", anchor="w")
        self.lbl_st_live.pack(fill="x", padx=10, pady=8)

        # --- bang mau ---
        wrap = tk.Frame(parent, bg=MAIN_BG)
        wrap.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        cols = ("t", "link", "loss", "redund", "hbtx", "hbrx", "crc", "frag", "drop", "pok", "pfail")
        heads = ("Giây", "Link", "LOSS‰", "REDUND", "ΔHB_TX", "ΔHB_RX", "ΔCRC", "ΔFRAG", "ΔDROP", "Bơm OK", "Bơm FAIL")
        self.tree_st = ttk.Treeview(wrap, columns=cols, show="headings", height=9)
        for c, h in zip(cols, heads):
            self.tree_st.heading(c, text=h)
            self.tree_st.column(c, width=70, anchor="center")
        sb = ttk.Scrollbar(wrap, orient="vertical", command=self.tree_st.yview)
        self.tree_st.configure(yscrollcommand=sb.set)
        self.tree_st.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

    def _stress_query(self):
        """Doc 1 bo so lieu tu board: 'rf stat' (2 dong) + 'bridge stat'.
        Tra ve dict hoac None neu khong doc duoc (timeout/mat ket noi)."""
        if not self.link.is_open():
            return None
        idx = self.link.snapshot_len()
        self.link.send("rf stat")
        line = self.link.wait_for("RF_LINK=", 1.5)
        if not line:
            return None
        m = RF_STAT_RE.search(line)
        if not m:
            return None
        d = dict(up=(m.group(1) == "UP"), peer=int(m.group(2)),
                 age=int(m.group(3)), loss=int(m.group(4)),
                 redund=int(m.group(5)), hb_tx=int(m.group(6)),
                 hb_rx=int(m.group(7)),
                 crc=0, frag=0, dup=0, rf_tx=0, rx_ok=0, drop=0)
        for ln in self.link.lines_since(idx):
            mc = RF_CNT_RE.search(ln)
            if mc:
                d.update(rf_tx=int(mc.group(1)), rx_ok=int(mc.group(2)),
                         dup=int(mc.group(3)), crc=int(mc.group(4)),
                         frag=int(mc.group(5)))
                break
        idx2 = self.link.snapshot_len()
        self.link.send("bridge stat")
        lb = self.link.wait_for("BRIDGE=", 1.5)
        if lb:
            mb = BR_CNT_RE.search(lb)
            if mb:
                d["drop"] = int(mb.group(3)) + int(mb.group(4))
        return d

    def stress_start(self):
        if not self._require_conn() or self.stress_running:
            return
        try:
            dur_s = max(0, int(float(self.ent_st_dur.get() or "0") * 60))
            intv_s = max(1.0, float(self.ent_st_intv.get() or "5"))
            pump_ms = max(50, int(self.ent_st_pms.get() or "200"))
            plen = min(63, max(8, int(self.ent_st_plen.get() or "32")))
        except ValueError:
            messagebox.showerror(APP_TITLE, "Tham số không hợp lệ.")
            return
        pump_mode = {0: None, 1: "rf", 2: "rs485"}[self.cbo_st_pump.current()]
        self.stress_rows = []
        for it in self.tree_st.get_children():
            self.tree_st.delete(it)
        self.lbl_st_verdict.config(text="ĐANG CHẠY...", fg=RUN_FG)
        self.stress_stop_evt.clear()
        self.stress_running = True
        self.btn_st_start.config(state="disabled")
        self.btn_st_stop.config(state="normal")
        threading.Thread(target=self._stress_worker,
                         args=(dur_s, intv_s, pump_mode, pump_ms, plen),
                         daemon=True).start()

    def stress_stop_click(self):
        self.stress_stop_evt.set()

    def _stress_worker(self, dur_s, intv_s, pump_mode, pump_ms, plen):
        t0 = time.time()
        base = self._stress_query()
        if base is None:
            self.root.after(0, lambda: self._stress_done(None, 0, [], 0, 0, "Không đọc được rf stat từ board"))
            return
        prev = dict(base)
        was_up = base["up"]
        down_cnt = 0
        loss_list = []
        pump_n = pump_ok = pump_fail = 0
        mark = self.link.snapshot_len()
        next_pump = time.time()
        next_sample = time.time() + intv_s

        while not self.stress_stop_evt.is_set() and self.link.is_open():
            now = time.time()
            if dur_s > 0 and (now - t0) >= dur_s:
                break
            if pump_mode and now >= next_pump:
                payload = ("S%06d" % pump_n).ljust(plen, "A")[:63]
                self.link.send(("rf tx %s" % payload) if pump_mode == "rf"
                               else ("rs485 %s" % payload))
                pump_n += 1
                next_pump = now + pump_ms / 1000.0
            if now >= next_sample:
                d = self._stress_query()
                next_sample = time.time() + intv_s
                if d:
                    for ln in self.link.lines_since(mark):
                        if ln.startswith("RF_TX=OK"):
                            pump_ok += 1
                        elif ln.startswith("RF_TX=FAIL"):
                            pump_fail += 1
                    mark = self.link.snapshot_len()
                    if was_up and not d["up"]:
                        down_cnt += 1
                    was_up = d["up"]
                    loss_list.append(d["loss"])
                    t_s = int(now - t0)
                    row = (t_s, "UP" if d["up"] else "DOWN", d["loss"], d["redund"],
                           max(0, d["hb_tx"] - prev["hb_tx"]),
                           max(0, d["hb_rx"] - prev["hb_rx"]),
                           max(0, d["crc"] - prev["crc"]),
                           max(0, d["frag"] - prev["frag"]),
                           max(0, d["drop"] - prev["drop"]),
                           pump_ok, pump_fail)
                    self.stress_rows.append(row)
                    prev = dict(d)
                    self.root.after(0, self._stress_gui_update, row, d, base,
                                    down_cnt, loss_list[:], pump_n, pump_ok, pump_fail, t_s)
            time.sleep(0.02)

        d_end = prev
        self.root.after(0, lambda: self._stress_done(
            d_end, down_cnt, loss_list, pump_n, pump_fail,
            None, base=base, pump_ok=pump_ok))

    def _stress_gui_update(self, row, d, base, down_cnt, loss_list, pump_n,
                           pump_ok, pump_fail, t_s):
        self.tree_st.insert("", 0, values=row)
        avg_loss = sum(loss_list) / max(1, len(loss_list))
        self.lbl_st_live.config(text=(
            "Mẫu: %d | %02d:%02d | LINK: %s | Lần rớt link: %d\n"
            "LOSS‰ hiện/tb/max: %d / %.0f / %d | REDUND: %d\n"
            "Tổng ΔCRC: %d | ΔFRAGDROP: %d | ΔDROP queue: %d | "
            "Bơm: %d gửi / %d OK / %d FAIL"
        ) % (len(loss_list), t_s // 60, t_s % 60,
             "UP" if d["up"] else "DOWN", down_cnt,
             d["loss"], avg_loss, max(loss_list), d["redund"],
             max(0, d["crc"] - base["crc"]), max(0, d["frag"] - base["frag"]),
             max(0, d["drop"] - base["drop"]), pump_n, pump_ok, pump_fail))

    def _stress_done(self, d_end, down_cnt, loss_list, pump_n, pump_fail,
                     err, base=None, pump_ok=0):
        self.stress_running = False
        self.btn_st_start.config(state="normal")
        self.btn_st_stop.config(state="disabled")
        if err or d_end is None or not loss_list:
            self.lbl_st_verdict.config(text=err or "KHÔNG CÓ DỮ LIỆU", fg=FAIL_FG)
            return
        avg_loss = sum(loss_list) / len(loss_list)
        max_loss = max(loss_list)
        d_crc = max(0, d_end["crc"] - base["crc"])
        d_frag = max(0, d_end["frag"] - base["frag"])
        d_drop = max(0, d_end["drop"] - base["drop"])
        reasons = []
        if down_cnt > 0:
            reasons.append("rớt link %d lần" % down_cnt)
        if avg_loss > 50:
            reasons.append("LOSS tb %.0f‰ > 50‰" % avg_loss)
        if max_loss > 150:
            reasons.append("LOSS max %d‰ > 150‰" % max_loss)
        if d_crc > 0:
            reasons.append("CRC lỗi %d" % d_crc)
        if d_frag > 0:
            reasons.append("mất mảnh %d" % d_frag)
        if d_drop > 0:
            reasons.append("drop queue %d" % d_drop)
        if pump_n > 0 and pump_fail > 0:
            reasons.append("bơm FAIL %d/%d" % (pump_fail, pump_n))
        if reasons:
            self.lbl_st_verdict.config(text="KHÔNG ĐẠT: " + "; ".join(reasons), fg=FAIL_FG)
        else:
            self.lbl_st_verdict.config(
                text="ĐẠT (LOSS tb %.0f‰, %d mẫu, không rớt link)" % (avg_loss, len(loss_list)),
                fg=PASS_FG)

    def stress_export_csv(self):
        if not self.stress_rows:
            messagebox.showinfo(APP_TITLE, "Chưa có dữ liệu stress test.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV", "*.csv")],
            initialfile="stress_rf_%s.csv" % datetime.now().strftime("%Y%m%d_%H%M%S"))
        if not path:
            return
        with open(path, "w", encoding="utf-8-sig") as f:
            f.write("t_s,link,loss_promille,redund,d_hb_tx,d_hb_rx,d_crc,d_frag,d_drop,pump_ok,pump_fail\n")
            for r in self.stress_rows:
                f.write(",".join(str(x) for x in r) + "\n")
        messagebox.showinfo(APP_TITLE, "Đã lưu: %s" % path)

    def _build_terminal_tab(self, parent):
        self.term_log = tk.Text(parent, bg=TERM_BG, fg=TERM_FG, font=FONT_MONO,
                                 insertbackground=WHITE)
        self.term_log.pack(fill="both", expand=True, padx=8, pady=(8, 4))

        quick = tk.Frame(parent, bg=MAIN_BG)
        quick.pack(fill="x", padx=8, pady=(0, 4))
        for cmd in ["id", "ver", "dip", "flash", "rtc", "rf id", "rf stat",
                     "bridge stat", "bridge log on", "bridge log off"]:
            sec_btn(quick, cmd, command=lambda c=cmd: self._send_and_echo(c),
                    font=FONT_SM).pack(side="left", padx=2, pady=2)

        entry_row = tk.Frame(parent, bg=MAIN_BG)
        entry_row.pack(fill="x", padx=8, pady=(0, 8))
        entry = tk.Entry(entry_row, font=FONT_MONO)
        entry.pack(side="left", fill="x", expand=True)
        self.term_entry = entry

        history = []
        hist_idx = [0]

        def send_entry(event=None):
            cmd = entry.get().strip()
            if not cmd:
                return
            history.append(cmd)
            hist_idx[0] = len(history)
            self._send_and_echo(cmd)
            entry.delete(0, "end")

        def hist_up(event):
            if history and hist_idx[0] > 0:
                hist_idx[0] -= 1
                entry.delete(0, "end")
                entry.insert(0, history[hist_idx[0]])

        def hist_down(event):
            if history and hist_idx[0] < len(history) - 1:
                hist_idx[0] += 1
                entry.delete(0, "end")
                entry.insert(0, history[hist_idx[0]])
            else:
                hist_idx[0] = len(history)
                entry.delete(0, "end")

        entry.bind("<Return>", send_entry)
        entry.bind("<Up>", hist_up)
        entry.bind("<Down>", hist_down)

        flat_btn(entry_row, "Gửi", PRIMARY, PRIMARY_HOV, command=send_entry).pack(side="left", padx=(6, 0))

    def _send_and_echo(self, cmd):
        if not self._require_conn():
            return
        self.term_log.insert("end", "> %s\n" % cmd)
        self.term_log.see("end")
        self.link.send(cmd)


def main():
    if "--modbus" in sys.argv:
        import modbus_poll_app
        modbus_poll_app.main()
        return

    root = tk.Tk()
    app = MbwTestApp(root)

    def on_close():
        if getattr(app, "_xlog_on", False):  # dang ghi log -> luu not phan dem
            app._xlog_on = False
            try:
                app._fwd_log_flush(final=True)
            except Exception:
                pass
        app.link.disconnect()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()


if __name__ == "__main__":
    main()
