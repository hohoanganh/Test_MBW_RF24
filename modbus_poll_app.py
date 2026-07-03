#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
modbus_poll_app.py - App Test Modbus RTU "kieu Modbus Poll" (tich hop cho MBW RF24 RS485 2.0)
------------------------------------------------------------
Sao chep tu _reference/App_Test_Mobus_Pull/modbus_pull_app.py (giu nguyen logic,
chi doi ten file + tieu de) de tich hop lam CONG CU MASTER MODBUS RTU THAT, dung
kem mbw_test_app.py: mo cua so nay tu nut "Modbus Poll Test" trong tab "Giam sat
Forward" (giong cach opms_test_app.py mo ch348_test_app.py - chay lai chinh file
nay voi co --modbus, tien trinh/cua so rieng).

- Ket noi RS485/Modbus RTU qua cong COM THAT cua may (KHONG phai console CLI cua
  board MBW) - dung de bom/doc du lieu Modbus RTU that xuyen qua cau RS485<->
  Wireless: 1 may chay app nay o vai Modbus MASTER noi vao cong RS485 vat ly cua
  Board A, may kia (hoac 1 thiet bi Modbus Slave that) noi vao cong RS485 vat ly
  cua Board B - day la phep test THUC TE NHAT cho chuc nang chinh cua san pham.
- Chi dung pyserial, tu cai dat giao thuc Modbus RTU FC3/FC6 + CRC16, khong phu
  thuoc thu vien pymodbus.
- Doc lien tuc (poll) danh sach thanh ghi theo "Register Map" cau hinh san
  (mac dinh theo cam bien nhiet do/do am ES35-SW cua EPCB - co the sua lai
  DEFAULT_REGISTER_MAP cho dung thiet bi Modbus that dang test).
- Hien thi bang du lieu + do thi realtime.
- Ghi log ra file Excel (.xlsx) theo thoi gian thuc bang openpyxl.
- Cho phep ghi (Write Single Register - FC6) cac thanh ghi Read/Write
  (Address ID, Baudrate, Temperature Correction, Humidity Correction).

Tac gia: Claude (Cowork) cho Hoang Anh - EPCB Vietnam
"""

import os
import sys
import time
import queue
import threading
import struct
from datetime import datetime

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog

try:
    import serial
    from serial.tools import list_ports
except ImportError:
    serial = None
    list_ports = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

try:
    import matplotlib
    matplotlib.use("TkAgg")
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

# Dong bo giao dien voi mbw_test_app.py: dung chung theme (mau/font/nut) tu
# mbw_theme.py thay vi theme "clam" mac dinh cua ttk.
from mbw_theme import (
    MAIN_BG, WHITE, CARD_BD, HDR_BG, HDR_SUB, PRIMARY, PRIMARY_HOV,
    SEC_BG, SEC_HOV, SEC_TX, INK, PASS_FG, FAIL_FG, RUN_FG, WARN_FG, DIS_FG,
    TERM_BG, TERM_FG, TERM_TX, FONT, FONT_B, FONT_SM, FONT_CARD, FONT_MONO,
    flat_btn, sec_btn,
)

APP_TITLE = "MBW RF24 RS485 2.0 - Modbus Poll Test (RS485 that)"
APP_VERSION = "1.0"

# ----------------------------------------------------------------------
# 1) REGISTER MAP - co the chinh sua de dung cho thiet bi Modbus RTU khac
#    type: "int16" | "uint16" | "int32" | "uint32" | "float32"
#    scale: gia tri thuc = raw / scale
#    rw: "R" (read only, FC3) hoac "RW" (read/write, FC3 + FC6)
# ----------------------------------------------------------------------
DEFAULT_REGISTER_MAP = [
    {"name": "Temperature",           "address": 0,   "type": "int16",  "scale": 10, "unit": "°C",  "rw": "R"},
    {"name": "Humidity",              "address": 1,   "type": "int16",  "scale": 10, "unit": "%RH", "rw": "R"},
    {"name": "Address Device (ID)",   "address": 100, "type": "uint16", "scale": 1,  "unit": "",    "rw": "RW"},
    {"name": "Baudrate",              "address": 101, "type": "uint16", "scale": 1,  "unit": "code","rw": "RW"},
    {"name": "Parity",                "address": 102, "type": "uint16", "scale": 1,  "unit": "",    "rw": "R"},
    {"name": "DataBits",              "address": 103, "type": "uint16", "scale": 1,  "unit": "bit", "rw": "R"},
    {"name": "StopBits",              "address": 104, "type": "uint16", "scale": 1,  "unit": "bit", "rw": "R"},
    {"name": "Setting Mode",          "address": 105, "type": "uint16", "scale": 1,  "unit": "",    "rw": "R"},
    {"name": "Temperature Correction","address": 106, "type": "int16",  "scale": 1,  "unit": "°C",  "rw": "RW"},
    {"name": "Humidity Correction",   "address": 107, "type": "int16",  "scale": 1,  "unit": "%RH", "rw": "RW"},
]

BAUDRATE_CODE_MAP = {
    2: 4800, 3: 9600, 4: 14400, 5: 19200,
    6: 38400, 7: 56000, 8: 57600, 9: 115200,
}
BAUDRATE_LIST = [4800, 9600, 14400, 19200, 38400, 56000, 57600, 115200]


# ----------------------------------------------------------------------
# 2) MODBUS RTU MASTER (thuan pyserial + CRC16, khong dung pymodbus)
# ----------------------------------------------------------------------
def crc16_modbus(data: bytes) -> bytes:
    """Tinh CRC16 chuan Modbus RTU. Tra ve 2 byte little-endian (Lo, Hi)."""
    crc = 0xFFFF
    for b in data:
        crc ^= b
        for _ in range(8):
            if crc & 0x0001:
                crc >>= 1
                crc ^= 0xA001
            else:
                crc >>= 1
    return struct.pack("<H", crc)


class ModbusError(Exception):
    pass


class ModbusTimeoutError(ModbusError):
    pass


class ModbusCRCError(ModbusError):
    pass


class ModbusExceptionResponse(ModbusError):
    def __init__(self, function_code, exception_code):
        self.function_code = function_code
        self.exception_code = exception_code
        super().__init__(
            "Modbus exception resp: FC=0x%02X code=0x%02X" % (function_code, exception_code)
        )


class ModbusRTUMaster:
    """Modbus RTU master toi gian: ho tro FC3 (Read Holding Registers)
    va FC6 (Write Single Register). Tu quan ly ket noi serial."""

    def __init__(self, log_callback=None):
        self.ser = None
        self.lock = threading.Lock()
        self.log_callback = log_callback  # callback(direction, raw_bytes)

    def connect(self, port, baudrate=9600, parity="N", databits=8, stopbits=1, timeout=1.0):
        if serial is None:
            raise RuntimeError("Thu vien 'pyserial' chua duoc cai dat. Chay: pip install pyserial")
        parity_map = {"N": serial.PARITY_NONE, "E": serial.PARITY_EVEN, "O": serial.PARITY_ODD}
        stopbits_map = {1: serial.STOPBITS_ONE, 2: serial.STOPBITS_TWO}
        self.ser = serial.Serial(
            port=port,
            baudrate=baudrate,
            bytesize=serial.EIGHTBITS if databits == 8 else serial.SEVENBITS,
            parity=parity_map.get(parity, serial.PARITY_NONE),
            stopbits=stopbits_map.get(stopbits, serial.STOPBITS_ONE),
            timeout=timeout,
            write_timeout=timeout,
        )

    def disconnect(self):
        if self.ser and self.ser.is_open:
            try:
                self.ser.close()
            except Exception:
                pass
        self.ser = None

    @property
    def is_connected(self):
        return self.ser is not None and self.ser.is_open

    def _log(self, direction, raw: bytes):
        if self.log_callback:
            hex_str = " ".join("%02X" % b for b in raw)
            self.log_callback(direction, hex_str)

    def _transact(self, request: bytes, expected_min_len=5, retries=1):
        """Gui request va nhan response, co retry, tra ve payload bytes day du."""
        if not self.is_connected:
            raise ModbusError("Chua ket noi cong COM")
        last_err = None
        for attempt in range(retries + 1):
            with self.lock:
                try:
                    self.ser.reset_input_buffer()
                    self.ser.write(request)
                    self._log("TX", request)
                    # doc toi thieu 3 byte, roi doc them theo do dai da khai bao
                    resp = self.ser.read(3)
                    if len(resp) < 3:
                        raise ModbusTimeoutError("Khong nhan duoc phan hoi tu thiet bi (timeout)")
                    slave_addr, func_code = resp[0], resp[1]
                    if func_code & 0x80:
                        # exception response: slave, func|0x80, exception_code, crc(2)
                        rest = self.ser.read(2)
                        full = resp + rest
                        self._log("RX", full)
                        exc_code = resp[2]
                        raise ModbusExceptionResponse(func_code, exc_code)
                    if func_code == 0x03:
                        byte_count = resp[2]
                        rest = self.ser.read(byte_count + 2)
                        full = resp + rest
                        if len(full) < 3 + byte_count + 2:
                            raise ModbusTimeoutError("Phan hoi thieu du lieu (timeout)")
                        self._log("RX", full)
                        if crc16_modbus(full[:-2]) != full[-2:]:
                            raise ModbusCRCError("Sai CRC trong phan hoi")
                        return full
                    elif func_code == 0x06:
                        rest = self.ser.read(5)
                        full = resp + rest
                        if len(full) < 8:
                            raise ModbusTimeoutError("Phan hoi thieu du lieu (timeout)")
                        self._log("RX", full)
                        if crc16_modbus(full[:-2]) != full[-2:]:
                            raise ModbusCRCError("Sai CRC trong phan hoi")
                        return full
                    else:
                        rest = self.ser.read(64)
                        full = resp + rest
                        self._log("RX", full)
                        raise ModbusError("Function code khong ho tro: 0x%02X" % func_code)
                except (ModbusTimeoutError, ModbusCRCError) as e:
                    last_err = e
                    continue
        raise last_err if last_err else ModbusError("Loi khong xac dinh")

    def read_holding_registers(self, slave_id, address, count, retries=1):
        req = bytearray()
        req.append(slave_id & 0xFF)
        req.append(0x03)
        req += struct.pack(">H", address)
        req += struct.pack(">H", count)
        req += crc16_modbus(bytes(req))
        resp = self._transact(bytes(req), retries=retries)
        byte_count = resp[2]
        data = resp[3:3 + byte_count]
        values = list(struct.unpack(">%dH" % (byte_count // 2), data))
        return values

    def write_single_register(self, slave_id, address, value, retries=1):
        req = bytearray()
        req.append(slave_id & 0xFF)
        req.append(0x06)
        req += struct.pack(">H", address)
        req += struct.pack(">H", value & 0xFFFF)
        req += crc16_modbus(bytes(req))
        resp = self._transact(bytes(req), retries=retries)
        return resp


def raw_to_value(raw_words, reg_type, scale):
    """Chuyen danh sach word (uint16) tho thanh gia tri theo kieu du lieu."""
    if reg_type == "uint16":
        v = raw_words[0]
    elif reg_type == "int16":
        v = raw_words[0]
        if v >= 0x8000:
            v -= 0x10000
    elif reg_type in ("uint32", "int32", "float32"):
        combined = (raw_words[0] << 16) | raw_words[1]
        if reg_type == "uint32":
            v = combined
        elif reg_type == "int32":
            v = combined - 0x100000000 if combined >= 0x80000000 else combined
        else:  # float32
            v = struct.unpack(">f", struct.pack(">I", combined))[0]
            return round(v, 4)
    else:
        v = raw_words[0]
    if scale and scale != 1:
        return round(v / scale, 4)
    return v


def reg_word_count(reg_type):
    return 2 if reg_type in ("uint32", "int32", "float32") else 1


# ----------------------------------------------------------------------
# 3) EXCEL LOGGER - ghi log realtime bang openpyxl
#    - Sheet "AllData": moi dong = 1 lan poll, cot Timestamp + tat ca thanh ghi
#    - Sheet rieng cho tung thanh ghi (giong Insight Sensor: Temperature/Humidity)
# ----------------------------------------------------------------------
class ExcelLogger:
    HEADER_FILL = "1F4E78"
    HEADER_FONT_COLOR = "FFFFFF"

    COMM_LOG_MAX_ROWS = 50000  # gioi han an toan, tranh file Excel qua nang khi log rat lau

    def __init__(self, filepath, register_map):
        self.filepath = filepath
        self.register_map = register_map
        self.wb = None
        self.ws_all = None
        self.ws_by_name = {}
        self.ws_comm = None
        self.comm_row_count = 0
        self.comm_log_full_warned = False
        self.pending_rows = 0
        self.lock = threading.Lock()

    def _style_header(self, ws, ncols):
        if Workbook is None:
            return
        fill = PatternFill(start_color=self.HEADER_FILL, end_color=self.HEADER_FILL, fill_type="solid")
        font = Font(color=self.HEADER_FONT_COLOR, bold=True)
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")

    def create_new(self):
        if Workbook is None:
            raise RuntimeError("Thu vien 'openpyxl' chua duoc cai dat. Chay: pip install openpyxl")
        self.wb = Workbook()
        self.ws_all = self.wb.active
        self.ws_all.title = "AllData"
        headers = ["Date", "Time"] + [
            "%s (%s)" % (r["name"], r["unit"]) if r["unit"] else r["name"]
            for r in self.register_map
        ]
        self.ws_all.append(headers)
        self._style_header(self.ws_all, len(headers))
        self.ws_all.column_dimensions["A"].width = 12
        self.ws_all.column_dimensions["B"].width = 12
        for idx in range(len(headers) - 2):
            self.ws_all.column_dimensions[get_column_letter(idx + 3)].width = 20

        # sheet rieng cho tung thanh ghi so (de tien ve bieu do, giong app mau)
        self.ws_by_name = {}
        for r in self.register_map:
            safe_name = r["name"][:28]
            ws = self.wb.create_sheet(title=safe_name)
            unit_label = r["unit"] if r["unit"] else "Value"
            ws.append(["Date", "Time", "Value", "Unit"])
            self._style_header(ws, 4)
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 10
            self.ws_by_name[r["name"]] = ws

        # sheet rieng ghi lai tung frame giao tiep TX/RX (giong khung Data Log trong app)
        self.ws_comm = self.wb.create_sheet(title="CommLog")
        self.ws_comm.append(["Date", "Time", "Direction", "Frame (HEX)"])
        self._style_header(self.ws_comm, 4)
        self.ws_comm.column_dimensions["A"].width = 12
        self.ws_comm.column_dimensions["B"].width = 14
        self.ws_comm.column_dimensions["C"].width = 10
        self.ws_comm.column_dimensions["D"].width = 55
        self.comm_row_count = 0
        self.comm_log_full_warned = False

        self.save()

    def open_existing_or_create(self):
        if os.path.exists(self.filepath):
            try:
                self.wb = load_workbook(self.filepath)
                self.ws_all = self.wb["AllData"] if "AllData" in self.wb.sheetnames else self.wb.active
                self.ws_by_name = {}
                for r in self.register_map:
                    name = r["name"][:28]
                    if name in self.wb.sheetnames:
                        self.ws_by_name[r["name"]] = self.wb[name]
                    else:
                        ws = self.wb.create_sheet(title=name)
                        ws.append(["Date", "Time", "Value", "Unit"])
                        self._style_header(ws, 4)
                        self.ws_by_name[r["name"]] = ws

                if "CommLog" in self.wb.sheetnames:
                    self.ws_comm = self.wb["CommLog"]
                    self.comm_row_count = max(0, self.ws_comm.max_row - 1)
                else:
                    self.ws_comm = self.wb.create_sheet(title="CommLog")
                    self.ws_comm.append(["Date", "Time", "Direction", "Frame (HEX)"])
                    self._style_header(self.ws_comm, 4)
                    self.comm_row_count = 0
                self.comm_log_full_warned = False
                return
            except Exception:
                pass
        self.create_new()

    def append_row(self, timestamp: datetime, values: dict):
        """values: dict {reg_name: value}. Ghi vao bo nho, chua save ngay."""
        with self.lock:
            date_str = timestamp.strftime("%d/%m/%Y")
            time_str = timestamp.strftime("%H:%M:%S")
            row = [date_str, time_str]
            for r in self.register_map:
                row.append(values.get(r["name"], ""))
            self.ws_all.append(row)

            for r in self.register_map:
                ws = self.ws_by_name.get(r["name"])
                if ws is not None and r["name"] in values:
                    ws.append([date_str, time_str, values[r["name"]], r["unit"]])
            self.pending_rows += 1

    def append_comm(self, timestamp: datetime, direction: str, hex_str: str):
        """Ghi 1 dong frame TX/RX vao sheet CommLog. Tra ve False neu da dat
        gioi han an toan COMM_LOG_MAX_ROWS (de tranh file Excel phinh to qua muc)."""
        with self.lock:
            if self.ws_comm is None:
                return True
            if self.comm_row_count >= self.COMM_LOG_MAX_ROWS:
                return False
            date_str = timestamp.strftime("%d/%m/%Y")
            time_str = timestamp.strftime("%H:%M:%S.%f")[:-3]
            self.ws_comm.append([date_str, time_str, direction, hex_str])
            self.comm_row_count += 1
            self.pending_rows += 1
            return True

    def save(self):
        with self.lock:
            if self.wb is None:
                return
            try:
                self.wb.save(self.filepath)
                self.pending_rows = 0
            except PermissionError as e:
                # Khong reset pending_rows: du lieu van giu trong bo nho, lan flush
                # ke tiep se thu ghi lai toan bo (khong mat du lieu).
                raise PermissionError(
                    "Khong ghi duoc file '%s' (Permission denied). File co the dang mo trong Excel, "
                    "bi khoa boi chuong trinh khac, hoac khong co quyen ghi vao thu muc nay. "
                    "Du lieu van dang duoc giu trong bo nho va se tu dong thu luu lai o lan sau."
                    % os.path.basename(self.filepath)
                ) from e
            except OSError as e:
                raise OSError(
                    "Loi he thong file khi luu '%s': %s. Du lieu van dang giu trong bo nho."
                    % (os.path.basename(self.filepath), e)
                ) from e

    def flush_if_needed(self, min_rows=1):
        if self.pending_rows >= min_rows:
            self.save()


# ----------------------------------------------------------------------
# 4) GUI APPLICATION
# ----------------------------------------------------------------------
class ModbusPollApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE + "  v" + APP_VERSION)
        self.geometry("1180x760")
        self.minsize(980, 640)
        self.configure(bg=MAIN_BG)

        self.register_map = [dict(r) for r in DEFAULT_REGISTER_MAP]
        self.master_modbus = ModbusRTUMaster(log_callback=self._on_modbus_log)

        self.gui_queue = queue.Queue()
        self.poll_thread = None
        self.poll_running = threading.Event()
        self.auto_scan_thread = None
        self.auto_scan_cancel = threading.Event()

        self.excel_logger = None
        self.logging_enabled = tk.BooleanVar(value=False)
        self.logging_active = False
        self.log_filepath_var = tk.StringVar(value="")
        self._last_excel_error_ts = 0.0
        self._excel_error_throttle_sec = 5.0

        self.history_max_points = 300
        self.history_time = []
        self.history_values = {r["name"]: [] for r in self.register_map if r["unit"] in ("°C", "%RH")}

        self._build_style()
        self._build_ui()
        self._refresh_ports()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.after(100, self._process_gui_queue)

    # ------------------------------------------------------------------
    # UI BUILD
    # ------------------------------------------------------------------
    def _build_style(self):
        """Dong bo style ttk voi mbw_test_app.py: cung bang mau/font Industrial
        Flat (nen sang #F4F6F8, card trang, header xanh than, Segoe UI)."""
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure(".", background=MAIN_BG, foreground=INK, font=FONT)
        style.configure("TFrame", background=MAIN_BG)
        style.configure("TLabel", background=MAIN_BG, foreground=INK, font=FONT)
        style.configure("TLabelframe", background=MAIN_BG, foreground=INK, font=FONT_CARD)
        style.configure("TLabelframe.Label", background=MAIN_BG, foreground=INK, font=FONT_CARD)
        style.configure("TCheckbutton", background=MAIN_BG, foreground=INK, font=FONT)
        style.configure("TEntry", font=FONT_SM)
        style.configure("TCombobox", font=FONT_SM)
        style.configure("TSpinbox", font=FONT_SM)
        style.configure("TButton", font=FONT)

        style.configure("TNotebook", background=MAIN_BG, borderwidth=0, tabmargins=(2, 4, 2, 0))
        style.configure("TNotebook.Tab", background=SEC_BG, foreground=SEC_TX,
                        font=FONT, padding=(14, 6), borderwidth=0)
        style.map("TNotebook.Tab",
                  background=[("selected", PRIMARY)],
                  foreground=[("selected", WHITE)],
                  expand=[("selected", (1, 1, 1, 0))])

        # QUAN TRONG: theme "clam" mac dinh ve them 1 "focus ring" (vien net
        # dut) rieng BEN TRONG tab moi khi tab duoc bam/focus - phan tu nay
        # chiem thêm padding tam thoi lam tab bi "nhay"/lech vi tri ngay luc
        # bam vao (dung y loi user bao cao). Ghi de lai layout TNotebook.Tab
        # bo hang "Notebook.focus" di la het loi, khong anh huong gi khac.
        style.layout("TNotebook.Tab", [
            ("Notebook.tab", {
                "sticky": "nswe",
                "children": [
                    ("Notebook.padding", {
                        "side": "top",
                        "sticky": "nswe",
                        "children": [
                            ("Notebook.label", {"side": "top", "sticky": ""})
                        ],
                    })
                ],
            })
        ])

        style.configure("Treeview", rowheight=24, font=FONT, background=WHITE,
                        fieldbackground=WHITE, foreground=INK, borderwidth=0)
        style.configure("Treeview.Heading", font=FONT_CARD, background=HDR_BG,
                        foreground=WHITE, relief="flat")
        style.map("Treeview.Heading", background=[("active", HDR_BG)])
        style.map("Treeview", background=[("selected", PRIMARY)],
                  foreground=[("selected", WHITE)])

        # style rieng cho card nen trang (thanh ket noi)
        style.configure("Card.TFrame", background=WHITE)
        style.configure("Card.TLabel", background=WHITE, foreground=INK, font=FONT_SM)

    def _build_header(self):
        hdr = tk.Frame(self, bg=HDR_BG, height=56)
        hdr.pack(fill=tk.X, side=tk.TOP)
        tk.Label(hdr, text="MBW RF24 RS485 2.0 - Modbus Poll Test", bg=HDR_BG, fg=WHITE,
                  font=("Segoe UI", 14, "bold")).pack(side=tk.LEFT, padx=16, pady=10)
        tk.Label(hdr, text="Modbus RTU Master thật qua cổng RS485 vật lý  |  v%s" % APP_VERSION,
                  bg=HDR_BG, fg=HDR_SUB, font=FONT_SM).pack(side=tk.LEFT, padx=4)

    def _build_ui(self):
        self._build_header()
        self._build_connection_bar()

        main_pane = ttk.PanedWindow(self, orient=tk.VERTICAL)
        main_pane.pack(fill=tk.BOTH, expand=True, padx=8, pady=(4, 8))

        top_frame = ttk.Frame(main_pane)
        main_pane.add(top_frame, weight=3)

        notebook = ttk.Notebook(top_frame)
        notebook.pack(fill=tk.BOTH, expand=True)

        self.tab_data = ttk.Frame(notebook)
        notebook.add(self.tab_data, text="Bang du lieu (Data Table)")
        self._build_data_tab(self.tab_data)

        self.tab_chart = ttk.Frame(notebook)
        notebook.add(self.tab_chart, text="Bieu do (Chart)")
        self._build_chart_tab(self.tab_chart)

        self.tab_settings = ttk.Frame(notebook)
        notebook.add(self.tab_settings, text="Ghi thanh ghi (Write Register)")
        self._build_settings_tab(self.tab_settings)

        bottom_frame = ttk.Frame(main_pane)
        main_pane.add(bottom_frame, weight=2)
        self._build_bottom_panel(bottom_frame)

        self.status_var = tk.StringVar(value="San sang. Chua ket noi.")
        status_bar = tk.Label(self, textvariable=self.status_var, bg=SEC_BG, fg=SEC_TX,
                               font=FONT_SM, anchor="w", padx=8, pady=3)
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)

    def _build_connection_bar(self):
        card = tk.Frame(self, bg=WHITE, highlightbackground=CARD_BD, highlightthickness=1)
        card.pack(fill=tk.X, padx=12, pady=8)

        tk.Label(card, text="Kết nối Modbus RTU (RS485)", bg=WHITE, fg=INK, font=FONT_CARD).grid(
            row=0, column=0, columnspan=8, sticky="w", padx=10, pady=(8, 2))

        tk.Label(card, text="Cổng (Port):", bg=WHITE, fg=INK, font=FONT_SM).grid(
            row=1, column=0, padx=(10, 4), pady=(0, 10), sticky="e")
        self.port_var = tk.StringVar()
        # rong 32 ky tu + kem MO TA thiet bi (dong bo voi app chinh
        # mbw_test_app.py) de chon dung cong khi may co nhieu COM
        self.port_combo = ttk.Combobox(card, textvariable=self.port_var, width=32,
                                        state="readonly", font=FONT_SM)
        self.port_combo.grid(row=1, column=1, padx=4, pady=(0, 10))
        sec_btn(card, "↻", command=self._refresh_ports, width=3).grid(
            row=1, column=2, padx=2, pady=(0, 10))

        tk.Label(card, text="Baudrate:", bg=WHITE, fg=INK, font=FONT_SM).grid(
            row=1, column=3, padx=(16, 4), pady=(0, 10), sticky="e")
        self.baud_var = tk.IntVar(value=9600)
        ttk.Combobox(card, textvariable=self.baud_var, width=9, state="readonly",
                     values=BAUDRATE_LIST, font=FONT_SM).grid(row=1, column=4, padx=4, pady=(0, 10))

        tk.Label(card, text="Parity:", bg=WHITE, fg=INK, font=FONT_SM).grid(
            row=1, column=5, padx=(16, 4), pady=(0, 10), sticky="e")
        self.parity_var = tk.StringVar(value="N")
        ttk.Combobox(card, textvariable=self.parity_var, width=4, state="readonly",
                     values=["N", "E", "O"], font=FONT_SM).grid(row=1, column=6, padx=4, pady=(0, 10))

        tk.Label(card, text="Slave ID:", bg=WHITE, fg=INK, font=FONT_SM).grid(
            row=1, column=7, padx=(16, 4), pady=(0, 10), sticky="e")
        self.slave_id_var = tk.IntVar(value=1)
        ttk.Spinbox(card, from_=1, to=247, textvariable=self.slave_id_var, width=6,
                    font=FONT_SM).grid(row=1, column=8, padx=4, pady=(0, 10))

        tk.Label(card, text="Scan rate (ms):", bg=WHITE, fg=INK, font=FONT_SM).grid(
            row=1, column=9, padx=(16, 4), pady=(0, 10), sticky="e")
        self.scan_rate_var = tk.IntVar(value=1000)
        ttk.Spinbox(card, from_=200, to=60000, increment=100, textvariable=self.scan_rate_var,
                    width=8, font=FONT_SM).grid(row=1, column=10, padx=4, pady=(0, 10))

        self.connect_btn = flat_btn(card, "Connect", PRIMARY, PRIMARY_HOV,
                                     command=self._toggle_connect)
        self.connect_btn.grid(row=1, column=11, padx=(16, 4), pady=(0, 10))

        self.auto_scan_btn = sec_btn(card, "Auto Scan", command=self._toggle_auto_scan)
        self.auto_scan_btn.grid(row=1, column=12, padx=4, pady=(0, 10))

        self.conn_indicator = tk.Canvas(card, width=16, height=16, bg=WHITE, highlightthickness=0)
        self.conn_indicator.grid(row=1, column=13, padx=(12, 4), pady=(0, 10))
        self._draw_indicator(DIS_FG)

        card.grid_columnconfigure(14, weight=1)

    def _build_data_tab(self, parent):
        columns = ("stt", "name", "address", "value", "unit", "rw")
        self.tree = ttk.Treeview(parent, columns=columns, show="headings", height=12)
        headers = {"stt": "#", "name": "Description", "address": "Address",
                   "value": "Value", "unit": "Unit", "rw": "Read/Write"}
        widths = {"stt": 40, "name": 240, "address": 90, "value": 120, "unit": 80, "rw": 100}
        for c in columns:
            self.tree.heading(c, text=headers[c])
            self.tree.column(c, width=widths[c], anchor="center" if c != "name" else "w")
        self.tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8, side=tk.LEFT)

        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.LEFT, fill=tk.Y, pady=8)

        for idx, r in enumerate(self.register_map, start=1):
            self.tree.insert("", "end", iid=r["name"], values=(idx, r["name"], r["address"], "-", r["unit"], r["rw"]))

        self.tree.bind("<Double-1>", self._on_tree_double_click)

    def _build_chart_tab(self, parent):
        if not HAS_MPL:
            tk.Label(parent, text="Thu vien matplotlib chua duoc cai dat. Chay: pip install matplotlib",
                      bg=MAIN_BG, fg=FAIL_FG, font=FONT).pack(padx=20, pady=20)
            return
        self.fig = Figure(figsize=(6, 4), dpi=90)
        self.ax_temp = self.fig.add_subplot(111)
        self.ax_hum = self.ax_temp.twinx()
        self.ax_temp.set_ylabel("Temperature (°C)", color="tab:green")
        self.ax_hum.set_ylabel("Humidity (%RH)", color="tab:blue")
        self.ax_temp.set_xlabel("Time")
        self.line_temp, = self.ax_temp.plot([], [], color="tab:green", label="Temperature")
        self.line_hum, = self.ax_hum.plot([], [], color="tab:blue", label="Humidity")
        self.fig.tight_layout()

        self.canvas = FigureCanvasTkAgg(self.fig, master=parent)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

    def _build_settings_tab(self, parent):
        info = tk.Label(
            parent,
            text=("Ghi gia tri xuong thanh ghi Read/Write (Function 06). "
                  "Chi ap dung khi thiet bi da o che do Setting Mode = 0 (Software)."),
            wraplength=1000, bg=MAIN_BG, fg=SEC_TX, font=FONT_SM, justify="left"
        )
        info.pack(anchor="w", padx=10, pady=(10, 4))

        frame = tk.Frame(parent, bg=MAIN_BG)
        frame.pack(fill=tk.X, padx=10, pady=10)

        writable = [r for r in self.register_map if r["rw"] == "RW"]
        self.write_entries = {}
        for i, r in enumerate(writable):
            tk.Label(frame, text="%s (reg %d):" % (r["name"], r["address"]), bg=MAIN_BG,
                      fg=INK, font=FONT).grid(row=i, column=0, sticky="e", padx=6, pady=6)
            var = tk.StringVar()
            entry = ttk.Entry(frame, textvariable=var, width=12, font=FONT_SM)
            entry.grid(row=i, column=1, padx=6, pady=6)
            self.write_entries[r["name"]] = var
            sec_btn(frame, "Apply", command=lambda reg=r, v=var: self._write_register(reg, v)).grid(
                row=i, column=2, padx=6, pady=6)

        hint = tk.Label(
            frame,
            text="Baudrate code: 2=4800, 3=9600, 4=14400, 5=19200, 6=38400, 7=56000, 8=57600, 9=115200",
            bg=MAIN_BG, fg=DIS_FG, font=FONT_SM
        )
        hint.grid(row=len(writable), column=0, columnspan=3, sticky="w", padx=6, pady=(10, 0))

    def _build_bottom_panel(self, parent):
        notebook = ttk.Notebook(parent)
        notebook.pack(fill=tk.BOTH, expand=True)

        # --- tab log TX/RX ---
        tab_log = tk.Frame(notebook, bg=MAIN_BG)
        notebook.add(tab_log, text="Data Log (TX/RX)")
        self.log_text = tk.Text(tab_log, height=10, font=FONT_MONO, bg=TERM_BG, fg=TERM_FG,
                                 insertbackground=TERM_TX)
        self.log_text.pack(fill=tk.BOTH, expand=True, side=tk.LEFT, padx=(8, 0), pady=8)
        log_scroll = ttk.Scrollbar(tab_log, command=self.log_text.yview)
        log_scroll.pack(side=tk.LEFT, fill=tk.Y, pady=8)
        self.log_text.configure(yscrollcommand=log_scroll.set, state="disabled")
        sec_btn(tab_log, "Xóa log", command=self._clear_log).pack(side=tk.LEFT, padx=8, anchor="n", pady=8)

        # --- tab excel logging ---
        tab_excel = ttk.Frame(notebook)
        notebook.add(tab_excel, text="Ghi log Excel")
        self._build_excel_tab(tab_excel)

    def _build_excel_tab(self, parent):
        row0 = tk.Frame(parent, bg=MAIN_BG)
        row0.pack(fill=tk.X, padx=10, pady=10)

        tk.Label(row0, text="File Excel:", bg=MAIN_BG, fg=INK, font=FONT).grid(
            row=0, column=0, sticky="e", padx=4)
        entry = ttk.Entry(row0, textvariable=self.log_filepath_var, width=70, font=FONT_SM)
        entry.grid(row=0, column=1, padx=4)
        sec_btn(row0, "Chọn file...", command=self._choose_excel_file).grid(row=0, column=2, padx=4)

        row1 = tk.Frame(parent, bg=MAIN_BG)
        row1.pack(fill=tk.X, padx=10, pady=4)
        self.start_log_btn = flat_btn(row1, "Bắt đầu ghi log (Start Logging)", PRIMARY, PRIMARY_HOV,
                                        command=self._toggle_logging)
        self.start_log_btn.grid(row=0, column=0, padx=4)
        sec_btn(row1, "Export snapshot Excel", command=self._export_snapshot).grid(row=0, column=1, padx=4)

        self.log_status_var = tk.StringVar(value="Chưa ghi log.")
        tk.Label(row1, textvariable=self.log_status_var, bg=MAIN_BG, fg=SEC_TX, font=FONT_SM).grid(
            row=0, column=2, padx=16)

        hint = tk.Label(
            parent,
            text=("Khi bat dau ghi log, moi gia tri doc duoc trong qua trinh poll se duoc them vao file Excel "
                  "(sheet AllData + sheet rieng cho tung thong so), tu dong luu dinh ky. "
                  "Tat ca frame TX/RX giao tiep voi thiet bi (giong khung Data Log ben tren) cung duoc ghi "
                  "vao sheet 'CommLog' trong cung file. Neu file bi khoa (VD: dang mo bang Excel) khi luu, "
                  "du lieu se KHONG bi mat - app se tu dong thu luu lai o lan doc tiep theo."),
            bg=MAIN_BG, fg=DIS_FG, font=FONT_SM, wraplength=1000, justify="left"
        )
        hint.pack(anchor="w", padx=10, pady=(4, 10))

    # ------------------------------------------------------------------
    # HELPERS - UI
    # ------------------------------------------------------------------
    def _draw_indicator(self, color):
        self.conn_indicator.delete("all")
        self.conn_indicator.create_oval(2, 2, 14, 14, fill=color, outline="")

    def _refresh_ports(self):
        # "COM9 - USB-SERIAL CH340" (dong bo voi app chinh) - khi ket noi se
        # tach lay ten cong bang _selected_port()
        ports = []
        if list_ports is not None:
            for p in list_ports.comports():
                desc = (p.description or "").strip()
                ports.append("%s - %s" % (p.device, desc) if desc else p.device)
        self.port_combo["values"] = ports
        if ports and not self.port_var.get():
            self.port_var.set(ports[0])

    def _selected_port(self):
        """Ten cong thuc (COM9) tu chuoi hien thi 'COM9 - USB-SERIAL CH340'."""
        return self.port_var.get().split(" - ")[0].strip()

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", tk.END)
        self.log_text.configure(state="disabled")

    # ------------------------------------------------------------------
    # MODBUS LOG CALLBACK (chay tren thread poll -> day vao queue)
    # ------------------------------------------------------------------
    def _on_modbus_log(self, direction, hex_str):
        ts_dt = datetime.now()
        ts = ts_dt.strftime("%H:%M:%S.%f")[:-3]
        self.gui_queue.put(("log", "[%s] %s: %s" % (ts, direction, hex_str)))

        # Neu dang bat Ghi log Excel, luu them frame nay vao sheet CommLog
        if self.logging_active and self.excel_logger is not None:
            try:
                still_ok = self.excel_logger.append_comm(ts_dt, direction, hex_str)
                if not still_ok and not self.excel_logger.comm_log_full_warned:
                    self.excel_logger.comm_log_full_warned = True
                    self.gui_queue.put(("error",
                        "CommLog da dat gioi han %d dong, ngung ghi them frame TX/RX vao Excel "
                        "(du lieu Temperature/Humidity van tiep tuc duoc ghi binh thuong)."
                        % self.excel_logger.COMM_LOG_MAX_ROWS))
            except Exception as e:
                self.gui_queue.put(("error", "Ghi CommLog loi: %s" % e))

    def _process_gui_queue(self):
        try:
            while True:
                kind, payload = self.gui_queue.get_nowait()
                if kind == "log":
                    self._append_log_line(payload)
                elif kind == "data":
                    self._update_data_table(payload)
                elif kind == "status":
                    self.status_var.set(payload)
                elif kind == "error":
                    self.status_var.set("Loi: %s" % payload)
                elif kind == "scan_progress":
                    self.status_var.set(payload)
                elif kind == "scan_done":
                    self._on_auto_scan_done(payload)
        except queue.Empty:
            pass
        self.after(80, self._process_gui_queue)

    def _append_log_line(self, line):
        self.log_text.configure(state="normal")
        self.log_text.insert(tk.END, line + "\n")
        # gioi han so dong hien thi de tranh cham may
        if float(self.log_text.index("end-1c").split(".")[0]) > 2000:
            self.log_text.delete("1.0", "500.0")
        self.log_text.see(tk.END)
        self.log_text.configure(state="disabled")

    # ------------------------------------------------------------------
    # CONNECT / DISCONNECT
    # ------------------------------------------------------------------
    def _toggle_connect(self):
        if self.master_modbus.is_connected:
            self._disconnect()
        else:
            self._connect()

    def _connect(self):
        port = self._selected_port()
        if not port:
            messagebox.showwarning(APP_TITLE, "Vui long chon cong COM.")
            return
        try:
            self.master_modbus.connect(
                port=port,
                baudrate=self.baud_var.get(),
                parity=self.parity_var.get(),
                databits=8,
                stopbits=1,
                timeout=1.0,
            )
        except Exception as e:
            messagebox.showerror(APP_TITLE, "Khong the mo cong %s:\n%s" % (port, e))
            return

        self.connect_btn.configure(text="Disconnect", bg=SEC_BG, fg=SEC_TX, activebackground=SEC_HOV)
        self._draw_indicator(WARN_FG)
        self.status_var.set("Da mo cong %s. Dang doc du lieu ..." % port)

        self.poll_running.set()
        self.poll_thread = threading.Thread(target=self._poll_loop, daemon=True)
        self.poll_thread.start()

    def _disconnect(self):
        self.poll_running.clear()
        if self.poll_thread is not None:
            self.poll_thread.join(timeout=2)
        self.master_modbus.disconnect()
        self.connect_btn.configure(text="Connect", bg=PRIMARY, fg="white", activebackground=PRIMARY_HOV)
        self._draw_indicator(DIS_FG)
        self.status_var.set("Da ngat ket noi.")
        if self.logging_active:
            self._stop_logging()

    # ------------------------------------------------------------------
    # POLLING LOOP (chay tren thread rieng)
    # ------------------------------------------------------------------
    def _poll_loop(self):
        error_count = 0
        while self.poll_running.is_set():
            start_t = time.time()
            slave_id = self.slave_id_var.get()
            values = {}
            ok_any = False
            for reg in self.register_map:
                if not self.poll_running.is_set():
                    break
                try:
                    n = reg_word_count(reg["type"])
                    raw = self.master_modbus.read_holding_registers(slave_id, reg["address"], n, retries=1)
                    val = raw_to_value(raw, reg["type"], reg["scale"])
                    values[reg["name"]] = val
                    ok_any = True
                    error_count = 0
                except Exception as e:
                    values[reg["name"]] = "ERR"
                    error_count += 1
                    self.gui_queue.put(("error", str(e)))
                time.sleep(0.02)  # nghi ngan giua cac request tranh dung do RS485

            if ok_any:
                self.gui_queue.put(("status", "Dang doc - Slave ID %d - %s" %
                                    (slave_id, datetime.now().strftime("%H:%M:%S"))))
            self.gui_queue.put(("data", values))

            if self.logging_active and self.excel_logger is not None:
                try:
                    numeric_values = {k: v for k, v in values.items() if v != "ERR"}
                    self.excel_logger.append_row(datetime.now(), numeric_values)
                    self.excel_logger.flush_if_needed(min_rows=5)
                except (PermissionError, OSError) as e:
                    # File dang bi khoa (thuong do dang mo bang Excel). Du lieu KHONG mat -
                    # van nam trong bo nho cua excel_logger va se duoc thu ghi lai o vong poll ke tiep.
                    # Chi bao loi toi da 1 lan moi N giay de tranh spam trang thai/gui_queue.
                    now_ts = time.time()
                    if now_ts - self._last_excel_error_ts >= self._excel_error_throttle_sec:
                        self._last_excel_error_ts = now_ts
                        self.gui_queue.put(("error", str(e)))
                except Exception as e:
                    self.gui_queue.put(("error", "Ghi Excel loi: %s" % e))

            elapsed = time.time() - start_t
            wait = max(0.05, self.scan_rate_var.get() / 1000.0 - elapsed)
            slept = 0.0
            while slept < wait and self.poll_running.is_set():
                time.sleep(0.05)
                slept += 0.05

    def _update_data_table(self, values: dict):
        if self.master_modbus.is_connected and values:
            ok_any = any(v != "ERR" for v in values.values())
            self._draw_indicator(PASS_FG if ok_any else FAIL_FG)

        for reg in self.register_map:
            val = values.get(reg["name"], "-")
            self.tree.set(reg["name"], "value", val)

        ts = datetime.now()
        temp = values.get("Temperature")
        hum = values.get("Humidity")
        if isinstance(temp, (int, float)) or isinstance(hum, (int, float)):
            self.history_time.append(ts)
            if "Temperature" in self.history_values:
                self.history_values["Temperature"].append(temp if isinstance(temp, (int, float)) else None)
            if "Humidity" in self.history_values:
                self.history_values["Humidity"].append(hum if isinstance(hum, (int, float)) else None)
            if len(self.history_time) > self.history_max_points:
                self.history_time.pop(0)
                for k in self.history_values:
                    self.history_values[k].pop(0)
            self._update_chart()

    def _update_chart(self):
        if not HAS_MPL:
            return
        if not self.history_time:
            return
        x = list(range(len(self.history_time)))
        labels = [t.strftime("%H:%M:%S") for t in self.history_time]
        temp_series = self.history_values.get("Temperature", [])
        hum_series = self.history_values.get("Humidity", [])

        self.line_temp.set_data(x, temp_series)
        self.line_hum.set_data(x, hum_series)

        self.ax_temp.relim()
        self.ax_temp.autoscale_view()
        self.ax_hum.relim()
        self.ax_hum.autoscale_view()

        step = max(1, len(x) // 8)
        self.ax_temp.set_xticks(x[::step])
        self.ax_temp.set_xticklabels(labels[::step], rotation=30, ha="right", fontsize=7)

        title_bits = []
        if temp_series and temp_series[-1] is not None:
            title_bits.append("Temp: %.1f °C" % temp_series[-1])
        if hum_series and hum_series[-1] is not None:
            title_bits.append("Humidity: %.1f %%RH" % hum_series[-1])
        self.ax_temp.set_title("   |   ".join(title_bits), fontsize=10)

        try:
            self.canvas.draw_idle()
        except Exception:
            pass

    # ------------------------------------------------------------------
    # AUTO SCAN
    # ------------------------------------------------------------------
    def _toggle_auto_scan(self):
        if self.auto_scan_thread is not None and self.auto_scan_thread.is_alive():
            self.auto_scan_cancel.set()
            return
        port = self._selected_port()
        if not port:
            messagebox.showwarning(APP_TITLE, "Vui long chon cong COM.")
            return
        if self.master_modbus.is_connected:
            self._disconnect()
        self.auto_scan_cancel.clear()
        self.auto_scan_btn.configure(text="Dung Scan")
        self.auto_scan_thread = threading.Thread(target=self._auto_scan_worker, args=(port,), daemon=True)
        self.auto_scan_thread.start()

    def _auto_scan_worker(self, port):
        """Quet qua cac baudrate pho bien va slave id 1..15 (dia chi phan cung)
        de tim thiet bi dang phan hoi thanh ghi Temperature (address 0)."""
        bauds_to_try = [9600, 19200, 4800, 14400, 38400, 57600, 115200, 56000]
        found = None
        total = len(bauds_to_try) * 15
        step = 0
        temp_master = ModbusRTUMaster(log_callback=self._on_modbus_log)
        for baud in bauds_to_try:
            if self.auto_scan_cancel.is_set():
                break
            try:
                temp_master.connect(port=port, baudrate=baud, parity="N", timeout=0.3)
            except Exception as e:
                self.gui_queue.put(("error", "Auto scan: khong mo duoc cong: %s" % e))
                break
            for sid in range(1, 16):
                if self.auto_scan_cancel.is_set():
                    break
                step += 1
                self.gui_queue.put(("scan_progress",
                                    "Auto scan %.0f%% - baud=%d, slave=%d" % (100.0 * step / total, baud, sid)))
                try:
                    raw = temp_master.read_holding_registers(sid, 0, 1, retries=0)
                    if raw:
                        found = (port, baud, sid)
                        break
                except Exception:
                    continue
            temp_master.disconnect()
            if found:
                break
        self.gui_queue.put(("scan_done", found))

    def _on_auto_scan_done(self, found):
        self.auto_scan_btn.configure(text="Auto Scan")
        if found:
            port, baud, sid = found
            self.baud_var.set(baud)
            self.slave_id_var.set(sid)
            self.status_var.set("Tim thay thiet bi: port=%s, baud=%d, slave id=%d" % (port, baud, sid))
            messagebox.showinfo(APP_TITLE, "Da tim thay thiet bi Modbus!\nBaudrate=%d, Slave ID=%d" % (baud, sid))
        else:
            self.status_var.set("Auto scan: khong tim thay thiet bi phan hoi.")
            if not self.auto_scan_cancel.is_set():
                messagebox.showwarning(APP_TITLE, "Khong tim thay thiet bi Modbus phan hoi tren cong da chon.")

    # ------------------------------------------------------------------
    # WRITE SINGLE REGISTER (tab Settings)
    # ------------------------------------------------------------------
    def _write_register(self, reg, var):
        if not self.master_modbus.is_connected:
            messagebox.showwarning(APP_TITLE, "Vui long ket noi thiet bi truoc khi ghi thanh ghi.")
            return
        raw_text = var.get().strip()
        if not raw_text:
            messagebox.showwarning(APP_TITLE, "Vui long nhap gia tri can ghi.")
            return
        try:
            value = int(float(raw_text) * (reg["scale"] if reg["scale"] else 1))
        except ValueError:
            messagebox.showerror(APP_TITLE, "Gia tri khong hop le.")
            return

        if not messagebox.askyesno(
            APP_TITLE,
            "Ghi gia tri %s vao thanh ghi %d (%s)?\nThiet bi co the can re-power de ap dung."
            % (raw_text, reg["address"], reg["name"])
        ):
            return

        def do_write():
            try:
                self.master_modbus.write_single_register(self.slave_id_var.get(), reg["address"], value)
                self.gui_queue.put(("status", "Ghi thanh cong: %s = %s" % (reg["name"], raw_text)))
            except Exception as e:
                self.gui_queue.put(("error", "Ghi thanh ghi loi: %s" % e))

        threading.Thread(target=do_write, daemon=True).start()

    # ------------------------------------------------------------------
    # TREEVIEW DOUBLE CLICK -> quick edit cho thanh ghi RW
    # ------------------------------------------------------------------
    def _on_tree_double_click(self, event):
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return
        reg = next((r for r in self.register_map if r["name"] == item_id), None)
        if reg is None or reg["rw"] != "RW":
            return
        current = self.tree.set(item_id, "value")
        new_val = simpledialog.askstring(
            APP_TITLE, "Nhap gia tri moi cho '%s' (hien tai: %s):" % (reg["name"], current), parent=self
        )
        if new_val:
            var = tk.StringVar(value=new_val)
            self._write_register(reg, var)

    # ------------------------------------------------------------------
    # EXCEL LOGGING
    # ------------------------------------------------------------------
    def _choose_excel_file(self):
        default_name = "ModbusLog_%s.xlsx" % datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            title="Chon vi tri luu file Excel log",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")]
        )
        if path:
            self.log_filepath_var.set(path)

    def _toggle_logging(self):
        if self.logging_active:
            self._stop_logging()
        else:
            self._start_logging()

    def _start_logging(self):
        filepath = self.log_filepath_var.get().strip()
        if not filepath:
            self._choose_excel_file()
            filepath = self.log_filepath_var.get().strip()
        if not filepath:
            return

        if self.excel_logger is not None and self.excel_logger.filepath == filepath:
            # Cung 1 file voi lan truoc (vi du lan Stop truoc bi loi Permission denied):
            # KHONG tao logger moi de khong mat cac dong dang cho trong bo nho, chi thu save lai.
            try:
                self.excel_logger.save()
            except Exception as e:
                messagebox.showerror(
                    APP_TITLE,
                    "File van dang bi khoa / khong the ghi:\n%s\n\n"
                    "Hay dong file Excel neu dang mo (hoac kiem tra quyen ghi thu muc), roi thu lai. "
                    "Du lieu chua luu van dang duoc giu trong bo nho." % e
                )
                return
        else:
            if self.excel_logger is not None and self.excel_logger.pending_rows > 0:
                if not messagebox.askyesno(
                    APP_TITLE,
                    "Con %d dong du lieu chua luu duoc vao file cu (%s).\n"
                    "Neu chuyen sang file moi, cac dong nay se bi MAT.\nBan co chac muon tiep tuc khong?"
                    % (self.excel_logger.pending_rows, os.path.basename(self.excel_logger.filepath))
                ):
                    return
            try:
                self.excel_logger = ExcelLogger(filepath, self.register_map)
                self.excel_logger.open_existing_or_create()
            except Exception as e:
                messagebox.showerror(APP_TITLE, "Khong the tao/mo file Excel:\n%s" % e)
                return

        self.logging_active = True
        self.start_log_btn.configure(text="Dừng ghi log (Stop Logging)", bg=PASS_FG,
                                      activebackground=PASS_FG)
        self.log_status_var.set("Dang ghi log vao: %s" % os.path.basename(filepath))

    def _stop_logging(self):
        self.logging_active = False
        self.start_log_btn.configure(text="Bắt đầu ghi log (Start Logging)", bg=PRIMARY,
                                      activebackground=PRIMARY_HOV)
        save_error = None
        if self.excel_logger is not None:
            try:
                self.excel_logger.save()
            except Exception as e:
                save_error = e
        self.log_status_var.set("Da dung ghi log. File: %s" %
                                (os.path.basename(self.log_filepath_var.get()) if self.log_filepath_var.get() else ""))
        if save_error is not None:
            messagebox.showwarning(
                APP_TITLE,
                "Da dung ghi log nhung LUU FILE LAN CUOI BI LOI:\n%s\n\n"
                "Hay dong file Excel (neu dang mo) roi bam 'Bat dau ghi log' lai de app "
                "thu luu lai toan bo du lieu con trong bo nho." % save_error
            )

    def _export_snapshot(self):
        """Xuat 1 ban ghi hien tai (snapshot) ra file Excel moi, khong can dang logging."""
        path = filedialog.asksaveasfilename(
            title="Xuat snapshot Excel",
            defaultextension=".xlsx",
            initialfile="ModbusSnapshot_%s.xlsx" % datetime.now().strftime("%Y%m%d_%H%M%S"),
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not path:
            return
        try:
            logger = ExcelLogger(path, self.register_map)
            logger.create_new()
            values = {r["name"]: self.tree.set(r["name"], "value") for r in self.register_map}
            numeric_values = {}
            for k, v in values.items():
                try:
                    numeric_values[k] = float(v)
                except (ValueError, TypeError):
                    numeric_values[k] = v
            logger.append_row(datetime.now(), numeric_values)
            logger.save()
            messagebox.showinfo(APP_TITLE, "Da xuat snapshot ra:\n%s" % path)
        except Exception as e:
            messagebox.showerror(APP_TITLE, "Loi khi xuat Excel:\n%s" % e)

    # ------------------------------------------------------------------
    def _on_close(self):
        try:
            self.poll_running.clear()
            self.auto_scan_cancel.set()
            if self.logging_active:
                self._stop_logging()
            if self.master_modbus.is_connected:
                self.master_modbus.disconnect()
        finally:
            self.destroy()


def main():
    missing = []
    if serial is None:
        missing.append("pyserial")
    if Workbook is None:
        missing.append("openpyxl")
    if missing:
        print("CANH BAO: thieu thu vien: %s" % ", ".join(missing))
        print("Cai dat bang lenh: pip install " + " ".join(missing))
    app = ModbusPollApp()
    app.mainloop()


if __name__ == "__main__":
    main()
